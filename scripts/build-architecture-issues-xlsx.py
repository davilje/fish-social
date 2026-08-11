#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Build or update 服务器架构问题与修复方案-v0.5.xlsx with all 14 issue sheets."""
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill

ROOT = Path(__file__).resolve().parents[1]
REPORTS = ROOT / "docs/planning/reports"
OUT = REPORTS / "服务器架构问题与修复方案-v0.5.xlsx"
OUT_ALT = REPORTS / "服务器架构问题与修复方案-v0.5-完整.xlsx"


def style_sheet(ws) -> None:
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 120
    for c in ws[1]:
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="1F4E78")
        c.alignment = Alignment(horizontal="center", vertical="center")
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=2):
        row[0].font = Font(bold=True)
        row[0].fill = PatternFill("solid", fgColor="D9E1F2")
        row[0].alignment = Alignment(vertical="top", wrap_text=True)
        row[1].alignment = Alignment(vertical="top", wrap_text=True)


def add_sheet(wb, name: str, content: list[tuple[str, str]]) -> None:
    if name in wb.sheetnames:
        del wb[name]
    ws = wb.create_sheet(name)
    ws.append(["字段", "内容"])
    for r in content:
        ws.append(list(r))
    style_sheet(ws)


OVERVIEW_ROWS = [
    ["R0-1", "无身份鉴权（playerId可伪造）", "P0", "账号安全/数据篡改", "已实现", "01-R0-1-鉴权缺失"],
    ["R0-2", "在线核心状态内存化（重启丢失）", "P0", "在线一致性/恢复", "已实现", "02-R0-2-在线态丢失"],
    ["R0-3", "定时器模型分散", "P0", "时序一致性", "已实现", "03-R0-3-定时器分散"],
    ["R0-4", "咬钩路径双轨（stateMachine/session）", "P0", "行为一致性", "已实现", "04-R0-4-咬钩双轨"],
    ["R1-1", "index.ts职责过载", "P1", "可维护性/改动风险", "已实现", "05-R1-1-入口过载"],
    ["R1-2", "会话映射碎片化", "P1", "重连排查/一致性", "已实现", "06-R1-2-映射碎片化"],
    ["R1-3", "全局可变状态跨模块暴露", "P1", "耦合度/回归风险", "已实现", "07-R1-3-全局可变状态"],
    ["R1-4", "缺少优雅停机资源回收", "P1", "退出一致性", "已实现", "08-R1-4-优雅停机"],
    ["R1-5", "缺少请求级链路追踪", "P1", "线上排障效率", "已实现", "09-R1-5-链路追踪"],
    ["R2-1", "高频全量snapshot广播/扫描", "P2", "性能扩展性", "已实现", "10-R2-1-snapshot全量"],
    ["R2-2", "SQLite写压力集中", "P2", "吞吐/延迟", "已实现", "11-R2-2-SQLite写压"],
    ["R2-3", "单实例架构扩展受限", "P2", "水平扩展", "已文档化", "12-R2-3-单实例上限"],
    ["R2-4", "Admin默认密钥+宽松CORS", "P2", "管理面安全", "已实现", "13-R2-4-Admin安全"],
    ["R2-5", "Bot与真人资源竞争复杂", "P2", "业务稳定性", "已实现", "14-R2-5-Bot竞争"],
]

SHEET_CONTENT: dict[str, list[tuple[str, str]]] = {
    "01-R0-1-鉴权缺失": [
        ("问题编号", "R0-1"),
        ("问题标题", "无身份鉴权（客户端可伪造 playerId）"),
        ("风险级别", "P0"),
        ("现状", "Socket/REST 多数写操作直接信任请求中的 playerId，缺少 token 鉴权与服务端身份绑定。"),
        ("触发路径", "register_player / join_pond / 社交与背包相关接口"),
        ("根因分析", "系统以联调便捷为导向，尚未建立统一认证边界；playerId 同时承担展示ID与身份ID，导致可伪造。"),
        ("修复目标", "所有会修改状态的请求必须由服务端认证身份决定，不信任 body/params 中的 playerId。"),
        ("修复方案-步骤1", "引入 JWT：登录/注册后签发短期 access token，payload 至少含 playerId、iat、exp。"),
        ("修复方案-步骤2", "Socket 鉴权：连接时通过 auth.token 校验，校验失败拒绝连接；连接上下文挂载 authPlayerId。"),
        ("修复方案-步骤3", "HTTP 鉴权中间件：对写接口统一校验 Authorization Bearer token。"),
        ("修复方案-步骤4", "接口收敛：写操作使用 token.playerId 作为唯一身份来源；忽略或校验 body 中 playerId 不一致。"),
        ("修复方案-步骤5", "审计日志：新增 auth_failed / identity_mismatch 结构化日志与 metrics。"),
        ("回归测试建议", "1) 正常 token 成功 2) 伪造 playerId 被拒 3) 过期 token 拒绝 4) Socket 重连鉴权通过"),
        ("实施优先级", "第一优先，本周完成"),
        ("预估工作量", "0.5~1.5 天（后端）"),
        ("验收标准", "关键写接口全部经鉴权；伪造 playerId 场景无法修改他人数据；日志可追踪鉴权失败原因。"),
        ("状态", "已实现"),
    ],
    "02-R0-2-在线态丢失": [
        ("问题编号", "R0-2"),
        ("问题标题", "在线核心状态内存化（进程重启后丢失）"),
        ("风险级别", "P0"),
        ("现状", "钓位 spotId、fishingPhase、断线宽限 disconnectTimers、hook 上下文 hookContextByUser、pending catch 锁、鱼塘聊天 pondChats 等关键在线态全部存于内存 Map；服务重启或崩溃后无法恢复。"),
        ("触发路径", "进程重启 / SIGTERM 强退 / 未捕获异常导致进程退出"),
        ("根因分析", "架构采用「单进程权威服 + 内存实时态 + SQLite 补充持久化」；仅生态鱼群、背包、玩家资料、metrics 落库，玩家「在塘内正在做什么」未设计 checkpoint。"),
        ("影响范围", "玩家离塘、钓位被释放、待领取鱼丢失、断线 60s 宽限失效、挂机恢复失败、排查时无法还原重启前现场"),
        ("修复目标", "重启后可恢复高价值在线态：塘内身份、钓位、phase、pending catch；至少做到 best-effort 恢复，避免静默丢状态。"),
        ("修复方案-阶段A（最小可行）", "新增 player_pond_session 表：字段 pondId, playerId, userId, spotId, fishingPhase, phaseEndsAt, hookEndsAt, disconnectedAt, updatedAt。在 phase 变更、占座/离席、disconnect、lockPendingCatch 时 upsert；accept/expired/leave 时删除。"),
        ("修复方案-阶段B（启动恢复）", "玩家 join_pond / reconnect 时先查 checkpoint：若存在未过期会话则复用 userId 与 phase 链恢复，而非一律新建 userId。"),
        ("修复方案-阶段C（pending 持久化）", "pending catch 在 lock 时写 SQLite；超时或 accept 时删除；重启后扫描过期项并补打 pending_catch_expired。"),
        ("修复方案-阶段D（聊天可选）", "pondChats 若需保留可限长落库（每塘最近 N 条）；非 P0，可二期。"),
        ("不改范围", "不做全量多实例热迁移；不做 Redis 在线态（当前阶段）。"),
        ("涉及文件", "gameState.ts · fishingStateMachine.ts · inventory.ts · db.ts · migrations · index.ts"),
        ("回归测试建议", "占座+waiting 后 kill 进程再启 reconnect 恢复；hooked 宽限内重启可续；pending 重启后仍可 expired"),
        ("实施优先级", "第一优先（与 R0-3 可并行）"),
        ("预估工作量", "2~4 天"),
        ("验收标准", "重启后 reconnect 可恢复 spotId 与合法 phase；pending 不静默丢失；verify 覆盖 kill-restart。"),
        ("状态", "已实现"),
    ],
    "03-R0-3-定时器分散": [
        ("问题编号", "R0-3"),
        ("问题标题", "定时器模型分散（全局 interval + 多处 per-user setTimeout）"),
        ("风险级别", "P0"),
        ("现状", "index.ts 内 200ms/1s/咬钩/生态全局循环；fishingStateMachine.disconnectTimers（60s）；fishingSession.hookStateByUser；inventory.pending 超时 timer。取消逻辑分散在 disconnect/reconnect/leave/stop 等多处。"),
        ("触发路径", "断线重连、离塘、stop_fishing、hook 超时、pending 超时"),
        ("根因分析", "各模块自行 setTimeout/setInterval，缺少统一注册与生命周期管理；历史上已出现重连未取消 disconnect timer 类 bug。"),
        ("影响范围", "重连误踢、重复 phase 推进、hook/pending 双触发、状态漂移"),
        ("修复目标", "所有 per-user / per-socket 定时器经统一注册表管理；leave/disconnect/reconnect 走统一 cancelAll 入口。"),
        ("修复方案-步骤1", "新建 timerRegistry.ts：register、cancelByUser、cancelBySocket、cancelByKind。"),
        ("修复方案-步骤2", "迁移 disconnectTimers、hookStateByUser timer、pending timeout 到 registry。"),
        ("修复方案-步骤3", "handleDisconnect、restoreDisconnectedUser、leavePond、clearSession 统一调用 cancel。"),
        ("修复方案-步骤4", "shutdown 时 clear 全局 interval；registry.listActive() 供 Admin debug。"),
        ("修复方案-步骤5", "埋点 timer_registered / timer_cancelled（可选 P2）。"),
        ("涉及文件", "timerRegistry.ts（新）· fishingStateMachine.ts · fishingSession.ts · inventory.ts · index.ts"),
        ("回归测试建议", "verify-disconnect-reconnect 全通过；重连 60s 内不被踢；leave 后无残留 timer"),
        ("实施优先级", "第一优先，与 R0-2 可并行"),
        ("预估工作量", "1~2 天"),
        ("验收标准", "用户态逻辑无分散裸 setTimeout；重连/离塘 timer 泄漏为 0。"),
        ("状态", "已实现"),
    ],
    "04-R0-4-咬钩双轨": [
        ("问题编号", "R0-4"),
        ("问题标题", "咬钩与 Hook 逻辑双轨（fishingStateMachine vs fishingSession）"),
        ("风险级别", "P0"),
        ("现状", "真人咬钩走 fishingStateMachine + hookContextByUser；bot 走 fishingSession.processBotBiteTick；hook 倒计时在 fishingSession.hookStateByUser；index.ts 多处手动 cancelHookResolution。"),
        ("触发路径", "waiting 咬钩、hooked 超时、bot 咬钩、stop/disconnect/leave"),
        ("根因分析", "C6 状态机上线后 fishingSession 未完全收敛，形成过渡架构。"),
        ("影响范围", "修一侧漏另一侧；hook 与 phase 不同步；bot/真人行为不一致"),
        ("修复目标", "单一咬钩内核；hook 超时纳入状态机；fishingSession 退化为纯概率/选鱼计算。"),
        ("修复方案-步骤1", "抽取 rollBiteHook 等纯函数到 biteEngine.ts 或保留在 fishingSession。"),
        ("修复方案-步骤2", "bot 咬钩调用与真人相同的 processWaitingBiteTick 或共享 inner 函数。"),
        ("修复方案-步骤3", "废弃 hookStateByUser：hooked 超时由 tickFishingPhases 根据 phaseEndsAt 推进。"),
        ("修复方案-步骤4", "删除 index.ts 分散的 cancelHookResolution，收敛到 leave/disconnect。"),
        ("涉及文件", "fishingStateMachine.ts · fishingSession.ts · bots.ts · index.ts · fishingDebug.ts"),
        ("回归测试建议", "真人/bot 咬钩脱钩回归；hook 超时进 resolving"),
        ("实施优先级", "第一优先，建议在 R0-3 之后"),
        ("预估工作量", "2~3 天"),
        ("验收标准", "无 hookStateByUser 并行 timer；bot 与真人共用咬钩主路径。"),
        ("状态", "已实现"),
    ],
    "05-R1-1-入口过载": [
        ("问题编号", "R1-1"),
        ("问题标题", "index.ts 职责过载（上帝对象）"),
        ("风险级别", "P1"),
        ("现状", "index.ts ~600 行：HTTP、Socket 全事件、4 类全局循环、重连编排、perf 日志、shutdown。"),
        ("根因分析", "早期快速迭代将所有编排堆在入口文件。"),
        ("修复目标", "index.ts 仅 boot + wiring；业务拆到独立模块。"),
        ("修复方案-步骤1", "socketPondHandlers.ts：join/leave/fishing/chat/accept_catch。"),
        ("修复方案-步骤2", "socketLifecycle.ts：connection/disconnect/register_player。"),
        ("修复方案-步骤3", "serverLoops.ts：phase tick、bite loop、ecology、perf。"),
        ("修复方案-步骤4", "createApp.ts：Express 组装。"),
        ("修复方案-步骤5", "index.ts 行数目标 <150。"),
        ("实施优先级", "第二期"),
        ("预估工作量", "2~3 天"),
        ("验收标准", "index.ts 无大段业务实现；verify 全绿。"),
        ("状态", "已实现"),
    ],
    "06-R1-2-映射碎片化": [
        ("问题编号", "R1-2"),
        ("问题标题", "会话身份映射碎片化（三套 Map）"),
        ("风险级别", "P1"),
        ("现状", "sessions、socketByUserId、playerSockets 分散维护，写入时机不一致。"),
        ("修复目标", "单一 SessionRegistry 管理 bind/unbind/rebind。"),
        ("修复方案-步骤1", "新建 sessionRegistry.ts。"),
        ("修复方案-步骤2", "register_player / join reconnect / disconnect 统一走 registry。"),
        ("修复方案-步骤3", "打 session_rebound 日志；identity_mismatch 检测。"),
        ("修复方案-步骤4", "socialRoutes / bots 改为 registry.resolveSocketByPlayer。"),
        ("实施优先级", "第二期，与 R0-1 协同"),
        ("预估工作量", "1~2 天"),
        ("验收标准", "仅 registry 可写映射；session_rebound 可观测。"),
        ("状态", "已实现"),
    ],
    "07-R1-3-全局可变状态": [
        ("问题编号", "R1-3"),
        ("问题标题", "全局可变状态跨模块 export（隐式耦合）"),
        ("风险级别", "P1"),
        ("现状", "playerSockets、pondUsers、sessions 等 Map 被多模块直接读写。"),
        ("修复目标", "状态访问经 service/API；Map 不 export 或只读视图。"),
        ("修复方案-步骤1", "gameState 只 export 函数，内部 Map 私有化。"),
        ("修复方案-步骤2", "playerSockets 迁入 SessionRegistry。"),
        ("修复方案-步骤3", "禁止新业务直接 import 可变 Map。"),
        ("实施优先级", "第二期，与 R1-2 同步"),
        ("预估工作量", "1~2 天"),
        ("验收标准", "无业务文件直接 import 可变 Map。"),
        ("状态", "已实现"),
    ],
    "08-R1-4-优雅停机": [
        ("问题编号", "R1-4"),
        ("问题标题", "缺少优雅停机与资源回收"),
        ("风险级别", "P1"),
        ("现状", "shutdown 仅 io.close + 2s 强退；未清 interval、未 drain timer、未 checkpoint、未 db.close。"),
        ("修复目标", "拒新连接 → 停循环 → 取消 timer → checkpoint → 关 DB → 退出。"),
        ("修复方案-步骤1", "serverLoops 返回 stopHandles，shutdown 时 clear。"),
        ("修复方案-步骤2", "timerRegistry.cancelAll()。"),
        ("修复方案-步骤3", "若 R0-2 已做：shutdown 前写 checkpoint。"),
        ("修复方案-步骤4", "db.close()；超时延长至 5~10s 可配置。"),
        ("实施优先级", "第二期"),
        ("预估工作量", "0.5~1 天"),
        ("验收标准", "停机分阶段日志；无未 clear 的全局 interval。"),
        ("状态", "已实现"),
    ],
    "09-R1-5-链路追踪": [
        ("问题编号", "R1-5"),
        ("问题标题", "缺少请求级链路追踪（correlationId）"),
        ("风险级别", "P1"),
        ("现状", "有 event 级埋点，但 Socket/HTTP 无法串联同一操作链。"),
        ("修复目标", "每个 socket 连接一个 correlationId；关键 logs/metrics 携带。"),
        ("修复方案-步骤1", "connection 时生成 correlationId 挂 socket.data。"),
        ("修复方案-步骤2", "logStructuredEvent / recordFishingMetric 自动附加。"),
        ("修复方案-步骤3", "HTTP 支持 X-Request-Id。"),
        ("修复方案-步骤4", "Admin timeline 支持按 correlationId 过滤（可选）。"),
        ("实施优先级", "第二期"),
        ("预估工作量", "1 天"),
        ("验收标准", "P0 路径日志含 correlationId。"),
        ("状态", "已实现"),
    ],
    "10-R2-1-snapshot全量": [
        ("问题编号", "R2-1"),
        ("问题标题", "高频全量 snapshot 构建与广播"),
        ("风险级别", "P2"),
        ("现状", "1s 循环每塘 buildSnapshot 全量广播；咬钩循环每轮全塘扫描。"),
        ("修复目标", "广播只推变更用户；咬钩维护 waiting 用户索引。"),
        ("修复方案-步骤1", "markUserDirty；1s 循环只 emit dirty 用户。"),
        ("修复方案-步骤2", "waitingUsersByPond 索引，phase 进出 waiting 时更新。"),
        ("修复方案-步骤3", "buildSnapshot 仅在 join/客户端请求时全量。"),
        ("修复方案-步骤4", "结合 snapshot_build_duration_ms 验证效果。"),
        ("实施优先级", "第三期"),
        ("预估工作量", "2~3 天"),
        ("验收标准", "1s 循环不再每塘全量 buildSnapshot。"),
        ("状态", "已实现"),
    ],
    "11-R2-2-SQLite写压": [
        ("问题编号", "R2-2"),
        ("问题标题", "SQLite 写压力集中"),
        ("风险级别", "P2"),
        ("现状", "fishing_metrics 每条同步 insert；单 db 连接 WAL。"),
        ("修复目标", "metrics 异步批量 flush；生态写合并事务。"),
        ("修复方案-步骤1", "fishingMetrics 内存队列 + 批量 INSERT。"),
        ("修复方案-步骤2", "recordFishingMetric 改为 enqueue，退出时 drain。"),
        ("修复方案-步骤3", "生态 tick 写操作单事务。"),
        ("实施优先级", "第三期"),
        ("预估工作量", "1~2 天"),
        ("验收标准", "metrics 不再每条同步 insert。"),
        ("状态", "已实现"),
    ],
    "12-R2-3-单实例上限": [
        ("问题编号", "R2-3"),
        ("问题标题", "单实例架构扩展受限"),
        ("风险级别", "P2"),
        ("现状", "内存权威态 + Socket.io 单节点；无多实例设计。"),
        ("修复目标", "明确容量上限；远期多实例方案文档化，短期不实施。"),
        ("修复方案-前置条件", "R0-2 checkpoint · SessionRegistry · Redis adapter · 分布式锁"),
        ("修复方案-远期", "Redis 会话、socket.io redis-adapter、nginx ip_hash、分塘分片。"),
        ("修复方案-短期", "文档标注并发上限；监控活跃连接；超阈值告警。"),
        ("实施优先级", "第四期 / 按需"),
        ("预估工作量", "1~2 周（完整多实例）"),
        ("验收标准", "短期有容量指标与告警；远期双实例联调通过。"),
        ("状态", "已文档化"),
    ],
    "13-R2-4-Admin安全": [
        ("问题编号", "R2-4"),
        ("问题标题", "Admin 默认密钥与宽松 CORS"),
        ("风险级别", "P2"),
        ("现状", "ADMIN_SECRET 默认 fish-social-debug；CORS origin *。"),
        ("修复目标", "生产强制强密钥；CORS 白名单；破坏性操作审计。"),
        ("修复方案-步骤1", "production 未设 ADMIN_SECRET 则拒绝启动。"),
        ("修复方案-步骤2", "CORS 改为 ALLOWED_ORIGINS 环境变量。"),
        ("修复方案-步骤3", "clear/reset 记录审计日志。"),
        ("修复方案-步骤4", "文档：Admin 仅内网/VPN 暴露。"),
        ("实施优先级", "第三期（公网前必须）"),
        ("预估工作量", "0.5 天"),
        ("验收标准", "生产无默认密钥；CORS 可配置。"),
        ("状态", "已实现"),
    ],
    "14-R2-5-Bot竞争": [
        ("问题编号", "R2-5"),
        ("问题标题", "Bot 与真人资源竞争逻辑复杂"),
        ("风险级别", "P2"),
        ("现状", "满员 evictBotsForHuman；bot 与真人共享生态与不同咬钩路径。"),
        ("修复目标", "咬钩路径统一（R0-4）；bot 配额可配置；metrics 区分 isBot。"),
        ("修复方案-步骤1", "收敛 bot 咬钩到统一引擎。"),
        ("修复方案-步骤2", "配置 MAX_BOTS_PER_POND、BOT_EVICT_POLICY。"),
        ("修复方案-步骤3", "生态补充可选降低 bot 权重。"),
        ("修复方案-步骤4", "可选埋点 bot_spawn / bot_despawn。"),
        ("实施优先级", "第三期，依赖 R0-4"),
        ("预估工作量", "1~2 天"),
        ("验收标准", "bot/真人咬钩同路径；踢 bot 可配置有 metrics。"),
        ("状态", "已实现"),
    ],
}


def build_workbook() -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = "总览"
    headers = ["编号", "问题", "风险级别", "影响范围", "当前状态", "对应sheet"]
    ws.append(headers)
    for row in OVERVIEW_ROWS:
        ws.append(row)
    for col, w in {"A": 10, "B": 38, "C": 10, "D": 22, "E": 16, "F": 24}.items():
        ws.column_dimensions[col].width = w
    for c in ws[1]:
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="1F4E78")
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=6):
        for c in row:
            c.alignment = Alignment(vertical="top", wrap_text=True)

    sheet_order = [r[5] for r in OVERVIEW_ROWS]
    for name in sheet_order:
        add_sheet(wb, name, SHEET_CONTENT[name])
    return wb


def main() -> None:
    REPORTS.mkdir(parents=True, exist_ok=True)
    wb = build_workbook()
    targets = [OUT, OUT_ALT]
    saved = None
    for target in targets:
        try:
            wb.save(target)
            saved = target
            break
        except PermissionError:
            continue
    if not saved:
        raise SystemExit("无法写入 xlsx（文件可能被 Excel 占用），请关闭后重试")
    print(f"Wrote {saved} ({len(wb.sheetnames)} sheets)")


if __name__ == "__main__":
    main()
