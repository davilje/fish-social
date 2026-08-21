#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Export 钓鱼玩法固定数值表.xlsx → shared/generated/game-data + Unity Resources/GameData."""
from __future__ import annotations

import json
import shutil
import subprocess
import sys
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[2]
XLSX = ROOT / "钓鱼玩法固定数值表.xlsx"
OUT_SHARED = ROOT / "shared" / "generated" / "game-data"
OUT_UNITY = ROOT / "fish-social-unity" / "Assets" / "Resources" / "GameData"
BUILD_SCRIPT = Path(__file__).resolve().parent / "build_balance_xlsx.py"
SKIP_EXPORT_SHEETS = {"字段说明"}


def cell_value(v):
    if v is None:
        return None
    if isinstance(v, bool):
        return v
    if isinstance(v, float) and v == int(v):
        return int(v)
    return v


def is_zh_header_row(cells) -> bool:
    nonempty = [c for c in cells if c is not None and str(c).strip() != ""]
    if not nonempty:
        return False
    if any(isinstance(c, (int, float)) and not isinstance(c, bool) for c in nonempty):
        return False
    return any("\u4e00" <= ch <= "\u9fff" for c in nonempty for ch in str(c))


def sheet_to_rows(ws) -> list[dict]:
    rows_iter = ws.iter_rows(values_only=True)
    headers = None
    for raw in rows_iter:
        if raw is None or all(c is None or (isinstance(c, str) and not str(c).strip()) for c in raw):
            continue
        headers = [str(h).strip() if h is not None else f"col{i}" for i, h in enumerate(raw)]
        break
    if not headers:
        return []
    out: list[dict] = []
    skipped_zh = False
    for raw in rows_iter:
        if raw is None or all(c is None or (isinstance(c, str) and not str(c).strip()) for c in raw):
            continue
        if not skipped_zh and is_zh_header_row(raw):
            skipped_zh = True
            continue
        row: dict = {}
        for i, key in enumerate(headers):
            val = cell_value(raw[i]) if i < len(raw) else None
            if val is None or val == "":
                continue
            row[key] = val
        if row:
            out.append(row)
    return out


def meta_to_object(rows: list[dict]) -> dict:
    """_meta is key/value pairs → object."""
    obj: dict = {}
    for row in rows:
        if "key" in row and "value" in row:
            obj[str(row["key"])] = row["value"]
        else:
            # fallback: first two values
            items = list(row.items())
            if len(items) >= 2:
                obj[str(items[0][1])] = items[1][1]
    return obj


def ensure_xlsx() -> None:
    if XLSX.is_file():
        return
    print(f"Missing {XLSX.name}; running build…")
    subprocess.check_call([sys.executable, str(BUILD_SCRIPT)], cwd=str(ROOT))
    if not XLSX.is_file():
        raise SystemExit(f"Build did not produce {XLSX}")


def write_json(path: Path, data) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(
        json.dumps(data, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )


def export() -> None:
    ensure_xlsx()
    wb = load_workbook(XLSX, data_only=True, read_only=True)

    OUT_SHARED.mkdir(parents=True, exist_ok=True)
    OUT_UNITY.mkdir(parents=True, exist_ok=True)

    sheet_names = list(wb.sheetnames)
    version = "0.0.0"
    exported: dict[str, Path] = {}

    for name in sheet_names:
        if name in SKIP_EXPORT_SHEETS:
            continue
        ws = wb[name]
        rows = sheet_to_rows(ws)
        if name == "_meta":
            payload = meta_to_object(rows)
            version = str(payload.get("version", version))
        else:
            payload = rows
        rel = f"{name}.json"
        shared_path = OUT_SHARED / rel
        write_json(shared_path, payload)
        exported[name] = shared_path

    index = {
        "version": version,
        "source": XLSX.name,
        "exportedAt": datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ"),
        "sheets": [name for name in sheet_names if name not in SKIP_EXPORT_SHEETS],
    }
    write_json(OUT_SHARED / "_index.json", index)

    # Copy all JSON (including _index) to Unity Resources
    for src in OUT_SHARED.glob("*.json"):
        dst = OUT_UNITY / src.name
        shutil.copy2(src, dst)

    wb.close()
    print(f"Exported {len(sheet_names)} sheets (+ _index) → {OUT_SHARED}")
    print(f"Copied to {OUT_UNITY}")
    print(f"version={version}")


def main() -> None:
    export()


if __name__ == "__main__":
    main()
