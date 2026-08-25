#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Build root 钓鱼玩法固定数值表.xlsx (FEAT-PROG-01 balance pipeline)."""
from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from fish_cn_data import (
    FISH_QUALITY_STATS,
    FISHING_FORMULA_CONSTANTS,
    FISH_SPECIES_CN,
    build_category_quality_weight_rows,
    build_pond_fish_pool_rows,
    ecology_fields_for_pond,
    pond_extra_fields,
    species_rows_for_xlsx,
)
from spot_clue_data import (
    SPOT_CLUE_TEXTS,
    SPOT_TAG_DEFS,
    build_pond_spot_tag_rows,
)

ROOT = Path(__file__).resolve().parents[2]
OUT = ROOT / "钓鱼玩法固定数值表.xlsx"

HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(bold=True, color="FFFFFF")

# ---------------------------------------------------------------------------
# Authoritative draft data (FEAT-PROG-01 + FEAT-FISH-CN-01 + FEAT-POOL-01)
# ---------------------------------------------------------------------------

META_ROWS = [
    ("key", "value"),
    ("version", "1.1.2"),
    ("SIZE_EXP", 1.15),
    ("maxFeeChargesPerDayDefault", 4),
    ("maxStackCount", 50),
    ("biteMulGlobalCap", 1.5),
    ("albumPinCap", 12),
    ("schemaNote", "FEAT-PROG-01+RETURN-02+FISH-CN-01+POOL-01;pool authority;ecology cols on ponds;quality=category∩speciesBand"),
]

PONDS_HEADER = [
    "pondId",
    "name",
    "pondCategory",
    "mapZoneId",
    "feePer2h",
    "feePer2hSellOnly",
    "feePer2hAutoReturn",
    "allowsAutoReturn",
    "maxFeeChargesPerDay",
    "unlock",
    "isOpen",
    "showOnWorldMap",
    "minPlayerLevel",
    "mapX",
    "mapY",
    "bioRegion",
    "waterType",
    "realWorldRef",
    "maxPopulation",
    "minPopulation",
    "initialPopulation",
]

# Loaded from fish_cn_data (pondId…minLv only for dual-fee loop)
from fish_cn_data import PONDS_CN as _PONDS_CN  # noqa: E402

PONDS = [p[:10] for p in _PONDS_CN]

# FEAT-RETURN-01：单行全局规则（回鱼准入与奖励）
RETURN_RULES = [
    # minQuality, minSizeRatio, maxSizeRatio, goldMulVsSell, playerXp, pondXp,
    # sizeGainMinM, sizeGainMaxM, sizeGainMode, autoMinQuality, autoMinSizeRatio
    ("purple", 0.75, 1.0, 1.50, 8, 4, 0.02, 0.05, "uniform_random", "purple", 0.75),
]

PLAYER_LEVELS = [
    # level, xpToNext, pondXpPerHour, maxPondLevel
    (1, 100, 12, 1),
    (2, 150, 14, 2),
    (3, 220, 16, 3),
    (4, 300, 18, 4),
    (5, 400, 22, 5),
    (6, 520, 26, 5),
    (7, 650, 30, 6),
    (8, 800, 34, 6),
    (9, 980, 38, 7),
    (10, 1200, 42, 7),
    (11, 1450, 48, 8),
    (12, 1750, 54, 8),
    (13, 2100, 60, 9),
    (14, 2500, 66, 9),
    (15, 3000, 72, 10),
    (16, 3600, 80, 10),
    (17, 4300, 88, 10),
    (18, 5100, 96, 10),
    (19, 6000, 104, 10),
    (20, 0, 112, 10),
]

POND_LEVEL_XP = [80, 120, 180, 260, 360, 480, 620, 800, 1000, 0]

FISH_SPECIES = [
    # speciesId, name, diet, catchGroup, typicalMinM, typicalMaxM  (compat for fish_xp loop)
    (sid, name, diet, group, mn, mx)
    for sid, name, diet, group, mn, mx, _tier, _nw in FISH_SPECIES_CN
]

FISH_SPECIES_FULL = species_rows_for_xlsx()
POND_FISH_POOL = build_pond_fish_pool_rows()
POND_CATEGORY_QUALITY_WEIGHTS = build_category_quality_weight_rows()

XP_QUALITY_BASE = {
    "gray": 12,
    "green": 20,
    "blue": 35,
    "purple": 55,
    "red": 90,
    "orange": 140,
    "gold": 220,
}
CATCH_GROUP_COEFF = {
    "still_bait": 0.9,
    "stream_light": 0.85,
    "lure_predator": 1.1,
    "cast_heavy": 1.25,
    "giant_game": 1.6,
}
QUALITIES = list(XP_QUALITY_BASE.keys())

POND_MODIFIERS = [
    # category, biteRateMul, escapeRateMul, infoRevealMul, qualityWeightSkew, sizeCapMul, pondXpMul,
    # fineChancePerHour, fineGold, policeWarningMs
    ("advanced", 1, 1, 1, 1, 1, 1, 0, 0, 0),
    ("novice", 1, 1, 1, 1, 1, 1, 0, 0, 0),
    ("veteran", 1, 1, 1, 1, 1, 1, 0, 0, 0),
    ("wilderness", 0.75, 1.10, 0.70, 0.85, 0.95, 1.10, 0, 0, 0),
    ("reservoir", 0.65, 1.0, 0.55, 0.85, 1.0, 1.15, 0, 0, 0),
    ("forbidden", 0.90, 1, 1, 1, 1, 1, 0.15, 800, 10000),
    ("giant", 1, 1, 1, 1, 1, 1, 0, 0, 0),
]

# rods: fitStillBait / fitStreamLight / fitLurePredator / fitCastHeavy / fitGiantGame
# + quality fits (weak ramp by tier)
RODS = [
    # rodId, name, subType, priceGold, biteBonus, escapeReduction, breakSizeM, breakMaxLandings,
    # fitGray..fitGold, fitStill, fitStream, fitLure, fitCast, fitGiant
    (
        "rod-bamboo", "竹制手竿", "手竿入门", 0, 0, 0, 0.45, 3,
        1.1, 1.05, 1.0, 0.7, 0.5, 0.4, 0.3,
        1.0, 0.7, 0.7, 0.7, 0.7,
    ),
    (
        "rod-tai", "台钓碳素竿", "台钓", 2500, 0.03, 0.04, 0.8, 5,
        1.05, 1.05, 1.05, 0.9, 0.7, 0.55, 0.4,
        1.15, 0.75, 0.75, 0.75, 0.7,
    ),
    (
        "rod-stream", "溪流竿", "溪流", 3200, 0.04, 0.03, 0.55, 4,
        1.05, 1.05, 1.0, 0.85, 0.65, 0.5, 0.35,
        0.75, 1.2, 0.8, 0.75, 0.7,
    ),
    (
        "rod-iso", "矶钓竿", "矶钓", 5500, 0.03, 0.06, 1.2, 5,
        0.95, 1.0, 1.05, 1.0, 0.85, 0.7, 0.55,
        0.75, 0.8, 0.85, 1.1, 0.75,
    ),
    (
        "rod-surf", "海竿", "抛竿", 4800, 0.02, 0.07, 1.5, 6,
        0.9, 1.0, 1.05, 1.05, 0.9, 0.75, 0.6,
        0.7, 0.75, 0.8, 1.15, 0.8,
    ),
    (
        "rod-lure", "路亚竿", "路亚", 9000, 0.05, 0.05, 1.0, 5,
        0.85, 0.95, 1.05, 1.1, 0.95, 0.8, 0.65,
        0.7, 0.8, 1.2, 0.9, 0.75,
    ),
    (
        "rod-heavy", "重型路亚竿", "重路亚", 22000, 0.04, 0.10, 2.5, 8,
        0.75, 0.85, 0.95, 1.1, 1.15, 1.05, 0.9,
        0.7, 0.75, 1.1, 1.1, 0.9,
    ),
    (
        "rod-giant", "巨物竿", "巨物", 45000, 0.03, 0.12, 4.0, 10,
        0.7, 0.8, 0.9, 1.05, 1.15, 1.2, 1.25,
        0.7, 0.7, 0.85, 0.95, 1.25,
    ),
]

BAITS = [
    # baitId, name, diet, unlockPlayerLevel, costGoldPerUse,
    # biteBonusHerbivore, biteBonusOmnivore, biteBonusCarnivore, isDefaultInfinite
    ("bait-basic", "基础杂饵", "any", 0, 0, 0, 0, 0, True),
    ("bait-veg", "谷物饵", "herbivore", 2, 15, 0.06, 0.01, 0.01, False),
    ("bait-mix", "腥香饵", "omnivore", 3, 20, 0.01, 0.06, 0.01, False),
    ("bait-meat", "荤腥饵", "carnivore", 4, 25, 0.01, 0.01, 0.07, False),
]

# FEAT-GROUND-01
GROUNDBAITS = [
    # groundbaitId, name, unlockPlayerLevel, costGoldPerUse, castDurationMs,
    # durationMin, maxBites, perStackBiteBonus, maxBonus, stackK,
    # sizeBonusPerStack, maxSizeBonus
    ("gb-basic", "基础窝料", 3, 30, 3000, 20, 8, 0.006, 0.08, 0.08, 0.003, 0.08),
    ("gb-mix", "混合窝料", 5, 50, 5000, 25, 10, 0.010, 0.12, 0.07, 0.004, 0.10),
    ("gb-premium", "精品窝料", 7, 80, 8000, 30, 12, 0.014, 0.18, 0.06, 0.005, 0.12),
]

# FEAT-ALBUM-01
ACHIEVEMENTS = [
    # achievementId, name, desc, iconKey, category, conditionType, conditionValue, sortOrder, isHidden
    ("ach-first-catch", "初出茅庐", "首次将鱼收入背包", "ach_first_catch", "catch", "catch_count", 1, 10, False),
    ("ach-codex-5", "图鉴新芽", "图鉴解锁达到 5 种", "ach_codex_5", "progress", "codex_count", 5, 20, False),
    ("ach-codex-20", "博物渔者", "图鉴解锁达到 20 种", "ach_codex_20", "progress", "codex_count", 20, 30, False),
    ("ach-return-1", "放生有情", "成功回鱼至少 1 次", "ach_return_1", "catch", "return_count", 1, 40, False),
    ("ach-big-08", "大鱼临门", "单次捕获尺寸达到 0.8m", "ach_big_08", "catch", "max_size", 0.8, 50, False),
    ("ach-album-3", "相册收藏家", "相册精选钉选达到 3 张", "ach_album_3", "album", "album_pins", 3, 60, False),
]

VESSELS = [
    ("vessel-raft", "筏钓日票", 12, 15000, 8, False),
    ("vessel-boat", "路亚小艇", 14, 35000, 12, False),
    ("vessel-trawler", "捕捞船", 18, 80000, 20, False),
]

# FEAT-SPOT-02: spot_clue_data.py (tag-matched clues + activitySignal)

# sheet, field, 中文名, 说明 — exported JSON still uses English keys in row 1.
FIELD_DOCS: list[tuple[str, str, str, str]] = [
    ("_meta", "key", "键", "全局常量名，如 SIZE_EXP、version。"),
    ("_meta", "value", "值", "对应常量的数值或备注。"),
    ("_meta", "maxStackCount", "打窝层数上限", "FEAT-GROUND-01 全局硬 cap，默认 50。"),
    ("_meta", "biteMulGlobalCap", "咬钩总倍率软帽", "相对无窝基线的全局 soft cap，建议 1.5。"),
    ("ponds", "pondId", "鱼塘ID", "程序主键，进塘/扣费/地图都用这个。"),
    ("ponds", "name", "鱼塘名", "中文显示名。"),
    ("ponds", "pondCategory", "鱼塘分级", "novice/advanced/veteran/wilderness/reservoir/forbidden/giant。"),
    ("ponds", "mapZoneId", "地图分区", "世界地图分区 ID。"),
    ("ponds", "feePer2h", "每2小时入场费", "金币；0 表示不收费；兼容字段=出售档。"),
    ("ponds", "feePer2hSellOnly", "出售档每2h费", "FEAT-RETURN-02 不可回鱼档扣费。"),
    ("ponds", "feePer2hAutoReturn", "回鱼档每2h费", "FEAT-RETURN-02 自动回鱼档扣费。"),
    ("ponds", "allowsAutoReturn", "允许双价选择", "TRUE 时进塘需选 sell_only/auto_return。"),
    ("ponds", "bioRegion", "生物区", "策划标注；出鱼以 pond_fish_pool 为准。"),
    ("ponds", "waterType", "水体类型", "lake/reservoir/river/coastal/estuary/pond。"),
    ("ponds", "realWorldRef", "现实原型", "策划备注用真实钓场。"),
    ("ponds", "maxFeeChargesPerDay", "每日扣费次数上限", "按 2 小时切片累计。"),
    ("ponds", "unlock", "解锁条件", "onboarding / level:N / guide_only。"),
    ("ponds", "isOpen", "是否开放", "FALSE 的塘不能进（如巨物塘壳）。"),
    ("ponds", "showOnWorldMap", "是否上地图", "FALSE 则世界地图不画（新手练习塘）。"),
    ("ponds", "minPlayerLevel", "最低钓鱼等级", "低于此等级不能进塘。"),
    ("ponds", "mapX", "地图X", "世界地图坐标 0~1。"),
    ("ponds", "mapY", "地图Y", "世界地图坐标 0~1。"),
    ("ponds", "maxPopulation", "人口上限", "塘内鱼实体上限（原 pond_ecology 已并入）。"),
    ("ponds", "minPopulation", "人口下限", "低于此触发补充。"),
    ("ponds", "initialPopulation", "初始人口", "开服 seed 数量。"),
    ("player_levels", "level", "钓鱼等级", "1~20。"),
    ("player_levels", "xpToNext", "升到下级所需经验", "20 级为 0。"),
    ("player_levels", "pondXpPerHour", "每小时鱼塘熟练度", "挂机时长折算塘经验的参考。"),
    ("player_levels", "maxPondLevel", "可达到的最高塘等级", "玩家等级对塘升级的软上限。"),
    ("return_rules", "minQuality", "最低品质", "低于此品质不可回鱼；gray=灰及以上。"),
    ("return_rules", "minSizeRatio", "最小体长比", "相对品质最大体长的下限，如 0.2。"),
    ("return_rules", "maxSizeRatio", "最大体长比", "相对品质最大体长的上限；1.0 表示满尺寸不可回。"),
    ("return_rules", "goldMulVsSell", "回鱼金倍率", "回鱼金 = floor(卖价 × 本倍率)，建议 1.50。"),
    ("return_rules", "playerXp", "玩家经验", "每次回鱼发给玩家的熟练度。"),
    ("return_rules", "pondXp", "鱼塘经验", "每次回鱼发给当前塘熟练度。"),
    ("return_rules", "sizeGainMinM", "增重下限m", "塘内实体增重下限。"),
    ("return_rules", "sizeGainMaxM", "增重上限m", "塘内实体增重上限。"),
    ("return_rules", "sizeGainMode", "增重模式", "uniform_random=区间均匀随机。"),
    ("return_rules", "autoMinQuality", "自动回鱼最低品质", "FEAT-RETURN-02 如 purple。"),
    ("return_rules", "autoMinSizeRatio", "自动回鱼最低体长比", "相对品质 max，如 0.75。"),
    ("pond_levels", "level", "鱼塘等级", "1~10。"),
    ("pond_levels", "xpToNext", "升到下级所需塘经验", "10 级为 0。"),
    ("fish_species", "speciesId", "鱼种ID", "程序主键。"),
    ("fish_species", "name", "中文名", "商店/图鉴显示。"),
    ("fish_species", "diet", "食性", "herbivore 草食 / omnivore 杂食 / carnivore 肉食。"),
    ("fish_species", "catchGroup", "钓组", "still_bait 静水底钓 / stream_light 溪流轻口 / lure_predator 路亚掠食 / cast_heavy 重抛 / giant_game 巨物。"),
    ("fish_species", "typicalMinM", "典型最小体长m", "图鉴展示参考，不参与体长结算。"),
    ("fish_species", "typicalMaxM", "典型最大体长m", "图鉴展示参考，不参与体长结算。"),
    ("fish_species", "rarityTier", "稀有度档", "common/uncommon/rare/legendary；决定 qualityMax。"),
    ("fish_species", "nationwide", "全国广布", "TRUE 则所有 bioRegion 可出。"),
    ("fish_species", "qualityMin", "品质下限序", "恒为 1（gray）；播种品质带下限。"),
    ("fish_species", "qualityMax", "品质上限序", "1=gray…7=gold；稀有度抬高上限，与塘权重求交。"),
    ("pond_fish_pool", "pondId", "鱼塘ID", "该塘可出的鱼（仅种，不含品质）。"),
    ("pond_fish_pool", "speciesId", "鱼种ID", "种池条目。"),
    ("pond_fish_pool", "speciesName", "中文名", "只读展示，与 fish_species.name 一致。"),
    ("pond_fish_pool", "spawnWeight", "种刷新权重", "抽种相对权重；品质另走 pond_category_quality_weights。"),
    ("pond_fish_pool", "enabled", "是否启用", "FALSE 则该种不参与刷新。"),
    ("pond_category_quality_weights", "pondCategory", "塘分级", "novice…giant；播种/补充抽品质用。"),
    ("pond_category_quality_weights", "quality", "品质", "gray…gold。"),
    ("pond_category_quality_weights", "spawnWeight", "品质权重", "选好种后按此表加权抽品质。"),
    ("fish_quality_stats", "quality", "品质", "gray…gold；玩法+卖价同表。"),
    ("fish_quality_stats", "sizeCapM", "尺寸上限m", "品质体长硬帽。"),
    ("fish_quality_stats", "biteBaseAtMaxSize", "满尺寸咬钩基数", "公式读表。"),
    ("fish_quality_stats", "displayName", "显示名", "中文品质名。"),
    ("fish_quality_stats", "QUALITY_BASE", "品质底价", "卖价公式底数；钓组不参与卖价。"),
    ("fish_quality_stats", "SIZE_REF", "体长参考m", "卖价公式尺寸归一化。"),
    ("fish_quality_stats", "MIN_SELL", "最低卖价", "向下取整后的保底。"),
    ("fishing_formula_constants", "key", "常数键", "fishing.ts 公式常数名。"),
    ("fishing_formula_constants", "value", "值", "数值。"),
    ("fishing_formula_constants", "notes", "说明", "该常数用途注释。"),
    ("fish_xp", "speciesId", "鱼种ID", "与 fish_species 对齐。"),
    ("fish_xp", "speciesName", "中文名", "便于策划阅读。"),
    ("fish_xp", "quality", "品质", "gray/green/blue/purple/red/orange/gold。"),
    ("fish_xp", "playerXp", "玩家经验", "钓上该品质该种鱼给玩家的经验。"),
    ("fish_xp", "pondXp", "鱼塘经验", "同时给当前塘熟练度的经验。"),
    ("pond_modifiers", "category", "鱼塘分级", "与 ponds.pondCategory 对应。"),
    ("pond_modifiers", "biteRateMul", "咬钩倍率", "塘级对咬钩率的乘区。"),
    ("pond_modifiers", "escapeRateMul", "脱钩倍率", "塘级对脱钩率的乘区。"),
    ("pond_modifiers", "infoRevealMul", "信息揭示倍率", "线索/信息量。"),
    ("pond_modifiers", "qualityWeightSkew", "品质权重倾斜", ">1 偏向高品质。"),
    ("pond_modifiers", "sizeCapMul", "体长上限倍率", "乘在鱼种尺寸帽上。"),
    ("pond_modifiers", "pondXpMul", "塘经验倍率", "该分级塘熟练度获取。"),
    ("pond_modifiers", "fineChancePerHour", "每小时出警概率", "仅 forbidden 使用，初值 0.15。其它分级填 0。"),
    ("pond_modifiers", "fineGold", "超时罚款金币", "不足则归零。仅 forbidden 使用，初值 800。"),
    ("pond_modifiers", "policeWarningMs", "出警时限毫秒", "时限内离塘免罚。仅 forbidden 使用，初值 10000。"),
    ("rods", "rodId", "钓竿ID", "程序主键，商店与装备用。"),
    ("rods", "name", "中文名", "商店显示名。"),
    ("rods", "subType", "竿型", "手竿入门/台钓/溪流/矶钓/路亚/海竿/重路亚/巨物。"),
    ("rods", "priceGold", "金币价格", "0 表示新手赠送，商店不可回购。"),
    ("rods", "biteBonus", "咬钩加成", "加在咬钩判定上的弱加成，如 0.03 = +3%。"),
    ("rods", "escapeReduction", "防脱", "降低脱钩率的弱减免，如 0.04 = -4%。"),
    ("rods", "breakSizeM", "超规格体长m", "成功钓上大于该体长的鱼计入超规格。"),
    ("rods", "breakMaxLandings", "超规格成功上限", "累计满 N 次后当前装备竿销毁，不能修理。"),
    ("rods", "fitGray", "灰品质适配", "乘区；>1 更合适，<1 不合适。"),
    ("rods", "fitGreen", "绿品质适配", "乘区。"),
    ("rods", "fitBlue", "蓝品质适配", "乘区。"),
    ("rods", "fitPurple", "紫品质适配", "乘区。"),
    ("rods", "fitRed", "红品质适配", "乘区。"),
    ("rods", "fitOrange", "橙品质适配", "乘区。"),
    ("rods", "fitGold", "金品质适配", "乘区。"),
    ("rods", "fitStillBait", "静水底钓适配", "对应 catchGroup=still_bait。"),
    ("rods", "fitStreamLight", "溪流轻口适配", "对应 catchGroup=stream_light。"),
    ("rods", "fitLurePredator", "路亚掠食适配", "对应 catchGroup=lure_predator。"),
    ("rods", "fitCastHeavy", "重抛适配", "对应 catchGroup=cast_heavy。"),
    ("rods", "fitGiantGame", "巨物适配", "对应 catchGroup=giant_game。"),
    ("baits", "baitId", "鱼饵ID", "程序主键。"),
    ("baits", "name", "中文名", "商店显示名。"),
    ("baits", "diet", "对口食性", "any 通用 / herbivore 草食 / omnivore 杂食 / carnivore 肉食。"),
    ("baits", "unlockPlayerLevel", "解锁钓鱼等级", "0 表示开局可用。"),
    ("baits", "costGoldPerUse", "每次扣金", "咬钩成功时扣除；0 表示不扣金。不进货、无库存。"),
    ("baits", "biteBonusHerbivore", "对草食咬钩加成", "如 0.06 = +6%。"),
    ("baits", "biteBonusOmnivore", "对杂食咬钩加成", "如 0.06 = +6%。"),
    ("baits", "biteBonusCarnivore", "对肉食咬钩加成", "如 0.07 = +7%。"),
    ("baits", "isDefaultInfinite", "是否无限基础饵", "TRUE 的饵永不短缺，也不扣金。"),
    ("groundbaits", "groundbaitId", "窝料ID", "程序主键。"),
    ("groundbaits", "name", "中文名", "Overlay 显示名。"),
    ("groundbaits", "unlockPlayerLevel", "解锁钓鱼等级", "低于此等级不可打窝。"),
    ("groundbaits", "costGoldPerUse", "单次金币", "打窝开始时扣除。"),
    ("groundbaits", "castDurationMs", "打窝等待毫秒", "groundbaiting phase 时长。"),
    ("groundbaits", "durationMin", "持续分钟", "与 maxBites 先到失效。"),
    ("groundbaits", "maxBites", "持续口数", "与 durationMin 先到失效。"),
    ("groundbaits", "perStackBiteBonus", "每层咬钩参考", "策划标注；实际用 maxBonus/stackK 曲线。"),
    ("groundbaits", "maxBonus", "咬钩加成渐近上限", "bonus=maxBonus*(1-exp(-stackK*stack))。"),
    ("groundbaits", "stackK", "曲线速率", "非线性叠层速率。"),
    ("groundbaits", "sizeBonusPerStack", "尺寸每层增益m", "临时，不写库。"),
    ("groundbaits", "maxSizeBonus", "尺寸增益总帽m", "临时 sizeBonus 上限。"),
    ("vessels", "vesselId", "船具ID", "程序主键。"),
    ("vessels", "name", "中文名", "商店显示名。"),
    ("vessels", "unlockPlayerLevel", "解锁钓鱼等级", "低于此等级不能买。"),
    ("vessels", "priceGold", "金币价格", "一次性购入。"),
    ("vessels", "placeholderCatchCount", "占位捕捞次数", "表内预留，当前不生效。"),
    ("vessels", "enabledUse", "是否允许使用", "FALSE：可买但不可装备、不可开船。"),
    ("spot_clue_texts", "clueId", "线索ID", "唯一主键。"),
    ("spot_clue_texts", "clueType", "线索类型", "habitat 鱼喜环境 / activity 鱼情观察。"),
    ("spot_clue_texts", "clueText", "线索文案", "坐席后聊天气泡展示的中文。"),
    ("spot_clue_texts", "weight", "权重", "加权随机相对权重，默认 1。"),
    ("spot_clue_texts", "minPlayerLevel", "最低钓鱼等级", "未达到不进入抽选池。"),
    ("spot_clue_texts", "minPondLevel", "最低塘熟练度", "未达到不进入抽选池。"),
    ("spot_clue_texts", "pondCategory", "塘分级过滤", "空=全塘；否则仅匹配该分级。"),
    ("spot_clue_texts", "spotTag", "点位标签", "匹配 pond_spot_tags；空=不入池。"),
    ("spot_clue_texts", "activitySignal", "鱼情信号档", "habitat/active_high/active_mid/active_low/inactive/disturbed；供实时鱼群匹配。"),
    ("spot_clue_texts", "enabled", "是否启用", "FALSE 不参与抽选。"),
    ("spot_tag_defs", "tagId", "标签ID", "程序主键。"),
    ("spot_tag_defs", "tagCategory", "标签大类", "terrain/water/light/wind/depth/shore。"),
    ("spot_tag_defs", "nameZh", "中文名", "UI/策划显示。"),
    ("spot_tag_defs", "descriptionZh", "说明", "标签含义参考。"),
    ("pond_spot_tags", "pondId", "鱼塘ID", "与 ponds 对齐。"),
    ("pond_spot_tags", "spotId", "钓位ID", "如 calm-spot-3。"),
    ("pond_spot_tags", "tags", "标签列表", "逗号分隔多标签；每点 4–6 个跨类标签。"),
    ("achievements", "achievementId", "成就ID", "程序主键。"),
    ("achievements", "name", "名称", "展示名。"),
    ("achievements", "desc", "描述", "条件说明；隐藏成就未解锁时客户端不剧透。"),
    ("achievements", "iconKey", "图标键", "资源键。"),
    ("achievements", "category", "分类", "catch/social/progress/album。"),
    ("achievements", "conditionType", "条件类型", "catch_count/codex_count/return_count/max_size/album_pins。"),
    ("achievements", "conditionValue", "条件阈值", "达标阈值。"),
    ("achievements", "sortOrder", "排序", "展示序，小在前。"),
    ("achievements", "isHidden", "隐藏成就", "TRUE：未解锁不展示条件。"),
]


def style_header(ws) -> None:
    for cell in ws[1]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def autosize(ws, max_width: int = 28) -> None:
    for idx, col in enumerate(ws.columns, start=1):
        width = 10
        for cell in col:
            if cell.value is not None:
                width = max(width, min(max_width, len(str(cell.value)) + 2))
        ws.column_dimensions[get_column_letter(idx)].width = width


def write_sheet(wb: Workbook, title: str, headers: list[str], rows: list[tuple]) -> None:
    ws = wb.create_sheet(title)
    ws.append(headers)
    zh_row = [
        next((zh for sheet, field, zh, _desc in FIELD_DOCS if sheet == title and field == h), "")
        for h in headers
    ]
    if any(zh_row):
        ws.append(zh_row)
        for cell in ws[2]:
            cell.font = Font(italic=True, color="1F4E78")
            cell.alignment = Alignment(wrap_text=True, vertical="center")
    comments_map = {
        field: (zh, desc)
        for sheet, field, zh, desc in FIELD_DOCS
        if sheet == title
    }
    for idx, header in enumerate(headers, start=1):
        meta = comments_map.get(header)
        if not meta:
            continue
        zh, desc = meta
        ws.cell(1, idx).comment = Comment(f"{zh}\n{desc}".strip(), "策划")
    for row in rows:
        ws.append(list(row))
    style_header(ws)
    autosize(ws)


def write_field_docs(wb: Workbook) -> None:
    ws = wb.create_sheet("字段说明", 0)
    ws.append(["表名", "字段名", "中文名", "说明"])
    for row in FIELD_DOCS:
        ws.append(list(row))
    style_header(ws)
    autosize(ws, 56)
    ws.column_dimensions["D"].width = 64
    ws.freeze_panes = "A2"


def build() -> Path:
    wb = Workbook()
    # Drop the empty default sheet so the first append is not preceded by a blank row.
    default = wb.active
    wb.remove(default)

    write_field_docs(wb)
    write_sheet(wb, "_meta", list(META_ROWS[0]), META_ROWS[1:])

    pond_rows = []
    dual_categories = {"advanced", "veteran", "forbidden"}
    for p in PONDS:
        fee = p[4]
        category = p[2]
        dual = fee > 0 and category in dual_categories
        auto_fee = round(fee * 1.75) if dual else 0
        bio, water, ref = pond_extra_fields(p[0])
        mx, mn, init = ecology_fields_for_pond(p[0])
        pond_rows.append((*p[:5], fee, auto_fee, dual, *p[5:], 0, 0, bio, water, ref, mx, mn, init))
    write_sheet(wb, "ponds", PONDS_HEADER, pond_rows)

    write_sheet(
        wb,
        "player_levels",
        ["level", "xpToNext", "pondXpPerHour", "maxPondLevel"],
        PLAYER_LEVELS,
    )

    write_sheet(
        wb,
        "return_rules",
        [
            "minQuality",
            "minSizeRatio",
            "maxSizeRatio",
            "goldMulVsSell",
            "playerXp",
            "pondXp",
            "sizeGainMinM",
            "sizeGainMaxM",
            "sizeGainMode",
            "autoMinQuality",
            "autoMinSizeRatio",
        ],
        RETURN_RULES,
    )

    write_sheet(
        wb,
        "pond_levels",
        ["level", "xpToNext"],
        [(i + 1, xp) for i, xp in enumerate(POND_LEVEL_XP)],
    )

    write_sheet(
        wb,
        "fish_species",
        [
            "speciesId",
            "name",
            "diet",
            "catchGroup",
            "typicalMinM",
            "typicalMaxM",
            "rarityTier",
            "nationwide",
            "qualityMin",
            "qualityMax",
        ],
        FISH_SPECIES_FULL,
    )

    fish_xp_rows = []
    for sid, sname, _diet, group, _mn, _mx in FISH_SPECIES:
        coeff = CATCH_GROUP_COEFF[group]
        for q in QUALITIES:
            player_xp = round(XP_QUALITY_BASE[q] * coeff)
            pond_xp = round(player_xp * 0.5)
            fish_xp_rows.append((sid, sname, q, player_xp, pond_xp))
    write_sheet(
        wb,
        "fish_xp",
        ["speciesId", "speciesName", "quality", "playerXp", "pondXp"],
        fish_xp_rows,
    )

    write_sheet(
        wb,
        "pond_modifiers",
        [
            "category",
            "biteRateMul",
            "escapeRateMul",
            "infoRevealMul",
            "qualityWeightSkew",
            "sizeCapMul",
            "pondXpMul",
            "fineChancePerHour",
            "fineGold",
            "policeWarningMs",
        ],
        POND_MODIFIERS,
    )

    write_sheet(
        wb,
        "pond_fish_pool",
        ["pondId", "speciesId", "speciesName", "spawnWeight", "enabled"],
        POND_FISH_POOL,
    )

    write_sheet(
        wb,
        "pond_category_quality_weights",
        ["pondCategory", "quality", "spawnWeight"],
        POND_CATEGORY_QUALITY_WEIGHTS,
    )

    write_sheet(
        wb,
        "fish_quality_stats",
        [
            "quality",
            "sizeCapM",
            "biteBaseAtMaxSize",
            "displayName",
            "QUALITY_BASE",
            "SIZE_REF",
            "MIN_SELL",
        ],
        FISH_QUALITY_STATS,
    )

    write_sheet(
        wb,
        "fishing_formula_constants",
        ["key", "value", "notes"],
        FISHING_FORMULA_CONSTANTS,
    )

    rod_headers = [
        "rodId",
        "name",
        "subType",
        "priceGold",
        "biteBonus",
        "escapeReduction",
        "breakSizeM",
        "breakMaxLandings",
        "fitGray",
        "fitGreen",
        "fitBlue",
        "fitPurple",
        "fitRed",
        "fitOrange",
        "fitGold",
        "fitStillBait",
        "fitStreamLight",
        "fitLurePredator",
        "fitCastHeavy",
        "fitGiantGame",
    ]
    write_sheet(wb, "rods", rod_headers, RODS)

    write_sheet(
        wb,
        "baits",
        [
            "baitId",
            "name",
            "diet",
            "unlockPlayerLevel",
            "costGoldPerUse",
            "biteBonusHerbivore",
            "biteBonusOmnivore",
            "biteBonusCarnivore",
            "isDefaultInfinite",
        ],
        BAITS,
    )

    write_sheet(
        wb,
        "groundbaits",
        [
            "groundbaitId",
            "name",
            "unlockPlayerLevel",
            "costGoldPerUse",
            "castDurationMs",
            "durationMin",
            "maxBites",
            "perStackBiteBonus",
            "maxBonus",
            "stackK",
            "sizeBonusPerStack",
            "maxSizeBonus",
        ],
        GROUNDBAITS,
    )

    write_sheet(
        wb,
        "vessels",
        [
            "vesselId",
            "name",
            "unlockPlayerLevel",
            "priceGold",
            "placeholderCatchCount",
            "enabledUse",
        ],
        VESSELS,
    )

    write_sheet(
        wb,
        "spot_tag_defs",
        ["tagId", "tagCategory", "nameZh", "descriptionZh"],
        SPOT_TAG_DEFS,
    )

    write_sheet(
        wb,
        "pond_spot_tags",
        ["pondId", "spotId", "tags"],
        build_pond_spot_tag_rows(),
    )

    write_sheet(
        wb,
        "spot_clue_texts",
        [
            "clueId",
            "clueType",
            "clueText",
            "weight",
            "minPlayerLevel",
            "minPondLevel",
            "pondCategory",
            "spotTag",
            "activitySignal",
            "enabled",
        ],
        SPOT_CLUE_TEXTS,
    )

    write_sheet(
        wb,
        "achievements",
        [
            "achievementId",
            "name",
            "desc",
            "iconKey",
            "category",
            "conditionType",
            "conditionValue",
            "sortOrder",
            "isHidden",
        ],
        ACHIEVEMENTS,
    )

    try:
        wb.save(OUT)
        print(f"Wrote {OUT}")
    except PermissionError:
        alt = OUT.with_name(OUT.stem + ".next.xlsx")
        wb.save(alt)
        print(f"WARNING: {OUT.name} is locked; wrote {alt.name} instead.")
        print("Close Excel and re-run: npm run game-data:build")
        print(f"Sheets: {', '.join(wb.sheetnames)}")
        return alt
    print(f"Sheets: {', '.join(wb.sheetnames)}")
    return OUT


def main() -> None:
    build()


if __name__ == "__main__":
    main()
