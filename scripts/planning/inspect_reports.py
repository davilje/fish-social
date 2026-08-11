import openpyxl
import os

BASE = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
REPORTS_DIR = os.path.join(BASE, 'docs', 'planning', 'reports')

for fname in sorted(os.listdir(REPORTS_DIR)):
    fpath = os.path.join(REPORTS_DIR, fname)
    if fname.endswith('.xlsx'):
        try:
            wb = openpyxl.load_workbook(fpath, read_only=True, data_only=True)
            print(f'\n=== {fname} ===')
            print(f'  Sheets: {wb.sheetnames}')
            for s in wb.sheetnames[:3]:
                ws = wb[s]
                rows = list(ws.iter_rows(min_row=1, max_row=min(3, ws.max_row), values_only=True))
                for i, r in enumerate(rows):
                    vals = [str(v)[:30] if v else '' for v in (r or [])]
                    print(f'  {s} R{i+1}: {vals}')
            wb.close()
        except Exception as e:
            print(f'\n=== {fname} === ERROR: {e}')
    elif fname.endswith('.md'):
        print(f'\n=== {fname} === (md)')
