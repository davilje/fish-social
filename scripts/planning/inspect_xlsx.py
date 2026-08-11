import openpyxl

wb = openpyxl.load_workbook('docs/planning/specs/项目开发需求计划表.xlsx')
for s in wb.sheetnames:
    ws = wb[s]
    print(f'Sheet: {s} ({ws.max_row}r x {ws.max_column}c)')
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=min(3, ws.max_row), values_only=True)):
        vals = [str(v)[:40] if v else '' for v in row]
        print(f'  R{i+1}: {vals}')
    print()

wb2 = openpyxl.load_workbook('docs/planning/reports/三层数据体系-开发需求清单.xlsx')
print('=== 三层数据体系 ===')
for s in wb2.sheetnames:
    ws = wb2[s]
    print(f'Sheet: {s} ({ws.max_row}r x {ws.max_column}c)')
