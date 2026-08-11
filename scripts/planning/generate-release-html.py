"""
Generate release HTML from master xlsx
Usage: python scripts/planning/generate-release-html.py [--version vX.Y.Z]
"""
import openpyxl, os, argparse
from datetime import datetime

BASE = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
PLAN_XLSX = os.path.join(BASE, 'docs', 'planning', '项目开发需求计划表.xlsx')

COLORS = {'数据平台':'#4A9EFF','架构':'#9B59B6','Bug修复':'#E74C3C','数值':'#F39C12'}

def load(path):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    data = {}
    for s in wb.sheetnames:
        ws = wb[s]
        rows = []
        for row in ws.iter_rows(min_row=1, values_only=True):
            rows.append([str(v) if v is not None else '' for v in row])
        data[s] = rows
    wb.close()
    return data

def gen(version='v0.5.0', date=None):
    data = load(PLAN_XLSX)
    if not date:
        date = datetime.now().strftime('%Y-%m-%d')

    plan = data.get('开发计划', [])
    headers = plan[0] if plan else []
    done_items = [r for r in plan[1:] if len(r) > 0 and r[0] == '已实现']

    by_type = {}
    for item in done_items:
        typ = item[2] if len(item) > 2 else ''
        by_type.setdefault(typ, [])
        by_type[typ].append({'code': item[1], 'name': item[3], 'desc': item[7] if len(item) > 7 else ''})

    # Build HTML
    parts = [f'''<!DOCTYPE html>
<html lang="zh-CN">
<head><meta charset="utf-8"><meta name="viewport" content="width=device-width,initial-scale=1">
<title>发布说明 {version} - Fish Social</title>
<style>
:root {{ --bg:#0f1419; --surface:#1a2332; --text:#e8edf2; --muted:#8b9cb3; --accent:#4a9eff; --border:#2d3a4d; --ok:#48bb78; }}
* {{ box-sizing:border-box; margin:0; padding:0; }}
body {{ font-family:"Segoe UI","PingFang SC",sans-serif; background:var(--bg); color:var(--text); padding:2rem; line-height:1.6; }}
.wrap {{ max-width:800px; margin:0 auto; }}
h1 {{ font-size:1.8rem; margin-bottom:.3rem; }}
.meta {{ color:var(--muted); font-size:.9rem; margin-bottom:2rem; }}
.summary {{ display:flex; gap:1rem; margin-bottom:2rem; flex-wrap:wrap; }}
.stat {{ background:var(--surface); border:1px solid var(--border); border-radius:8px; padding:1rem 1.5rem; text-align:center; min-width:100px; flex:1; }}
.stat-value {{ font-size:1.6rem; font-weight:bold; }}
.stat-label {{ font-size:.8rem; color:var(--muted); margin-top:.25rem; }}
.section {{ margin-bottom:2rem; }}
.section h2 {{ font-size:1.2rem; margin-bottom:1rem; padding-bottom:.5rem; border-bottom:2px solid var(--border); display:flex; align-items:center; gap:.5rem; }}
.tag {{ display:inline-block; padding:.15rem .6rem; border-radius:4px; font-size:.75rem; color:#fff; }}
.items {{ list-style:none; }}
.items li {{ background:var(--surface); border:1px solid var(--border); border-radius:6px; padding:.75rem 1rem; margin-bottom:.5rem; }}
.items li .code {{ font-family:monospace; color:var(--accent); font-size:.85rem; }}
.items li .desc {{ color:var(--muted); font-size:.85rem; margin-top:.25rem; }}
.footer {{ margin-top:3rem; padding-top:1rem; border-top:1px solid var(--border); font-size:.85rem; color:var(--muted); text-align:center; }}
</style></head><body><div class="wrap">
<h1>发布说明 {version}</h1>
<p class="meta">{date}</p>
<div class="summary">
<div class="stat"><div class="stat-value">{len(done_items)}</div><div class="stat-label">已完成</div></div>
<div class="stat"><div class="stat-value">{len(by_type)}</div><div class="stat-label">类型</div></div>
''']
    for typ, items in by_type.items():
        parts.append(f'<div class="stat"><div class="stat-value">{len(items)}</div><div class="stat-label">{typ}</div></div>\n')
    parts.append('</div>\n')

    for typ, items in by_type.items():
        c = COLORS.get(typ, '#666')
        parts.append(f'<div class="section"><h2><span class="tag" style="background:{c}">{typ}</span> ({len(items)})</h2><ul class="items">\n')
        for it in items:
            d = f'<div class="desc">{it["desc"]}</div>' if it['desc'] else ''
            parts.append(f'<li><span class="code">{it["code"]}</span> {it["name"]}{d}</li>\n')
        parts.append('</ul></div>\n')

    # List all sheets
    parts.append(f'<div class="section"><h2>总表 Sheets ({len(data)})</h2><ul class="items">\n')
    for s in data:
        n = len(data[s]) - 1
        parts.append(f'<li><span class="code">{s}</span> ({n} rows)</li>\n')
    parts.append('</ul></div>\n')

    parts.append(f'<div class="footer"><p>生成自 <code>项目开发需求计划表.xlsx</code> · ' +
                 f'<a href="../项目开发需求计划表.xlsx" style="color:var(--accent)">下载 xlsx</a></p></div>\n')
    parts.append('</div></body></html>')

    out_dir = os.path.join(BASE, 'docs', 'planning', 'releases')
    os.makedirs(out_dir, exist_ok=True)
    out = os.path.join(out_dir, f'{version}.html')
    with open(out, 'w', encoding='utf-8') as f:
        f.write('\n'.join(parts))
    print(f'Release HTML: {out}')

if __name__ == '__main__':
    p = argparse.ArgumentParser()
    p.add_argument('--version', default='v0.5.0')
    p.add_argument('--date', default=datetime.now().strftime('%Y-%m-%d'))
    args = p.parse_args()
    gen(args.version, args.date)
