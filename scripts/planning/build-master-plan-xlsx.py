#!/usr/bin/env python3
"""Regenerate the planning workbook without locale-dependent text handling.

The previous generator was a large, hand-maintained Python literal that had
become mojibake and syntactically invalid.  The workbook is the durable
planning artifact, so this generator updates that workbook in place, syncing
statuses from the authoritative spec metadata while preserving all sheets and
rows.
"""

from __future__ import annotations

import re
import shutil
import subprocess
import sys
from datetime import date
from pathlib import Path

import openpyxl


ROOT = Path(__file__).resolve().parents[2]
DOCS_PLANNING = ROOT / "docs" / "planning"
SPEC_DIR = DOCS_PLANNING / "specs"
TODAY = date.today().isoformat()

# Rows that must exist on the first planning sheet (upsert by 编号).
# Columns: 当前状态, 编号, 类型, 需求名称, 层级, 版本/阶段, 优先级, 说明, 文档路径, 设计时间, 完成时间
ENSURE_PLAN_ROWS: list[list[object]] = [
    [
        "已实现",
        "BUG-18",
        "Bug修复",
        "进塘首帧状态与演示降级",
        "—",
        "hotfix",
        "P0",
        "进塘清空旧态；禁静默DEMO；收杆ack最终额度；快照门禁",
        "docs/planning/specs/BUG修复-进塘首帧状态与演示降级.md",
        "2026-08-10",
        "2026-08-10",
    ],
    [
        "已实现",
        "BUG-19",
        "Bug修复",
        "每日额度单一口径重构",
        "—",
        "hotfix",
        "P0",
        "拆分base/session锚点；checkpoint不前移展示锚点；客户端禁反推",
        "docs/planning/specs/BUG修复-每日额度单一口径重构.md",
        "2026-08-10",
        "2026-08-10",
    ],
    [
        "已实现",
        "BUG-20",
        "Bug修复",
        "进塘与钓鱼剩余展示回归",
        "—",
        "hotfix",
        "P0",
        "BUG-19后：冻结基线+墙钟走动；未选钓点join ack种子防满额8h",
        "docs/planning/specs/BUG修复-进塘与钓鱼剩余展示回归.md",
        "2026-08-10",
        "2026-08-10",
    ],
    [
        "已实现",
        "BUG-21",
        "Bug修复",
        "桌面端关闭后进程残留",
        "—",
        "hotfix",
        "P0",
        "关闭按钮真正退出；托盘显式隐藏；单实例互斥；清理托盘线程",
        "docs/planning/specs/BUG修复-桌面端关闭后进程残留.md",
        "2026-08-12",
        "2026-08-12",
    ],
    [
        "已实现",
        "BUG-22",
        "Bug修复",
        "Steam Lobby 创建权限拒绝与状态残留",
        "—",
        "hotfix",
        "P0",
        "细分 Lobby 权限错误；校验 Steam 绑定；创建失败回滚 CurrentLobbyId/pondId；不影响鱼塘生命周期",
        "docs/planning/specs/BUG修复-SteamLobby创建权限与状态残留.md",
        "2026-08-13",
        "2026-08-13",
    ],
    [
        "已废弃",
        "STEAM-DESKTOP-06",
        "功能优化",
        "Steam Lobby 生命周期与邀请反馈优化",
        "—",
        "v1.0-steam-desktop",
        "P1",
        "本版本跳过；不进入开发排期，后续如有正式 Lobby 产品方案再重新立项",
        "docs/planning/specs/Steam Lobby生命周期与邀请反馈优化.md",
        "2026-08-13",
        "2026-08-13",
    ],
    [
        "已实现",
        "STEAM-DESKTOP-07A",
        "功能",
        "桌面宠物主视图与鱼塘入口",
        "—",
        "v1.0-steam-desktop",
        "P0",
        "显示自己的 2D 猫咪、钓鱼状态和鱼塘入口；采用序列帧+状态机，复用已完成桌面壳",
        "docs/planning/specs/Steam桌面宠物与多人鱼塘表现层.md",
        "2026-08-13",
        "2026-08-14",
    ],
    [
        "已实现",
        "STEAM-DESKTOP-07B",
        "功能",
        "2D 鱼塘环境与自己的猫咪",
        "—",
        "v1.0-steam-desktop",
        "P0",
        "显示池塘环境、钓位、自己的宠物和钓鱼表现",
        "docs/planning/specs/Steam桌面宠物与多人鱼塘表现层.md",
        "2026-08-13",
        "2026-08-14",
    ],
    [
        "已实现",
        "STEAM-DESKTOP-07C",
        "功能",
        "同塘玩家宠物与状态同步",
        "—",
        "v1.0-steam-desktop",
        "P0",
        "在 Overlay 渲染 pond_user_joined/left/updated；128×128 统一序列帧；fishingPhase→petVisualState；IPC 不传图、不新开协议",
        "docs/planning/specs/Steam桌面宠物与多人鱼塘表现层.md",
        "2026-08-13",
        "2026-08-15",
    ],
    [
        "已实现",
        "STEAM-DESKTOP-07D",
        "功能",
        "桌面宠物右键菜单",
        "—",
        "v1.0-steam-desktop",
        "P0",
        "在产品窗口/宠物区域提供鱼塘、好友、背包、图鉴、设置、托盘和退出入口",
        "docs/planning/specs/Steam桌面宠物与多人鱼塘表现层.md",
        "2026-08-13",
        "2026-08-16",
    ],
    [
        "已实现",
        "STEAM-DESKTOP-07E",
        "功能",
        "桌面宠物主窗口功能页签",
        "—",
        "v1.0-steam-desktop",
        "P0",
        "好友/聊天、背包、图鉴、设置为主窗口页签；Overlay 菜单切页且主窗口高于 Overlay；不用功能弹窗",
        "docs/planning/specs/Steam桌面端Web功能对齐设计.md",
        "2026-08-13",
        "2026-08-17",
    ],
    [
        "已实现",
        "STEAM-DESKTOP-07F",
        "功能",
        "桌面宠物主流程与恢复验收",
        "—",
        "v1.0-steam-desktop",
        "P0",
        "串联登录、进塘、挂机、通知、托盘、收鱼、断线恢复并完成 Windows 验收",
        "docs/planning/specs/Steam桌面宠物与多人鱼塘表现层.md",
        "2026-08-13",
        "2026-08-17",
    ],
    [
        "已实现",
        "STEAM-DESKTOP-07G",
        "功能",
        "原生桌面宠物 Overlay",
        "—",
        "v1.0-steam-desktop",
        "P1",
        "独立 WPF/Win32 Overlay；不启动第二个 Unity Player；通过 Named Pipe 接收状态并发送命令",
        "docs/planning/specs/Steam原生桌面宠物Overlay.md",
        "2026-08-14",
        "2026-08-17",
    ],
    [
        "已实现",
        "STEAM-DESKTOP-08A",
        "功能",
        "世界地图与鱼塘选择",
        "—",
        "v1.0-steam-desktop",
        "P0",
        "主窗口超大地图 Image：拖动/滚轮缩放/边界；pondId 坐标绑定；点击进塘并显示 Overlay",
        "docs/planning/specs/Steam桌面端-08A世界地图与鱼塘选择.md",
        "2026-08-17",
        "2026-08-17",
    ],
    [
        "已实现",
        "STEAM-DESKTOP-08B",
        "功能",
        "商店与装备",
        "—",
        "v1.0-steam-desktop",
        "P0",
        "PanelShop：鱼饵/渔具购买与装备；金币库存以服务端为准；Overlay 只 menu_shop 切页",
        "docs/planning/specs/Steam桌面端-08B商店与装备.md",
        "2026-08-17",
        "2026-08-18",
    ],
    [
        "已实现",
        "STEAM-DESKTOP-08C",
        "功能",
        "动态墙与好友动态",
        "—",
        "v1.0-steam-desktop",
        "P1",
        "PanelSocialFeed：公共/好友动态、分页加载、动态卡片 Prefab、点赞、评论、删除评论；不塞进 PanelSocial",
        "docs/planning/specs/Steam桌面端-08C动态墙与好友动态.md",
        "2026-08-17",
        "2026-08-19",
    ],
    [
        "已实现",
        "STEAM-DESKTOP-08D",
        "功能",
        "排行榜",
        "—",
        "v1.0-steam-desktop",
        "P1",
        "PanelLeaderboard：日/周/鱼塘/稀有四榜；Top3 固定 Image 领奖台；4 名起纵向滚动；分数以服务端为准",
        "docs/planning/specs/Steam桌面端-08D排行榜.md",
        "2026-08-17",
        "2026-08-20",
    ],
    [
        "已实现",
        "FEAT-SOC-03b",
        "功能",
        "排行榜入口与领奖台改版",
        "—",
        "v0.6.1",
        "P0",
        "首页入口、独立页、领奖台、仅日/周、删稀有 UI、周榜改最大鱼",
        "docs/planning/specs/排行榜-入口与领奖台改版.md",
        "2026-07-14",
        "2026-08-19",
    ],
    [
        "已实现",
        "STEAM-DESKTOP-08E",
        "功能",
        "个人中心与资料编辑",
        "—",
        "v1.0-steam-desktop",
        "P1",
        "PanelProfile/Edit：昵称头像展示鱼获；不改 SteamID、不迁移旧档",
        "docs/planning/specs/Steam桌面端-08E个人中心与资料编辑.md",
        "2026-08-17",
        "2026-08-18",
    ],
    [
        "已实现",
        "STEAM-DESKTOP-08F",
        "功能",
        "好友列表与申请 Prefab",
        "—",
        "v1.0-steam-desktop",
        "P0",
        "独立 PanelFriends；修复接受/拒绝重叠不可点；不与 PanelSocial 双开好友 UI",
        "docs/planning/specs/Steam桌面端-08F好友列表与申请Prefab.md",
        "2026-08-17",
        "2026-08-17",
    ],
    [
        "已实现",
        "STEAM-DESKTOP-08G",
        "功能",
        "Overlay 钓鱼操作栏",
        "—",
        "v1.0-steam-desktop",
        "P0",
        "Overlay 底部选钓位/开始/收杆/领鱼获；Pipe 发命令；Unity/Node 保留权威",
        "docs/planning/specs/Steam桌面端-08GOverlay钓鱼操作栏.md",
        "2026-08-17",
        "2026-08-17",
    ],
    [
        "已实现",
        "STEAM-DESKTOP-08H",
        "工程",
        "全量 UI 预制体化与动态内容组件",
        "—",
        "v1.0-steam-desktop",
        "P0",
        "壳层、功能页、列表行和 Grid item 全部使用合法 Prefab；运行时代码只绑定数据和事件",
        "docs/planning/specs/Steam桌面端-08H全量UI预制体化.md",
        "2026-08-17",
        "2026-08-17",
    ],
    [
        "已实现",
        "STEAM-DESKTOP-08I",
        "Bug 修复 / 功能优化",
        "鱼塘退出与跨塘切换优化",
        "—",
        "v1.0-steam-desktop",
        "P0",
        "Overlay 点击钓位直接入座；新增离席/退出鱼塘；收杆领鱼离席后跨塘切换；修复隐藏主窗口时 5 FPS 导致的 ACK 延迟，增加 ACK 优先队列、持久状态序号、重复状态丢弃及三段链路时间戳验收",
        "docs/planning/specs/Steam桌面端-08I鱼塘退出与跨塘切换优化.md",
        "2026-08-18",
        "2026-08-18",
    ],
    [
        "已实现",
        "STEAM-DESKTOP-09A",
        "功能",
        "Overlay 玩家右键菜单",
        "—",
        "v1.0-steam-desktop",
        "P1",
        "右键同塘玩家：资料、加好友、私聊、点赞；Overlay IPC → Unity 权威",
        "docs/planning/specs/Steam桌面端-09AOverlay玩家右键菜单.md",
        "2026-08-19",
        "2026-08-20",
    ],
    [
        "已实现",
        "STEAM-DESKTOP-09B",
        "功能",
        "Overlay 悬停状态与钓鱼时长",
        "—",
        "v1.0-steam-desktop",
        "P1",
        "悬停 300ms 仅时长 Tooltip；IPC sessionFishingMs/phase/hookDeadlineMs",
        "docs/planning/specs/Steam桌面端-09BOverlay悬停状态与钓鱼时长.md",
        "2026-08-19",
        "2026-08-20",
    ],
    [
        "已实现",
        "STEAM-DESKTOP-09C",
        "功能",
        "Overlay 鱼塘聊天气泡与输入",
        "—",
        "v1.0-steam-desktop",
        "P1",
        "Overlay 公屏最近 20 条气泡 + 紧凑输入；IPC recentChats/send_pond_chat；完整聊天仍主窗口",
        "docs/planning/specs/Steam桌面端-09COverlay鱼塘聊天气泡与输入.md",
        "2026-08-19",
        "2026-08-20",
    ],
    [
        "已实现",
        "STEAM-DESKTOP-09D",
        "功能",
        "Overlay 布局与角色表现优化",
        "—",
        "v1.0-steam-desktop",
        "P0",
        "960×560；宠物 64px；昵称在上；默认状态/上钩圆环；悬停仅时长；左上纵向收纳菜单；Bot 无「·机」",
        "docs/planning/specs/Steam桌面端-09DOverlay布局与角色表现优化.md",
        "2026-08-20",
        "2026-08-20",
    ],
    [
        "已确认",
        "EPIC-FISH-V07",
        "Epic",
        "钓鱼玩法扩展 v0.7 索引",
        "—",
        "v0.7",
        "P0",
        "15 条覆盖与 R1 拆票索引；Steam 优先",
        "docs/planning/specs/钓鱼玩法扩展-v0.7-Epic.md",
        "2026-08-21",
        "",
    ],
    [
        "已实现",
        "FEAT-PROG-01",
        "功能",
        "鱼塘分级与玩家成长",
        "—",
        "v0.7",
        "P0",
        "七类塘、双熟练度、每2h扣费、新手塘、卖价公式、固定数值表导出",
        "docs/planning/specs/鱼塘分级与玩家成长.md",
        "2026-08-21",
        "2026-08-21",
    ],
    [
        "已实现",
        "STEAM-DESKTOP-11",
        "功能",
        "新手引导本地教学关",
        "—",
        "v0.7",
        "P0",
        "Overlay 本地教学关；5秒必上钩/圆圈；自动入包弹窗",
        "docs/planning/specs/Steam桌面端-新手引导本地教学关.md",
        "2026-08-21",
        "2026-08-21",
    ],
    [
        "已实现",
        "FEAT-GEAR-01",
        "功能",
        "钓具与鱼饵配置",
        "—",
        "v0.7",
        "P0",
        "竿弱加成与断竿买新、饵等级解锁按次扣金、船仅商店不可用",
        "docs/planning/specs/钓具与鱼饵配置.md",
        "2026-08-21",
        "2026-08-22",
    ],
    [
        "已实现",
        "STEAM-DESKTOP-08A2",
        "功能",
        "世界地图分区与进塘扣费确认",
        "—",
        "v0.7",
        "P0",
        "六区展示锁态、巨物暂闭、进塘扣费确认；读 ponds.json",
        "docs/planning/specs/Steam桌面端-08A2世界地图分区与进塘扣费.md",
        "2026-08-21",
        "2026-08-22",
    ],
    [
        "已实现",
        "FEAT-RISK-01",
        "功能",
        "禁止钓鱼塘巡警事件",
        "—",
        "v0.7",
        "P1",
        "巡警气泡10秒离塘免罚；超时罚款归零+当日禁塘；开发版禁止塘Overlay一键出警",
        "docs/planning/specs/禁止钓鱼塘巡警事件.md",
        "2026-08-21",
        "2026-08-22",
    ],
    [
        "已实现",
        "FEAT-SPOT-01",
        "功能",
        "钓位点位线索文字泡",
        "—",
        "v0.7",
        "P1",
        "坐席后聊天泡；文案表随机；habitat/activity两类经验",
        "docs/planning/specs/钓位点位线索文字泡.md",
        "2026-08-21",
        "2026-08-22",
    ],
    [
        "已实现",
        "STEAM-DESKTOP-12",
        "功能",
        "玩法 Debug 菜单",
        "—",
        "v0.7",
        "P1",
        "开发版呼出菜单：升级/满级/塘熟练度/加钱/出警/鱼获/+2h扣费测",
        "docs/planning/specs/Steam桌面端-玩法Debug菜单.md",
        "2026-08-22",
        "2026-08-22",
    ],
    [
        "已实现",
        "STEAM-DESKTOP-12B",
        "功能",
        "Debug鱼塘钓位查看",
        "—",
        "hotfix",
        "P1",
        "Overlay Debug三列：塘/钓位鱼列表与属性、钓位数据、强制上钩；宽×1.3",
        "docs/planning/specs/Steam桌面端-Debug鱼塘钓位查看.md",
        "2026-08-25",
        "2026-08-26",
    ],
    [
        "已实现",
        "FEAT-RETURN-01",
        "功能",
        "回鱼机制",
        "—",
        "v0.7",
        "P0",
        "回塘增重、紫品+体长比≥75%准入、金=卖价×1.5+熟练度；仅收费塘可回",
        "docs/planning/specs/回鱼机制.md",
        "2026-08-23",
        "2026-08-23",
    ],
    [
        "已实现",
        "FEAT-RETURN-02",
        "功能",
        "双价塘与自动回鱼",
        "—",
        "v0.7",
        "P1",
        "进塘二选一收费；可回鱼档达标自动回鱼、不达标留包；免票塘不可回",
        "docs/planning/specs/双价塘与自动回鱼.md",
        "2026-08-23",
        "2026-08-24",
    ],
    [
        "已实现",
        "FEAT-FISH-CN-01",
        "功能",
        "中国鱼种与区域分布",
        "—",
        "v0.7",
        "P0",
        "50种本土鱼库；塘真实钓场名；pond_fish_pool出鱼；pondId不变",
        "docs/planning/specs/中国鱼种与区域分布.md",
        "2026-08-24",
        "2026-08-25",
    ],
    [
        "已实现",
        "FEAT-POOL-01",
        "功能",
        "塘生态数值表驱动",
        "—",
        "v0.7",
        "P0",
        "种池表驱动；按塘类型品质权重；品质尺寸/咬钩脱钩基数入表；公式不变",
        "docs/planning/specs/塘生态数值表驱动.md",
        "2026-08-24",
        "2026-08-25",
    ],
    [
        "已实现",
        "FEAT-POOL-02",
        "功能",
        "塘人口与品质权重调参",
        "—",
        "v0.7",
        "P1",
        "新手1条不补；公开塘上限100；补鱼60min；老手/野外水库/禁止品质占比定稿",
        "docs/planning/specs/塘人口与品质权重调参.md",
        "2026-08-27",
        "2026-08-27",
    ],
    [
        "已实现",
        "FEAT-SPOT-02",
        "功能",
        "钓位标签与线索库v2",
        "—",
        "v0.7",
        "P1",
        "22类标签×420点位；154条合规线索；activitySignal；标签过滤线索；体长XP^0.85",
        "docs/planning/specs/钓位标签与线索库-v2.md",
        "2026-08-25",
        "2026-08-25",
    ],
    [
        "已确认",
        "FEAT-SPOT-03",
        "内容",
        "钓位标签地形微调",
        "—",
        "v0.7",
        "P2",
        "xlsx pond_spot_tags 420点按Tile岸位与真实地形人工微调标签",
        "docs/planning/specs/钓位标签地形微调.md",
        "2026-08-25",
        "",
    ],
    [
        "已确认",
        "FEAT-SPOT-04",
        "内容",
        "线索文案策划审校",
        "—",
        "v0.7",
        "P2",
        "spot_clue_texts 增删文案；遵守 habitat/activity 禁区；每 tag≥10 条",
        "docs/planning/specs/线索文案策划审校.md",
        "2026-08-25",
        "",
    ],
    [
        "已确认",
        "FEAT-SPOT-05",
        "功能",
        "钓位鱼情联动线索",
        "—",
        "v0.7",
        "P1",
        "生态 v2：钓位鱼量/迁移/受惊→computedTier；activity 线索匹配 activitySignal",
        "docs/planning/specs/钓位鱼情联动线索.md",
        "2026-08-25",
        "",
    ],
    [
        "已实现",
        "FEAT-RETURN-03",
        "功能",
        "回鱼规则调优",
        "—",
        "v0.7",
        "P1",
        "准入 purple/0.75 统一；回鱼金=卖价×1.5；仅收费塘可回；不达标留包",
        "docs/planning/specs/回鱼规则调优.md",
        "2026-08-24",
        "2026-08-24",
    ],
    [
        "已实现",
        "FEAT-RETURN-04",
        "功能",
        "自动回鱼体验闭环",
        "—",
        "v0.7",
        "P1",
        "回鱼档钓到即回；结算页；fish_catch_settled；Debug顶层弹窗；Debug发鱼不自动回",
        "docs/planning/specs/自动回鱼体验闭环.md",
        "2026-08-25",
        "2026-08-25",
    ],
    [
        "已实现",
        "FEAT-RETURN-05",
        "功能",
        "回鱼体重分档",
        "—",
        "v0.7",
        "P1",
        "准入改体重>10斤；>100斤回鱼金×3、>10斤×1.5；废弃紫品/体长比门槛",
        "docs/planning/specs/回鱼体重分档.md",
        "2026-08-27",
        "2026-08-27",
    ],
    [
        "已实现",
        "BUG-23",
        "Bug修复",
        "日额度满后仍可开钓",
        "—",
        "hotfix",
        "P0",
        "扣满4次/满8h不可start；可落座；对齐Web",
        "docs/planning/specs/BUG修复-日额度满后仍可开钓.md",
        "2026-08-23",
        "2026-08-23",
    ],
    [
        "已实现",
        "FEAT-GROUND-01",
        "功能",
        "打窝机制",
        "—",
        "v0.7",
        "P1",
        "Overlay并列循环打窝；50层非线性；0.5~1.5%/层曲线；附近鱼小幅尺寸",
        "docs/planning/specs/打窝机制.md",
        "2026-08-23",
        "2026-08-24",
    ],
    [
        "已实现",
        "FEAT-ALBUM-01",
        "功能",
        "钓鱼相册与成就",
        "—",
        "v0.7",
        "P1",
        "个人中心大改：资料+展示柜+图鉴+相册墙+成就；他人可见精选",
        "docs/planning/specs/钓鱼相册与成就.md",
        "2026-08-23",
        "2026-08-23",
    ],
    [
        "已实现",
        "STEAM-DESKTOP-10",
        "功能",
        "公网联调与服务器地址配置",
        "—",
        "v1.0-steam-desktop",
        "P1",
        "可配置 serverBaseUrl（server.json/环境变量）；本机开发可迁云；联调手册",
        "docs/planning/specs/Steam桌面端公网联调与服务器地址配置.md",
        "2026-08-20",
        "2026-08-20",
    ],
    [
        "已废弃",
        "STEAM-DESKTOP-10A",
        "功能",
        "本机公网映射联调（方案 A）",
        "—",
        "v1.0-steam-desktop",
        "P1",
        "CGNAT不可行已废弃；设置页/检查脚本保留给局域网；外网改10B",
        "docs/planning/specs/Steam桌面端本机公网映射联调.md",
        "2026-08-20",
        "2026-08-20",
    ],
    [
        "已确认",
        "STEAM-DESKTOP-10B",
        "运维联调",
        "云服务器联调与切换（保留项）",
        "—",
        "v1.0-steam-desktop",
        "P2",
        "方案已就绪(Windows手册)；实施可暂缓；上云仅改server.json；见docs/ops/windows-cloud-deploy.md",
        "docs/planning/specs/Steam桌面端云服务器联调保留.md",
        "2026-08-20",
        "",
    ],
    [
        "已确认",
        "STEAM-DESKTOP-ART-01",
        "美术",
        "桌面宠物与鱼塘视觉资源替换",
        "—",
        "v1.0-steam-desktop",
        "P0",
        "提供猫咪宠物、鱼塘环境和基础视觉资源；其他 UI 先由程序使用通用资源",
        "docs/planning/specs/Steam桌面宠物UI需求拆分.md",
        "2026-08-13",
        "",
    ],
    [
        "已实现",
        "STEAM-DESKTOP-ART-02",
        "美术",
        "Overlay 场景布局管线",
        "—",
        "v1.0-steam-desktop",
        "P1",
        "Unity 960×560 Canvas Prefab 导出布局 JSON；Overlay 按像素表摆图与钓位，有表则停用 MapToScene；不含 HUD",
        "docs/planning/specs/Steam桌面Overlay场景布局管线.md",
        "2026-08-16",
        "2026-08-29",
    ],
    [
        "已实现",
        "STEAM-DESKTOP-ART-03",
        "美术",
        "分塘底图、猫咪序列帧与 HUD 同步",
        "—",
        "v1.0-steam-desktop",
        "P0",
        "每塘独立底图；六姿势序列帧目录与加载；Unity Overlay HUD Prefab 导出 JSON/PNG 后 Overlay 像素一一对应",
        "docs/planning/specs/Steam桌面Overlay分塘底图与HUD同步.md",
        "2026-08-28",
        "2026-08-29",
    ],
    [
        "已实现",
        "STEAM-DESKTOP-13",
        "功能",
        "Overlay HUD 聊天栏与预制体对齐",
        "—",
        "hotfix",
        "P0",
        "聊天栏默认一行向上展开；dock_chat 相对坐标；导出修复嵌套控件尺寸",
        "docs/planning/specs/Steam桌面端-13OverlayHUD聊天栏与预制体对齐.md",
        "2026-08-29",
        "2026-08-29",
    ],
    [
        "已实现",
        "STEAM-DESKTOP-13A",
        "功能",
        "Overlay 场景四边半透明渐隐",
        "—",
        "hotfix",
        "P0",
        "仅场景层 Absolute 渐隐；四边 40px；HUD 不参与；修复底部 mask 跟包围盒走",
        "docs/planning/specs/Steam桌面端-13AOverlay场景边缘半透明渐隐.md",
        "2026-08-29",
        "2026-08-29",
    ],
    [
        "已实现",
        "STEAM-DESKTOP-13B",
        "Bug修复",
        "Overlay 宠物标签、右键与登录窗",
        "—",
        "hotfix",
        "P0",
        "状态/名字徽章；waiting→钓鱼中；右键他人社交菜单；DesktopShell scale=0 修复",
        "docs/planning/specs/Steam桌面端-13BOverlay宠物标签右键与登录窗.md",
        "2026-08-29",
        "2026-08-29",
    ],
    [
        "已实现",
        "STEAM-DESKTOP-13C",
        "功能",
        "Overlay 打窝 HUD 收口",
        "—",
        "hotfix",
        "P0",
        "按钮仅「打窝n/50」；成功文案走 txt_error；txt_error 移到钓鱼条上方",
        "docs/planning/specs/Steam桌面端-13COverlay打窝HUD收口.md",
        "2026-08-29",
        "2026-08-29",
    ],
    [
        "已实现",
        "STEAM-DESKTOP-14",
        "Bug修复",
        "Overlay 透明区点击穿透回归",
        "—",
        "hotfix",
        "P0",
        "透明像素点穿桌面；保留13A四边40px渐隐；勿用#01000000整层抢命中",
        "docs/planning/specs/Steam桌面端-14Overlay透明点击穿透回归.md",
        "2026-08-29",
        "2026-08-29",
    ],
    [
        "已实现",
        "STEAM-DESKTOP-14A",
        "功能",
        "Overlay 座位预制体与空位显隐",
        "—",
        "hotfix",
        "P0",
        "OverlayPondActor为座位真相；塘内摆实例；猫跟actor-pet；空位半透明/落座隐藏",
        "docs/planning/specs/Steam桌面端-14AOverlay钓位座位图.md",
        "2026-08-29",
        "2026-08-30",
    ],
    [
        "已实现",
        "STEAM-DESKTOP-14B",
        "功能",
        "Overlay 横轴场景平移",
        "—",
        "hotfix",
        "P1",
        "场景宽默认1920视口960；左右箭头长按平滑平移；HUD与渐隐不跟移",
        "docs/planning/specs/Steam桌面端-14BOverlay横轴场景平移.md",
        "2026-08-29",
        "2026-08-30",
    ],
    [
        "已实现",
        "STEAM-DESKTOP-14C",
        "美术",
        "Overlay HUD 正式素材与文字对齐",
        "—",
        "hotfix",
        "P0",
        "Prefab换正式图；导出字体/字号/对齐；停止强制左对齐；动态文案不烤进PNG",
        "docs/planning/specs/Steam桌面端-14COverlayHUD正式素材与文字对齐.md",
        "2026-08-29",
        "2026-08-30",
    ],
    [
        "已实现",
        "STEAM-DESKTOP-14D",
        "功能",
        "Overlay 宠物状态图标与上钩圆环",
        "—",
        "hotfix",
        "P0",
        "取消头顶相位字；小图标；圆环套在64猫身外圈；时长仍仅悬停",
        "docs/planning/specs/Steam桌面端-14DOverlay宠物状态图标与上钩圆环.md",
        "2026-08-29",
        "2026-08-30",
    ],
    [
        "已实现",
        "STEAM-DESKTOP-15",
        "功能",
        "Overlay 钓鱼栏按钮时序",
        "—",
        "hotfix",
        "P0",
        "主按钮常驻改文案；打窝后开钓须藏打窝；收杆结束后同时出打窝+离席；打窝中离席可打断",
        "docs/planning/specs/Steam桌面端-15Overlay钓鱼栏按钮时序.md",
        "2026-08-31",
        "2026-08-31",
    ],
    [
        "已实现",
        "STEAM-DESKTOP-16",
        "功能",
        "Overlay 窗口视口缩放",
        "—",
        "hotfix",
        "P0",
        "右键切960/800/600视口；高度切顶；HUD随宽内收；禁止整窗比例缩放",
        "docs/planning/specs/Steam桌面端-16Overlay窗口视口缩放.md",
        "2026-08-31",
        "2026-08-31",
    ],
    [
        "已实现",
        "STEAM-DESKTOP-17",
        "功能",
        "占用座位不可顶机器人",
        "—",
        "hotfix",
        "P0",
        "点bot/真人占用座不take_spot、不踢该座bot；塘满腾位保留",
        "docs/planning/specs/Steam桌面端-17占用座位不可顶机器人.md",
        "2026-09-01",
        "2026-09-01",
    ],
    [
        "已实现",
        "BUG-24",
        "Bug修复",
        "Overlay形象进游戏不恢复",
        "—",
        "hotfix",
        "P0",
        "冷启动用已保存avatar；不必再打开个人中心保存",
        "docs/planning/specs/BUG修复-Overlay形象进游戏不恢复.md",
        "2026-09-01",
        "2026-09-01",
    ],
    [
        "已实现",
        "BUG-25",
        "Bug修复",
        "Overlay打窝按钮剩余口数",
        "—",
        "hotfix",
        "P0",
        "按钮只打窝n/50；去掉bitesLeft后缀（-12实为剩余口数）",
        "docs/planning/specs/BUG修复-Overlay打窝按钮剩余口数.md",
        "2026-09-01",
        "2026-09-01",
    ],
    [
        "已实现",
        "BUG-26",
        "Bug修复",
        "Overlay悬停与右键菜单指针命中",
        "—",
        "hotfix",
        "P0",
        "猫身命中同一条链；菜单StaysOpen+右键抬起打开；悬停用自有开关不信IsOpen；关菜单后GetCursorPos恢复",
        "docs/planning/specs/BUG修复-Overlay悬停与右键菜单指针命中.md",
        "2026-09-01",
        "2026-09-01",
    ],
    [
        "已实现",
        "STEAM-DESKTOP-18",
        "功能",
        "Overlay悬停浮窗条数与hint定位",
        "—",
        "hotfix",
        "P0",
        "浮窗两行时长+钓到N条/空军；锚在actor-hint；命中问题归BUG-26",
        "docs/planning/specs/Steam桌面端-18Overlay悬停浮窗条数与hint定位.md",
        "2026-09-01",
        "2026-09-01",
    ],
    [
        "已实现",
        "STEAM-DESKTOP-14E",
        "功能",
        "Overlay头顶状态图标与进度环",
        "—",
        "hotfix",
        "P0",
        "等鱼无图标；仅上钩/打窝出icon；环改actor-ring+ring-bg，不再套猫",
        "docs/planning/specs/Steam桌面端-14EOverlay头顶状态图标与进度环.md",
        "2026-09-01",
        "2026-09-01",
    ],
    # —— 2026-08-10 收尾验收：ARC / DP-C / OPS ——
    [
        "已实现",
        "ARC-06",
        "架构",
        "Docker 容器化部署",
        "—",
        "v0.8",
        "P0",
        "Dockerfile+compose+SQLite卷+/health；verify:deploy（Docker CLI 环境可SKIP）",
        "docs/planning/specs/phase2-开发计划.md",
        "2026-07-11",
        "2026-08-10",
    ],
    [
        "已实现",
        "ARC-07",
        "架构",
        "Mobile JWT Token 管理",
        "—",
        "v0.8",
        "P0",
        "SecureStore+刷新+API Authorization+Socket auth.token；verify:deploy",
        "docs/planning/specs/phase2-开发计划.md",
        "2026-07-11",
        "2026-08-10",
    ],
    [
        "已实现",
        "ARC-08",
        "架构",
        "gameState.ts 拆分",
        "—",
        "v0.8",
        "P1",
        "facade+pondSession/UserManager/Chat；import boundary；verify:engineering",
        "docs/planning/specs/phase2-开发计划.md",
        "2026-07-12",
        "2026-08-10",
    ],
    [
        "已实现",
        "ARC-09",
        "架构",
        "统一日志/指标 API",
        "—",
        "v0.8",
        "P1",
        "业务路径无裸logInfo/Warn；logStructuredEvent+metrics；verify:engineering",
        "docs/planning/specs/phase2-开发计划.md",
        "2026-07-12",
        "2026-08-10",
    ],
    [
        "已实现",
        "ARC-10",
        "架构",
        "安全加固补完",
        "—",
        "v0.8",
        "P1",
        "HTTP限流+Socket连接上限+dev-token仅localhost；verify:engineering",
        "docs/planning/specs/phase2-开发计划.md",
        "2026-07-12",
        "2026-08-10",
    ],
    [
        "已实现",
        "ARC-11",
        "架构",
        "单元测试 + CI",
        "—",
        "v0.8",
        "P1",
        "Vitest+GitHub Actions；BUG-08门禁改语义对齐isFishingActive；verify:engineering",
        "docs/planning/specs/phase2-开发计划.md",
        "2026-07-12",
        "2026-08-10",
    ],
    [
        "已实现",
        "BUG-08",
        "Bug修复",
        "F1 Modal 会话计时不冻结",
        "—",
        "v0.8",
        "P2",
        "Modal打开时基于fishingStartedAt本地插值；tick与服务端isFishingActive语义对齐；verify:engineering",
        "docs/planning/specs/phase2-开发计划.md",
        "2026-07-12",
        "2026-08-10",
    ],
    [
        "已实现",
        "D-L3-02",
        "数据平台",
        "线上实测 vs 模拟对照",
        "—",
        "v0.8",
        "P1",
        "live-vs-sim.html/json+deviationPct；日批样本；verify:data-platform-dp-c",
        "docs/planning/specs/数据平台-Phase2-稳定增长.md",
        "2026-07-12",
        "2026-08-10",
    ],
    [
        "已实现",
        "D-L3-09",
        "数据平台",
        "analytics 索引打通",
        "—",
        "v0.8",
        "P1",
        "manifest live-daily+index入口+日批持续生成+warehouse/latest；verify:data-platform-dp-c",
        "docs/planning/specs/数据平台-Phase2-稳定增长.md",
        "2026-07-12",
        "2026-08-10",
    ],
    [
        "已实现",
        "OPS-RELEASE-1",
        "运维",
        "发版与热更策略",
        "—",
        "hotfix",
        "P1",
        "单机Runbook：A配置/B发版/备份/迁移/冒烟/health/回滚；OTA暂不做",
        "docs/planning/specs/发版与热更策略.md",
        "2026-07-27",
        "2026-08-10",
    ],
    [
        "已文档化",
        "STEAM-DESKTOP-EPIC",
        "产品规划",
        "Steam 桌面端独立游戏转型",
        "—",
        "v1.0-steam-desktop",
        "P0",
        "规划已完成；实际 Unity 表现层拆分为 STEAM-DESKTOP-07，已完成子需求不重复开发",
        "docs/planning/specs/Steam桌面端独立游戏转型计划.md",
        "2026-08-11",
        "2026-08-13",
    ],
    [
        "已文档化",
        "STEAM-DESKTOP-01",
        "产品规划",
        "2D 多人社交桌面宠物定位、鱼塘场景与信息架构",
        "—",
        "v1.0-steam-desktop",
        "P0",
        "规划已完成；桌面宠物、多人鱼塘、右键菜单与弹窗实现转入 STEAM-DESKTOP-07",
        "docs/planning/specs/Steam桌面端产品定位与信息架构.md",
        "2026-08-11",
        "2026-08-13",
    ],
    [
        "已实现",
        "STEAM-DESKTOP-02",
        "架构",
        "Steam 身份、账号绑定与安全会话",
        "—",
        "v1.0-steam-desktop",
        "P0",
        "Steam Ticket 验证；SteamID64↔playerId；JWT/Refresh；权威数据与审计",
        "docs/planning/specs/Steam身份账号绑定与安全会话.md",
        "2026-08-11",
        "2026-08-12",
    ],
    [
        "已实现",
        "STEAM-DESKTOP-03",
        "功能",
        "Steam 好友、Lobby、邀请与鱼塘映射",
        "—",
        "v1.0-steam-desktop",
        "P1",
        "核心功能验收完成：Lobby 创建/加入、邀请与 pondId 映射、权限拒绝、Lobby 失效后重新进塘、房主离开不删除鱼塘；双 Steam 联调因缺少第二测试账号跳过",
        "docs/planning/specs/Steam好友Lobby邀请与鱼塘映射.md",
        "2026-08-11",
        "2026-08-14",
    ],
    [
        "已实现",
        "STEAM-DESKTOP-04",
        "Unity",
        "Unity Windows 桌面端基础壳",
        "—",
        "v1.0-steam-desktop",
        "P0",
        "窗口/托盘/主界面占位/通知接口；04A～04F Windows Development Build 与冒烟已完成",
        "docs/planning/specs/Unity Windows桌面端基础壳.md",
        "2026-08-11",
        "2026-08-12",
    ],
    [
        "已实现",
        "STEAM-DESKTOP-04A",
        "Unity",
        "Unity 工程基线与 Windows 构建",
        "—",
        "v1.0-steam-desktop",
        "P0",
        "目录约定、DesktopMain、Player Settings、gitignore、构建菜单",
        "docs/planning/specs/Unity Windows桌面端基础壳.md",
        "2026-08-11",
        "2026-08-11",
    ],
    [
        "已实现",
        "STEAM-DESKTOP-04B",
        "Unity",
        "应用生命周期与窗口模式",
        "—",
        "v1.0-steam-desktop",
        "P0",
        "普通窗口/全屏/无边框、配置保存、关闭进托盘",
        "docs/planning/specs/Unity Windows桌面端基础壳.md",
        "2026-08-11",
        "2026-08-11",
    ],
    [
        "已实现",
        "STEAM-DESKTOP-04C",
        "Unity",
        "托盘与后台挂机容器",
        "—",
        "v1.0-steam-desktop",
        "P0",
        "Win32 托盘菜单；隐藏后降帧；会话生命周期占位接口",
        "docs/planning/specs/Unity Windows桌面端基础壳.md",
        "2026-08-11",
        "2026-08-11",
    ],
    [
        "已实现",
        "STEAM-DESKTOP-04D",
        "Unity",
        "Unity 主界面与功能占位",
        "—",
        "v1.0-steam-desktop",
        "P0",
        "运行时 UGUI：鱼塘/好友/鱼获/设置四入口可进出",
        "docs/planning/specs/Unity Windows桌面端基础壳.md",
        "2026-08-11",
        "2026-08-11",
    ],
    [
        "已实现",
        "STEAM-DESKTOP-04E",
        "Unity",
        "桌面设置与通知接口",
        "—",
        "v1.0-steam-desktop",
        "P1",
        "通知偏好持久化；鱼咬钩/好友邀请/连接错误模拟",
        "docs/planning/specs/Unity Windows桌面端基础壳.md",
        "2026-08-11",
        "2026-08-11",
    ],
    [
        "已实现",
        "STEAM-DESKTOP-04F",
        "Unity",
        "Unity Windows 性能与发布验证",
        "—",
        "v1.0-steam-desktop",
        "P1",
        "Windows Development Build、窗口/托盘/设置/通知/后台降帧冒烟已完成",
        "docs/planning/specs/Unity Windows桌面端基础壳.md",
        "2026-08-11",
        "2026-08-12",
    ],
    [
        "已实现",
        "STEAM-DESKTOP-05",
        "架构",
        "空鱼塘休眠与生态离线补算",
        "—",
        "v1.0-steam-desktop",
        "P1",
        "lastSimulatedAt；鱼群迁移/成长补算；幂等唤醒；活跃鱼塘才运行高频 Tick",
        "docs/planning/specs/空鱼塘休眠与生态离线补算.md",
        "2026-08-11",
        "2026-08-12",
    ],
    [
        "已实现",
        "UNITY-P1",
        "架构",
        "Unity 移植 P1·契约工程化",
        "—",
        "unity-client",
        "P0",
        "Unity/Node/shared Socket 契约、C# DTO、服务端权威边界与兼容策略；Debug：补齐 phase/重连状态同步",
        "docs/planning/specs/Unity移植-分阶段需求清单.md",
        "2026-07-26",
        "2026-08-12",
    ],
    [
        "已实现",
        "UNITY-P2",
        "架构",
        "Unity 移植 P2·网络薄客户端",
        "—",
        "unity-client",
        "P0",
        "Steam JWT Socket 登录、进塘、钓鱼、咬钩、领取、背包更新与断线重连；Debug：修复 fish_bite 时序、重复收杆和空钓位 waiting",
        "docs/planning/specs/Unity移植-分阶段需求清单.md",
        "2026-07-26",
        "2026-08-12",
    ],
]

# ENSURE 行状态优先于「总 spec」元信息同步（避免 phase2 总文档把子编号打回已确认）
ENSURE_PLAN_IDS = {str(row[1]) for row in ENSURE_PLAN_ROWS}
REMOVE_PLAN_IDS = {
    "STEAM-DESKTOP-07",
    "STEAM-DESKTOP-07A",
    "STEAM-DESKTOP-07B",
    "STEAM-DESKTOP-07C",
    "STEAM-DESKTOP-07D",
    "STEAM-DESKTOP-07E",
    "STEAM-DESKTOP-07F",
    "STEAM-DESKTOP-07G",
    "STEAM-DESKTOP-08A",
    "STEAM-DESKTOP-08B",
    "STEAM-DESKTOP-08C",
    "STEAM-DESKTOP-08D",
    "STEAM-DESKTOP-08E",
    "STEAM-DESKTOP-08F",
    "STEAM-DESKTOP-08G",
    "STEAM-DESKTOP-08H",
    "STEAM-DESKTOP-08I",
    "STEAM-UI-01",
    "STEAM-UI-02",
    "STEAM-UI-03",
    "STEAM-UI-04",
    "STEAM-UI-05",
    "STEAM-UI-06",
    "STEAM-UI-07",
    "STEAM-UI-PROG-01",
    "STEAM-UI-ART-01",
    "STEAM-UI-PROG-02",
    "STEAM-UI-ART-02",
    "STEAM-UI-PROG-03",
    "STEAM-UI-ART-03",
    "STEAM-UI-PROG-04",
    "STEAM-UI-ART-04",
    "STEAM-UI-PROG-05",
    "STEAM-UI-ART-05",
    "STEAM-UI-PROG-06",
    "STEAM-UI-ART-06",
    "STEAM-UI-PROG-07",
    "STEAM-UI-ART-07",
}
NORMALIZE_PLAN_TYPES = {
    "UNITY-P0": "架构",
    "UNITY-P1": "架构",
    "UNITY-P2": "架构",
    "UNITY-P3": "架构",
}


def is_metrics_workbook(path: Path) -> bool:
    name = path.name.lower()
    return "埋点" in path.name or "metrics" in name or "event" in name


def find_master_workbook() -> Path:
    candidates = [
        p
        for p in ROOT.glob("*.xlsx")
        if not is_metrics_workbook(p) and not p.name.startswith("~$")
    ]
    if not candidates:
        raise FileNotFoundError("No planning workbook (*.xlsx) found in repository root")
    preferred = [p for p in candidates if "计划表" in p.name or "plan" in p.name.lower()]
    return sorted(preferred or candidates, key=lambda p: p.stat().st_mtime, reverse=True)[0]


def read_spec_metadata(spec_path: Path) -> tuple[str | None, str | None]:
    if not spec_path.exists():
        return None, None
    text = spec_path.read_text(encoding="utf-8-sig")
    status_match = re.search(r"\|\s*状态\s*\|\s*\*\*(.+?)\*\*", text)
    done_match = re.search(r"\|\s*完成时间\s*\|\s*\*\*(\d{4}-\d{2}-\d{2})\*\*", text)
    return (
        status_match.group(1).strip() if status_match else None,
        done_match.group(1) if done_match else None,
    )


def find_spec_path(row: list[object]) -> Path | None:
    for value in row:
        if not isinstance(value, str):
            continue
        match = re.search(r"(docs/planning/specs/[^'\"]+\.md)", value)
        if match:
            return ROOT / match.group(1).replace("/", "\\")
    return None


def status_column(ws: openpyxl.worksheet.worksheet.Worksheet) -> int | None:
    for cell in ws[1]:
        if cell.value in {"当前状态", "当前状态"}:
            return cell.column
    # Existing workbooks may have mojibake headers.  Known planning sheets use
    # the conventional status positions below.
    return {1: 1, 2: 5, 3: 5, 4: 5, 5: 5, 6: 5, 7: 5, 8: 5, 9: 2}.get(ws._parent.worksheets.index(ws) + 1)


def ensure_plan_rows(workbook: openpyxl.Workbook) -> int:
    if not workbook.worksheets:
        return 0
    ws = workbook.worksheets[0]
    changed = 0
    id_to_row: dict[str, int] = {}
    for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if len(row) > 1 and isinstance(row[1], str) and row[1].strip():
            id_to_row[row[1].strip()] = idx

    rows_to_remove = [
        idx
        for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2)
        if len(row) > 1 and isinstance(row[1], str) and row[1].strip() in REMOVE_PLAN_IDS
    ]
    for existing in reversed(rows_to_remove):
        ws.delete_rows(existing, 1)
        changed += 1

    id_to_row = {}
    for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if len(row) > 1 and isinstance(row[1], str) and row[1].strip():
            id_to_row[row[1].strip()] = idx

    for plan_row in ENSURE_PLAN_ROWS:
        plan_id = str(plan_row[1])
        existing = id_to_row.get(plan_id)
        if existing is None:
            ws.append(plan_row)
            changed += 1
            continue
        # ENSURE 为验收权威：状态/说明/路径/日期一并回写
        for col, value in enumerate(plan_row, start=1):
            cell = ws.cell(existing, col)
            if value in (None, ""):
                continue
            if cell.value != value:
                cell.value = value
                changed += 1

    for row in ws.iter_rows(min_row=2):
        if len(row) <= 2:
            continue
        if row[2].value == "Unity":
            row[2].value = "架构"
            changed += 1
        elif row[2].value == "UI":
            row[2].value = "功能"
            changed += 1
    return changed


def update_workbook(path: Path) -> int:
    workbook = openpyxl.load_workbook(path)
    changed = 0

    for ws in workbook.worksheets:
        status_col = status_column(ws)
        if status_col is None:
            continue
        for row in ws.iter_rows(min_row=2):
            values = [cell.value for cell in row]
            plan_id = values[1] if len(values) > 1 and isinstance(values[1], str) else None
            if plan_id in NORMALIZE_PLAN_TYPES and len(row) > 2:
                type_cell = row[2]
                normalized_type = NORMALIZE_PLAN_TYPES[plan_id]
                if type_cell.value != normalized_type:
                    type_cell.value = normalized_type
                    changed += 1
            elif len(row) > 2 and row[2].value == "Unity":
                row[2].value = "架构"
                changed += 1
            elif len(row) > 2 and row[2].value == "UI":
                row[2].value = "功能"
                changed += 1
            # 子编号若在 ENSURE 中，不以总文档元信息覆盖状态
            if plan_id and plan_id in ENSURE_PLAN_IDS:
                continue
            spec_path = find_spec_path(values)
            if spec_path is None:
                continue
            status, completed = read_spec_metadata(spec_path)
            if status is None:
                continue
            cell = row[status_col - 1]
            if cell.value != status:
                cell.value = status
                changed += 1
            if status == "已实现" and completed:
                # Completion date is the cell immediately after design date in
                # all planning sheets; leave unrelated sheets untouched.
                for index, value in enumerate(values):
                    if isinstance(value, str) and re.fullmatch(r"\d{4}-\d{2}-\d{2}", value):
                        target = row[index + 1] if index + 1 < len(row) else None
                        if target is not None and target.value in (None, ""):
                            target.value = completed
                            changed += 1
                break

    # 最后应用 ENSURE，保证验收收口状态不被总 spec 打回
    changed += ensure_plan_rows(workbook)

    workbook.save(path)
    return changed


def sync_copy(master: Path) -> Path:
    DOCS_PLANNING.mkdir(parents=True, exist_ok=True)
    copy_path = DOCS_PLANNING / master.name
    shutil.copy2(master, copy_path)
    return copy_path


def regenerate_board() -> None:
    board_builder = ROOT / "scripts" / "planning" / "build-producer-progress-html.py"
    if board_builder.exists():
        subprocess.run([sys.executable, str(board_builder)], cwd=ROOT, check=True)


def main() -> None:
    master = find_master_workbook()
    changed = update_workbook(master)
    copy_path = sync_copy(master)
    regenerate_board()
    print(f"Updated: {master}")
    print(f"Synced copy: {copy_path}")
    print(f"Status/date cells changed: {changed}")


if __name__ == "__main__":
    main()
