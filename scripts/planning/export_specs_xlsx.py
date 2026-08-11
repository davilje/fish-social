from pathlib import Path
import re
import subprocess

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill


ROOT = Path(__file__).resolve().parents[2]
SPECS_DIR = ROOT / "docs" / "planning" / "specs"
OUT_PATH = SPECS_DIR / "specs_requirements_summary.xlsx"


def extract_title(lines: list[str], fallback: str) -> str:
    for line in lines:
        m = re.match(r"^#\s+(.+?)\s*$", line.strip())
        if m:
            return m.group(1).strip()
    return fallback


def first_match(text: str, patterns: list[str]) -> str:
    for pattern in patterns:
        m = re.search(pattern, text, flags=re.MULTILINE)
        if m:
            return m.group(1).strip()
    return ""


def extract_brief(text: str, lines: list[str]) -> str:
    bg = first_match(text, [r"\*\*背景\*\*\s*[:：]\s*([^\n]+)"])
    if bg:
        return normalize_brief(bg)

    for line in lines:
        s = line.strip()
        if not s:
            continue
        if s.startswith(("#", "|", "```", ">", "---", "***")):
            continue
        clean = re.sub(r"\*\*|`", "", s).strip()
        if clean:
            return normalize_brief(clean)
    return ""


def normalize_brief(s: str) -> str:
    s = re.sub(r"\s+", " ", s).strip()
    if len(s) > 180:
        s = s[:177] + "..."
    return s


def main() -> None:
    files = sorted(
        p for p in SPECS_DIR.glob("*.md") if p.name.lower() != "readme.md"
    )

    enriched: list[dict[str, str | int]] = []
    for path in files:
        text = path.read_text(encoding="utf-8", errors="ignore")
        lines = text.splitlines()
        title = extract_title(lines, path.stem)
        version = first_match(
            text,
            [
                r"\*\*版本\*\*\s*[:：]\s*([^\n]+)",
                r"\|\s*目标版本\s*\|\s*([^|\n]+)\|",
            ],
        )
        status = first_match(
            text,
            [
                r"\*\*状态\*\*\s*[:：]\s*([^\n]+)",
                r"\|\s*Spec 状态\s*\|\s*([^|\n]+)\|",
                r"\|\s*状态\s*\|\s*([^|\n]+)\|",
            ],
        )
        brief = extract_brief(text, lines)

        created_ts = get_first_commit_ts(path)
        enriched.append(
            {
                "path": str(path.relative_to(ROOT)).replace("\\", "/"),
                "orig_version": version,
                "status": status,
                "title": title,
                "brief": brief,
                "created_ts": created_ts,
            }
        )

    # 按真实开发顺序（文件首次提交时间）排序；相同时间按路径稳定排序
    enriched.sort(key=lambda x: (int(x["created_ts"]), str(x["path"])))

    wb = Workbook()
    ws = wb.active
    ws.title = "specs开发顺序汇总"

    headers = [
        "开发顺序",
        "重排版本号",
        "文档路径",
        "原版本",
        "开发状态",
        "需求名称",
        "需求内容简要描述",
    ]
    ws.append(headers)
    for i, item in enumerate(enriched, start=1):
        new_version = f"0.{i:02d}"
        ws.append(
            [
                i,
                new_version,
                item["path"],
                item["orig_version"],
                item["status"],
                item["title"],
                item["brief"],
            ]
        )

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for c in ws[1]:
        c.fill = header_fill
        c.font = header_font
        c.alignment = Alignment(horizontal="center", vertical="center")

    widths = {"A": 10, "B": 12, "C": 56, "D": 16, "E": 20, "F": 40, "G": 95}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=7):
        for c in row:
            c.alignment = Alignment(vertical="top", wrap_text=True)

    ws.freeze_panes = "A2"
    wb.save(OUT_PATH)
    print(f"Wrote {OUT_PATH}")
    print(f"Rows: {len(enriched)}")


def get_first_commit_ts(path: Path) -> int:
    rel = str(path.relative_to(ROOT)).replace("\\", "/")
    try:
        cmd = [
            "git",
            "log",
            "--follow",
            "--diff-filter=A",
            "--format=%ct",
            "--",
            rel,
        ]
        out = subprocess.check_output(cmd, cwd=ROOT, text=True, stderr=subprocess.DEVNULL)
        lines = [ln.strip() for ln in out.splitlines() if ln.strip()]
        if lines:
            return int(lines[-1])
    except Exception:
        pass
    return int(path.stat().st_mtime)


if __name__ == "__main__":
    main()
