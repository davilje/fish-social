from __future__ import annotations

from pathlib import Path
import re
import subprocess

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill


ROOT = Path(__file__).resolve().parents[2]
SPECS_DIR = ROOT / "docs" / "planning" / "specs"
CHANGELOG_PATH = ROOT / "docs" / "planning" / "CHANGELOG.md"
OUT_PATH = SPECS_DIR / "项目开发需求计划表.xlsx"

# 以 CHANGELOG 为准覆盖 spec 元信息中的滞后状态
CHANGELOG_OVERRIDES: dict[str, str] = {
    "A0-数值重构.md": "已废弃",
    "A1-飘字广播.md": "已实现",
    "A2-Debug面板.md": "已实现",
    "B0-商店基础.md": "已实现",
    "B1-鱼饵偏好.md": "已实现",
    "UI体验修复-社交商店图鉴.md": "已实现",
    "BUG修复-四项体验问题.md": "已实现",
    "BUG修复-资料与鱼塘UI.md": "已实现",
    "BUG修复-鱼塘钓鱼时长显示.md": "已实现",
    "分析与修复-钓鱼概率与饵文案.md": "已实现",
    "BUG修复-挂机断线离位.md": "已实现",
    "排查-挂机断线诊断阶段2-4.md": "已实现",
    "数值重构-品质尺寸咬钩脱钩.md": "已实现",
    "生态调优-咬钩间隔与鱼群恢复.md": "已实现",
    "钓点鱼群流动性与分区咬钩.md": "已实现",
    "咬钩产量调优-单鱼抽样与脱钩.md": "已实现",
    "服务器架构缺陷与埋点设计-v0.4.4.md": "已实现",
    "v0.4.4-未完成项补完.md": "已实现",
    "v0.4.4-埋点缺口复核与补全.md": "已实现",
    "服务器架构优化路线图-v0.5.md": "已实现",
    "服务器维护-端口占用.md": "已实现",
    "他人主页优化.md": "已实现",
    "数值重构v2-成长咬钩与文案.md": "已实现",
    "钓鱼系统v2-生态与玩法重构.md": "已废弃",
    "v0.2.4-开发交接.md": "已实现",
    "v0.2.5-开发交接.md": "已实现",
    "v0.2.6-开发交接.md": "已实现",
    "v0.3.0-开发交接.md": "已实现",
    "v0.3.1-开发交接.md": "已实现",
    "v0.3.2-开发交接.md": "已实现",
    "v0.4.0-开发交接.md": "已实现",
    "v0.4.1-开发交接.md": "已实现",
    "钓鱼系统v2-开发交接.md": "已实现",
    "BUG修复-切页误离塘与计时中断.md": "已实现",
    "BUG修复-会话计时广播回归.md": "已实现",
    "BUG修复-tsx-watch启动挂死.md": "已实现",
    "状态机需求描述.md": "已定稿",
}

# 未实现 / 部分实现项的人工分析（按文档路径）
GAP_ANALYSIS: dict[str, dict[str, str]] = {
    "C-调优与状态机.md": {
        "gap": "C1~C7 子任务均未完整落地：热更双人审批、Bot 深度适配、金币 Sink 二期、繁衍遗传、图鉴增强、完整状态机动画、灰度周报。",
        "recommendation": "继续开发（P2 按需）",
        "priority": "P2",
        "note": "C3 渔具耐久已有雏形（C3_SINK_ENABLED）；C6 服务端 phase 已较完整，客户端动画仍简陋。",
    },
    "状态机需求描述.md": {
        "gap": "完整状态机与阶段动画的设计基线；非独立交付物。",
        "recommendation": "无需单独开发",
        "priority": "—",
        "note": "供 C6 引用；服务端 phase 转移、断线 60s 续接已基本具备。",
    },
    "数值重构v2-成长咬钩与文案.md": {
        "gap": "§6.3 Debug 方格面板在 v0.2.4 已实现；后续 v0.3.x 数值已多次迭代，本文档部分数值已 superseded。",
        "recommendation": "文档同步即可",
        "priority": "—",
        "note": "AdminPondFishDebugGrid 已落地；无需按 v2 原文重做。",
    },
    "钓鱼系统v2-生态与玩法重构.md": {
        "gap": "主 spec 已拆分为 A0~C 子文档，本文档仅作历史参考。",
        "recommendation": "无需继续开发",
        "priority": "—",
        "note": "保留归档；开发以子 spec + CHANGELOG 为准。",
    },
    "A0-数值重构.md": {
        "gap": "初版指数咬钩模型，已被 v0.3.x 品质×尺寸公式替代。",
        "recommendation": "无需继续开发",
        "priority": "—",
        "note": "标记 superseded，仅供追溯。",
    },
    "他人主页优化.md": {
        "gap": "简介、收藏品、动态列表已在 UserProfileModal 实现。",
        "recommendation": "文档同步即可",
        "priority": "—",
        "note": "spec 状态仍写「已确认」；建议改为「已实现」。",
    },
    "咬钩产量调优-单鱼抽样与脱钩.md": {
        "gap": "D10~D12 已在 v0.4.1 实现。",
        "recommendation": "文档同步即可",
        "priority": "—",
        "note": "spec 仍写「策划已确认」；建议改为「已实现」。",
    },
    "钓点鱼群流动性与分区咬钩.md": {
        "gap": "D9 已在 v0.4.0 实现。",
        "recommendation": "文档同步即可",
        "priority": "—",
        "note": "spec 仍写「策划已确认」；建议改为「已实现」。",
    },
    "钓鱼系统v2-开发交接.md": {
        "gap": "A0~B1 及后续生态/架构版本均已交付；交接文档本身非功能交付物。",
        "recommendation": "文档同步即可",
        "priority": "—",
        "note": "建议状态改为「已实现（交接完成）」。",
    },
    "v0.2.4-开发交接.md": {
        "gap": "Debug 方格面板任务已在 CHANGELOG 0.2.4 记录为已实现。",
        "recommendation": "文档同步即可",
        "priority": "—",
        "note": "交接文档无独立状态字段。",
    },
    "v0.2.5-开发交接.md": {
        "gap": "钓鱼时长显示修复已在 CHANGELOG 0.2.5 记录为已实现。",
        "recommendation": "文档同步即可",
        "priority": "—",
        "note": "README 索引仍误标「待开发」。",
    },
    "v0.2.6-开发交接.md": {
        "gap": "咬钩 30s 与饵文案修复已在 CHANGELOG 0.2.6 记录为已实现（后续 v0.3.1 又改为 300s）。",
        "recommendation": "文档同步即可",
        "priority": "—",
        "note": "README 索引仍误标「待开发」。",
    },
    "服务器架构优化路线图-v0.5.md": {
        "gap": "三期 14 项（R0~R2）均已实现；R2-3 单实例容量仅文档化。",
        "recommendation": "文档同步即可",
        "priority": "—",
        "note": "表头仍写「待开发」与正文矛盾；多实例/Redis 待规模化时再开。",
    },
    "生态调优-咬钩间隔与鱼群恢复.md": {
        "gap": "D1~D8 均已实现；spec 头部仍写 D7~D8 待开发。",
        "recommendation": "文档同步即可",
        "priority": "—",
        "note": "以 CHANGELOG 0.3.1 / 0.3.2 为准。",
    },
}

# v0.5 遗留但无独立 spec 的缺口
EXTRA_GAPS: list[dict[str, str]] = [
    {
        "name": "移动端 JWT 鉴权对接",
        "path": "（无独立 spec，见 v0.5 R0-1 备注）",
        "effective_status": "未实现",
        "gap": "服务端 JWT / Socket auth.token 已就绪，mobile 客户端尚未携带 token 登录。",
        "recommendation": "继续开发（上线前必须）",
        "priority": "P0",
        "note": "开发期可用 AUTH_DISABLED=1；生产环境需 mobile 发 token。",
    },
    {
        "name": "多实例 / Redis 水平扩展",
        "path": "服务器架构优化路线图-v0.5.md §R2-3",
        "effective_status": "规划中",
        "gap": "单实例 checkpoint 已具备，但 Socket.IO Redis Adapter 与分布式锁未做。",
        "recommendation": "暂缓（规模化触发）",
        "priority": "P2",
        "note": "当前推荐单实例 <200 活跃连接；压测达标后再立项。",
    },
]


def git_first_commit_ts(path: Path) -> int:
    rel = str(path.relative_to(ROOT)).replace("\\", "/")
    try:
        out = subprocess.check_output(
            ["git", "log", "--follow", "--diff-filter=A", "--format=%ct", "--", rel],
            cwd=ROOT,
            text=True,
            stderr=subprocess.DEVNULL,
        )
        rows = [x.strip() for x in out.splitlines() if x.strip()]
        if rows:
            return int(rows[-1])
    except Exception:
        pass
    return int(path.stat().st_mtime)


def extract_title(lines: list[str], fallback: str) -> str:
    for line in lines:
        m = re.match(r"^#\s+(.+?)\s*$", line.strip())
        if m:
            return m.group(1).strip()
    return fallback


def first_match(text: str, patterns: list[str]) -> str:
    for p in patterns:
        m = re.search(p, text, flags=re.MULTILINE)
        if m:
            return m.group(1).strip()
    return ""


def normalize_status(raw: str) -> str:
    s = raw.strip().lower()
    if not s:
        return "未标注"
    if "已实现" in s:
        return "已实现"
    if "待开发" in s:
        return "待开发"
    if "已确认" in s or "已定稿" in s:
        return "已确认"
    if "待评审" in s or "评审中" in s:
        return "评审中"
    if "superseded" in s or "废弃" in s:
        return "已废弃"
    if "草案" in s or "提案" in s:
        return "规划中"
    return "未标注"


def effective_status(filename: str, spec_status: str) -> str:
    if filename in CHANGELOG_OVERRIDES:
        return CHANGELOG_OVERRIDES[filename]
    return spec_status


def optimize_name(title: str) -> str:
    t = title.strip()
    t = re.sub(r"^v\d+(?:\.\d+){1,2}\s*[-—_：:]?\s*", "", t, flags=re.IGNORECASE)
    t = re.sub(r"\s*[-—_]\s*v\d+(?:\.\d+){1,2}\s*$", "", t, flags=re.IGNORECASE)
    t = re.sub(r"\s*\(.*?superseded.*?\)\s*", "", t, flags=re.IGNORECASE)
    t = t.replace("BUG修复-", "BUG修复：")
    t = t.replace("排查-", "排查：")
    t = t.replace("分析与修复-", "分析与修复：")
    t = re.sub(r"\s{2,}", " ", t).strip(" -_：:")
    return t or title


def classify_type(name: str, path_name: str) -> str:
    key = f"{name} {path_name}".lower()
    if "开发交接" in key:
        return "开发交接"
    if "bug修复" in key or "修复" in key:
        return "缺陷修复"
    if "排查" in key or "诊断" in key:
        return "故障排查"
    if "架构" in key or "埋点" in key:
        return "架构治理"
    if "路线图" in key:
        return "路线规划"
    if "数值" in key or "调优" in key or "咬钩" in key:
        return "数值调优"
    if "状态机" in key:
        return "系统设计"
    if "商店" in key or "图鉴" in key or "ui" in key or "体验" in key or "主页" in key:
        return "功能优化"
    return "功能需求"


def extract_brief(text: str, lines: list[str]) -> str:
    bg = first_match(text, [r"\*\*背景\*\*\s*[:：]\s*([^\n]+)"])
    if bg:
        s = bg
    else:
        s = ""
        for line in lines:
            c = line.strip()
            if not c:
                continue
            if c.startswith(("#", "|", "```", ">", "---", "***")):
                continue
            s = re.sub(r"\*\*|`", "", c).strip()
            if s:
                break
    s = re.sub(r"\s+", " ", s).strip()
    if len(s) > 180:
        s = s[:177] + "..."
    return s


def style_header_row(ws, headers: list[str], widths: dict[str, float]) -> None:
    fill = PatternFill("solid", fgColor="1F4E78")
    font = Font(color="FFFFFF", bold=True)
    for c in ws[1]:
        c.fill = fill
        c.font = font
        c.alignment = Alignment(horizontal="center", vertical="center")
    for col, w in widths.items():
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A2"


def wrap_rows(ws, max_col: int) -> None:
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=max_col):
        for c in row:
            c.alignment = Alignment(vertical="top", wrap_text=True)


def build_doc_lag_rows(items: list[dict]) -> list[dict]:
    rows: list[dict] = []
    for item in items:
        if item["spec_status"] == item["effective_status"]:
            continue
        filename = Path(item["path"]).name
        gap = GAP_ANALYSIS.get(filename, {})
        rows.append(
            {
                "name": item["name"],
                "path": item["path"],
                "spec_status": item["spec_status"],
                "effective_status": item["effective_status"],
                "recommendation": gap.get("recommendation", "文档同步即可"),
                "note": gap.get("note", "以 CHANGELOG 与代码为准，更新 spec/README 状态列。"),
            }
        )
    return rows


def build_gap_rows(items: list[dict]) -> list[dict]:
    rows: list[dict] = []
    for item in items:
        filename = Path(item["path"]).name
        eff = item["effective_status"]
        if eff == "已实现":
            continue
        gap = GAP_ANALYSIS.get(filename)
        if gap:
            rows.append(
                {
                    "name": item["name"],
                    "path": item["path"],
                    "effective_status": eff,
                    "gap": gap["gap"],
                    "recommendation": gap["recommendation"],
                    "priority": gap["priority"],
                    "note": gap["note"],
                }
            )
        elif eff in ("已确认", "待开发", "未标注", "评审中", "规划中"):
            rows.append(
                {
                    "name": item["name"],
                    "path": item["path"],
                    "effective_status": eff,
                    "gap": item["brief"] or "spec 标记未完成，需对照 CHANGELOG 与代码复核。",
                    "recommendation": "需人工复核",
                    "priority": "—",
                    "note": f"文档状态={item['spec_status']}；有效状态={eff}",
                }
            )
    rows.extend(EXTRA_GAPS)
    return rows


def main() -> None:
    files = sorted(p for p in SPECS_DIR.glob("*.md") if p.name.lower() != "readme.md")

    items: list[dict] = []
    for p in files:
        text = p.read_text(encoding="utf-8", errors="ignore")
        lines = text.splitlines()
        title = extract_title(lines, p.stem)
        version = first_match(
            text,
            [r"\*\*版本\*\*\s*[:：]\s*([^\n]+)", r"\|\s*目标版本\s*\|\s*([^|\n]+)\|"],
        )
        raw_status = first_match(
            text,
            [
                r"\*\*状态\*\*\s*[:：]\s*([^\n]+)",
                r"\|\s*Spec 状态\s*\|\s*([^|\n]+)\|",
                r"\|\s*状态\s*\|\s*([^|\n]+)\|",
            ],
        )
        spec_status = normalize_status(raw_status)
        eff = effective_status(p.name, spec_status)
        rel_path = str(p.relative_to(ROOT)).replace("\\", "/")
        items.append(
            {
                "ts": git_first_commit_ts(p),
                "path": rel_path,
                "filename": p.name,
                "version": version,
                "spec_status": spec_status,
                "effective_status": eff,
                "name": optimize_name(title),
                "type": classify_type(title, p.name),
                "brief": extract_brief(text, lines),
            }
        )

    items.sort(key=lambda x: (int(x["ts"]), str(x["path"])))

    wb = Workbook()

    # Sheet 1 — 全量计划
    ws_plan = wb.active
    ws_plan.title = "开发需求计划"
    plan_headers = [
        "开发顺序",
        "版本号",
        "需求名称",
        "需求类型",
        "文档状态",
        "有效状态",
        "需求内容简要描述",
        "文档路径",
    ]
    ws_plan.append(plan_headers)
    for i, item in enumerate(items, start=1):
        ws_plan.append(
            [
                i,
                f"0.{i:02d}",
                item["name"],
                item["type"],
                item["spec_status"],
                item["effective_status"],
                item["brief"],
                item["path"],
            ]
        )
    style_header_row(
        ws_plan,
        plan_headers,
        {"A": 10, "B": 10, "C": 40, "D": 14, "E": 12, "F": 12, "G": 70, "H": 52},
    )
    wrap_rows(ws_plan, 8)

    # Sheet 2 — 未实现分析
    gap_rows = build_gap_rows(items)
    ws_gap = wb.create_sheet("未实现分析")
    gap_headers = [
        "序号",
        "需求名称",
        "有效状态",
        "缺口描述",
        "是否继续开发",
        "优先级",
        "备注",
        "文档路径",
    ]
    ws_gap.append(gap_headers)
    for i, row in enumerate(gap_rows, start=1):
        ws_gap.append(
            [
                i,
                row["name"],
                row["effective_status"],
                row["gap"],
                row["recommendation"],
                row["priority"],
                row["note"],
                row["path"],
            ]
        )
    style_header_row(
        ws_gap,
        gap_headers,
        {"A": 8, "B": 36, "C": 12, "D": 58, "E": 18, "F": 10, "G": 42, "H": 52},
    )
    wrap_rows(ws_gap, 8)

    # Sheet 3 — 文档滞后（已实现但 spec 状态未同步）
    lag_rows = build_doc_lag_rows(items)
    ws_lag = wb.create_sheet("文档滞后")
    lag_headers = [
        "序号",
        "需求名称",
        "文档状态",
        "有效状态",
        "处理建议",
        "备注",
        "文档路径",
    ]
    ws_lag.append(lag_headers)
    for i, row in enumerate(lag_rows, start=1):
        ws_lag.append(
            [
                i,
                row["name"],
                row["spec_status"],
                row["effective_status"],
                row["recommendation"],
                row["note"],
                row["path"],
            ]
        )
    style_header_row(
        ws_lag,
        lag_headers,
        {"A": 8, "B": 36, "C": 12, "D": 12, "E": 16, "F": 48, "G": 52},
    )
    wrap_rows(ws_lag, 7)

    # Sheet 4 — 统计摘要
    ws_sum = wb.create_sheet("统计摘要")
    total = len(items)
    eff_counts: dict[str, int] = {}
    for item in items:
        eff_counts[item["effective_status"]] = eff_counts.get(item["effective_status"], 0) + 1

    continue_dev = sum(1 for r in gap_rows if "继续开发" in r["recommendation"])
    doc_sync = sum(1 for r in gap_rows if "文档同步" in r["recommendation"]) + len(lag_rows)
    no_need = sum(1 for r in gap_rows if "无需" in r["recommendation"])
    defer = sum(1 for r in gap_rows if "暂缓" in r["recommendation"])

    ws_sum.append(["指标", "数值", "说明"])
    ws_sum.append(["需求文档总数", total, "不含 README.md"])
    for status in ["已实现", "已确认", "已定稿", "待开发", "已废弃", "未标注", "规划中", "评审中"]:
        if status in eff_counts:
            ws_sum.append([f"有效状态={status}", eff_counts[status], ""])
    ws_sum.append(["未实现/缺口条目", len(gap_rows), "见「未实现分析」sheet"])
    ws_sum.append(["建议继续开发", continue_dev, "含 P2 按需与上线前 P0"])
    ws_sum.append(["文档滞后条目", len(lag_rows), "见「文档滞后」sheet"])
    ws_sum.append(["仅需文档同步", doc_sync, "代码已完成，spec/README 滞后"])
    ws_sum.append(["无需继续开发", no_need, "已废弃或设计参考文档"])
    ws_sum.append(["暂缓/规模化触发", defer, "多实例等"])
    ws_sum.append(
        [
            "结论",
            "",
            "核心玩法与架构 v0.5 已闭环；剩余主要为 C 期增强、文档同步、mobile 鉴权与规模化前置。",
        ]
    )
    style_header_row(ws_sum, ["指标", "数值", "说明"], {"A": 22, "B": 12, "C": 70})
    wrap_rows(ws_sum, 3)

    wb.save(OUT_PATH)

    implemented = eff_counts.get("已实现", 0)
    print(f"Wrote {OUT_PATH}")
    print(f"Rows: {total} | Implemented: {implemented} | Gap analysis rows: {len(gap_rows)} | Doc lag rows: {len(lag_rows)}")


if __name__ == "__main__":
    main()
