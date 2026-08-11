#!/usr/bin/env python3
"""从 CSV 导出按类别分 sheet 的埋点表 xlsx。

格式标准（以 docs/planning/reports/v0.4.4-埋点表清单.xlsx 手工版为准）：
- Sheet：总览 + 连接与会话 / 钓位与鱼塘 / 状态机 / 咬钩与产出 / 性能 / 慢测验证
- 列顺序：类别, 事件名, 类型, 触发时机, 必填字段(核心), 示例payload JSON, 优先级, 状态, 推荐落点文件
- 必填字段(核心)：每个字段单独一行（逗号后换行），末尾保留两个空行
- 单元格：wrap_text=True，按行数自动行高
- 更新流程：优先直接编辑 xlsx；若从 CSV 重导，导出后需核对列顺序与换行格式
"""
import csv
import subprocess
import sys
from collections import defaultdict
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment
    from openpyxl.utils import get_column_letter
except ImportError:
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl", "-q"])
    from openpyxl import Workbook
    from openpyxl.styles import Alignment
    from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[1]
CSV_PATH = ROOT / "docs/planning/reports/v0.4.4-埋点表清单.csv"
XLSX_PATH = ROOT / "docs/planning/reports/v0.4.4-埋点表清单.xlsx"

# sheet 顺序与 CSV「类别」列对应
SHEET_ORDER = [
    ("总览", None),
    ("连接与会话", "连接与会话"),
    ("钓位与鱼塘", "钓位与鱼塘"),
    ("状态机", "状态机"),
    ("咬钩与产出", "咬钩与产出"),
    ("性能", "性能"),
    ("慢测验证", "慢测验证"),
]

COL_WIDTHS = [14, 34, 16, 8, 18, 30, 52, 44, 90]


def read_rows() -> tuple[list[str], list[list[str]]]:
    with CSV_PATH.open(encoding="utf-8-sig", newline="") as f:
        reader = csv.reader(f)
        header = next(reader)
        rows = [row for row in reader if row]
    return header, rows


def write_sheet(ws, header: list[str], rows: list[list[str]]) -> None:
    ws.append(header)
    for row in rows:
        ws.append(row)
    ws.freeze_panes = "A2"
    for i, width in enumerate(COL_WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(i)].width = min(width, 100)
    for row in ws.iter_rows():
        max_lines = 1
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            if cell.value is not None:
                max_lines = max(max_lines, str(cell.value).count("\n") + 1)
        ws.row_dimensions[row[0].row].height = max(20, 15 * max_lines)


def main() -> None:
    header, all_rows = read_rows()
    by_category: dict[str, list[list[str]]] = defaultdict(list)
    for row in all_rows:
        if row:
            by_category[row[0]].append(row)

    wb = Workbook()
    wb.remove(wb.active)

    for sheet_name, category in SHEET_ORDER:
        ws = wb.create_sheet(title=sheet_name)
        if category is None:
            write_sheet(ws, header, all_rows)
        else:
            write_sheet(ws, header, by_category.get(category, []))

    wb.save(XLSX_PATH)
    print(f"Wrote {XLSX_PATH} ({len(all_rows)} events, {len(SHEET_ORDER)} sheets)")


if __name__ == "__main__":
    main()
