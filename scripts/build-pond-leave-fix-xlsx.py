#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""生成 切页误离塘修复方案-v0.5.1.xlsx"""
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

ROOT = Path(__file__).resolve().parents[1]
OUT = ROOT / "docs/planning/reports/切页误离塘修复方案-v0.5.1.xlsx"


def style_sheet(ws) -> None:
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 120
    for c in ws[1]:
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="1F4E78")
        c.alignment = Alignment(horizontal="center", vertical="center")
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=2):
        row[0].font = Font(bold=True)
        row[0].fill = PatternFill("solid", fgColor="D9E1F2")
        row[0].alignment = Alignment(vertical="top", wrap_text=True)
        row[1].alignment = Alignment(vertical="top", wrap_text=True)


def add_sheet(wb, name: str, rows: list[tuple[str, str]]) -> None:
    ws = wb.create_sheet(name)
    ws.append(["字段", "内容"])
    for r in rows:
        ws.append(list(r))
    style_sheet(ws)


OVERVIEW = [
    ["编号", "需求", "优先级", "影响", "状态", "对应sheet"],
    ["L1", "移除社交/资料跳转前 leave_pond", "P0", "计时中断/误离塘", "已实现", "01-L1-社交资料不离塘"],
    ["L2", "移除 unmount 默认 leave_pond", "P0", "Debug等路由误离塘", "已实现", "02-L2-卸载不离塘"],
    ["L3", "仅返回地图显式 leave_pond", "P0", "离塘语义", "已实现", "03-L3-显式离塘"],
    ["L4", "切页断 Socket 走 disconnect 宽限", "P0", "重连恢复", "已实现", "04-L4-断线宽限"],
    ["L5", "回塘 reconnect 与计时连续", "P0", "体验验收", "已实现", "05-L5-重连续计时"],
    ["L6", "修订诊断文档与 verify", "P1", "回归/策划一致", "已实现", "06-L6-文档验收"],
    ["F1", "Modal 计时仍停排查", "P2", "背包/商店/图鉴", "待开发", "07-F1-Modal排查"],
]

SHEETS: dict[str, list[tuple[str, str]]] = {
    "01-L1-社交资料不离塘": [
        ("需求编号", "L1"),
        ("标题", "移除社交 / 资料跳转前的 leave_pond"),
        ("优先级", "P0"),
        ("现状", "pond/[id].tsx：handleLeaveToSocial、onEditProfile 在 router.push 前调用 leavePondWithReason(navigation_social|navigation_profile)。"),
        ("问题", "用户仅切换页面，服务端却清塘内会话、释放钓位，sessionFishingMs 停止更新。"),
        ("修复方案", "1) handleLeaveToSocial 仅 router.push('/social')，不 leave\n2) onEditProfile 仅 router.push('/profile')，不 leave\n3) 保留 auth_redirect 离塘逻辑"),
        ("涉及文件", "mobile/app/pond/[id].tsx"),
        ("验收标准", "钓鱼中点社交/资料：服务端日志无 leave_pond；可有 socket disconnect。"),
        ("状态", "已实现"),
    ],
    "02-L2-卸载不离塘": [
        ("需求编号", "L2"),
        ("标题", "移除 usePondSocket unmount 默认 leave_pond"),
        ("优先级", "P0"),
        ("现状", "usePondSocket cleanup：若未 leftPondRef，则 emitLeavePond('unmount') 再 socket.disconnect()。"),
        ("问题", "进入 /admin、/social 等导致鱼塘页卸载，必然 leave_pond(unmount)。"),
        ("修复方案", "1) 删除 cleanup 内 emitLeavePond('unmount')\n2) cleanup 仅 socket.disconnect()\n3) leftPondRef 仅由显式 leavePondWithReason 置位"),
        ("涉及文件", "mobile/lib/usePondSocket.ts"),
        ("验收标准", "路由离开鱼塘页：无 leave_pond 事件；metrics/timeline 无 unmount reason。"),
        ("状态", "已实现"),
    ],
    "03-L3-显式离塘": [
        ("需求编号", "L3"),
        ("标题", "仅「返回地图」等明确动作为显式离塘"),
        ("优先级", "P0"),
        ("产品规则", "发送 leave_pond 的场景：navigation_back（←地图）、auth_redirect；禁止：navigation_social、navigation_profile、unmount。"),
        ("修复方案", "1) handleLeaveToMap 保留 leavePondWithReason('navigation_back')\n2) leavePondWithReason 文档注释：仅显式离塘入口调用\n3) 可选重命名 API 为 leavePondExplicit 提高可读性"),
        ("涉及文件", "mobile/app/pond/[id].tsx · mobile/lib/usePondSocket.ts"),
        ("验收标准", "仅返回地图产生 leave_pond(navigation_back)；其余导航无 leave。"),
        ("状态", "已实现"),
    ],
    "04-L4-断线宽限": [
        ("需求编号", "L4"),
        ("标题", "切页导致 Socket 断开时依赖服务端 disconnect 宽限"),
        ("优先级", "P0"),
        ("现状", "页面卸载会 socket.disconnect()；若已去掉 leave，服务端应走 handleDisconnect 而非 leavePond。"),
        ("修复方案", "1) 客户端：卸载只 disconnect，不 leave\n2) 服务端：验证 disconnect 后 60s 内 spot/phase 保留（已有）\n3) 日志区分 disconnect vs leave_pond"),
        ("涉及文件", "mobile/lib/usePondSocket.ts · server fishingStateMachine（验证）"),
        ("验收标准", "切页后服务端有 disconnect 日志、无 leave_pond；60s 内 user 仍在 disconnected 态。"),
        ("状态", "已实现"),
    ],
    "05-L5-重连续计时": [
        ("需求编号", "L5"),
        ("标题", "回塘 reconnect 与会话计时连续"),
        ("优先级", "P0"),
        ("现状", "leave 后再 join 为 joinKind=fresh、新 userId，计时归零。"),
        ("修复方案", "1) L1–L4 完成后，60s 内回塘应 joinKind=reconnect\n2) sessionFishingMs / spotId / fishingPhase 与离开前一致或合理续接\n3) 结合 player_pond_session checkpoint（v0.5）"),
        ("验收标准", "T3/T4：join_pond_success joinKind=reconnect；头顶计时无明显归零（除非宽限过期）。"),
        ("测试", "手测 + verify:disconnect-reconnect + verify:session-checkpoint"),
        ("状态", "已实现"),
    ],
    "06-L6-文档验收": [
        ("需求编号", "L6"),
        ("标题", "修订策划文档与自动化验收"),
        ("优先级", "P1"),
        ("修复方案", "1) 更新 排查-挂机断线诊断阶段2-4.md §阶段3：社交不再预期 leave\n2) 更新 verify-afk-diag.ts\n3) 可选新增 verify-pond-navigation.ts：钓鱼→社交→无 leave→回塘 reconnect"),
        ("涉及文件", "docs/planning/specs/*.md · scripts/verify-afk-diag.ts"),
        ("验收标准", "文档与代码一致；verify 脚本通过。"),
        ("状态", "已实现"),
    ],
    "07-F1-Modal排查": [
        ("需求编号", "F1"),
        ("标题", "背包/商店/图鉴 Modal 计时仍停（跟进）"),
        ("优先级", "P2"),
        ("说明", "三者为同页 Modal，理论上不应 leave 或断 Socket。若 L1–L5 后仍反馈计时停，排查 pond_user_updated 广播与 PondCharacter 重渲染。"),
        ("排查步骤", "1) 打开 Modal 时抓服务端是否仍广播 sessionFishingMs\n2) 查 React 是否卸载 PondScene\n3) 查 demoMode 误触发"),
        ("状态", "待开发"),
    ],
}


def main() -> None:
    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws0 = wb.active
    ws0.title = "总览"
    for row in OVERVIEW:
        ws0.append(row)
    for col, w in {"A": 8, "B": 36, "C": 8, "D": 18, "E": 10, "F": 22}.items():
        ws0.column_dimensions[col].width = w
    for c in ws0[1]:
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="1F4E78")
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for row in ws0.iter_rows(min_row=2, max_row=ws0.max_row):
        for c in row:
            c.alignment = Alignment(vertical="top", wrap_text=True)

    for name, content in SHEETS.items():
        add_sheet(wb, name, content)

    wb.save(OUT)
    print(f"Wrote {OUT} ({len(SHEETS) + 1} sheets)")


if __name__ == "__main__":
    main()
