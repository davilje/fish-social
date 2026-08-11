#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
生成「三层数据体系 · 开发需求清单」xlsx
输出：docs/planning/reports/三层数据体系-开发需求清单.xlsx

【权威脚本】本文件为单一来源。旧版 scripts/build-data-platform-requirements-xlsx.py 已弃用。
npm run planning:data-platform-xlsx

关联既有文档（同目录）：
  - v0.4.4-埋点表清单.xlsx
  - 服务器架构问题与修复方案-v0.5.xlsx
"""
from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

ROOT = Path(__file__).resolve().parents[1]
REPORTS = ROOT / "docs/planning/reports"
OUT = REPORTS / "三层数据体系-开发需求清单.xlsx"

LAYER_FILL = {
    "L1": PatternFill("solid", fgColor="D9E1F2"),
    "L2": PatternFill("solid", fgColor="E2EFDA"),
    "L3": PatternFill("solid", fgColor="FCE4D6"),
}

STATUS_DONE = {"已实现", "已文档化"}


def style_kv_sheet(ws) -> None:
    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 118
    for c in ws[1]:
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="1F4E78")
        c.alignment = Alignment(horizontal="center", vertical="center")
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=2):
        row[0].font = Font(bold=True)
        row[0].fill = PatternFill("solid", fgColor="D9E1F2")
        row[0].alignment = Alignment(vertical="top", wrap_text=True)
        row[1].alignment = Alignment(vertical="top", wrap_text=True)


def add_kv_sheet(wb: Workbook, name: str, rows: list[tuple[str, str]]) -> None:
    if name in wb.sheetnames:
        del wb[name]
    ws = wb.create_sheet(name)
    ws.append(["字段", "内容"])
    for r in rows:
        ws.append(list(r))
    style_kv_sheet(ws)


# ─── 需求明细（每需求一张 sheet）────────────────────────────────────────────
REQUIREMENTS: dict[str, list[tuple[str, str]]] = {
    "D-L1-01-结构化日志": [
        ("需求编号", "D-L1-01"),
        ("数据层级", "L1 运行日志"),
        ("需求标题", "统一结构化 Logger 模块"),
        ("优先级", "P0"),
        ("实施阶段", "Phase 0（上线前）"),
        ("当前状态", "部分实现"),
        ("现状与缺口", "logStructuredEvent / logError / console.log 分散；无统一 level、service、输出目标；终端 stdout 无法检索与归档。"),
        ("商业影响", "线上排障依赖 SSH 看终端，无法 SLA 承诺；多实例无法汇总。"),
        ("建设目标", "全服务统一 JSON 日志：ts, level, service, eventType, correlationId, playerId, pondId, durationMs, error。"),
        ("方案-步骤1", "新建 server/src/logger.ts：info/warn/error/debug，禁止业务直接 console.log。"),
        ("方案-步骤2", "logStructuredEvent、logError、timedDbQuery 慢查询统一走 logger。"),
        ("方案-步骤3", "开发环境 pretty-print；生产环境单行 JSON。"),
        ("涉及文件", "logger.ts（新）· fishingObservability.ts · errorLog.ts · db.ts · index.ts"),
        ("依赖需求", "D-L1-06（correlationId）"),
        ("关联文档", "docs/planning/specs/服务器架构缺陷与埋点设计-v0.4.4.md §3.2"),
        ("关联xlsx", "服务器架构问题与修复方案-v0.5.xlsx → 09-R1-5-链路追踪"),
        ("验收标准", "P0 路径无裸 console.log；日志 JSON 可被 jq 解析；含 correlationId。"),
        ("预估工作量", "1~2 天"),
    ],
    "D-L1-02-日志级别策略": [
        ("需求编号", "D-L1-02"),
        ("数据层级", "L1 运行日志"),
        ("需求标题", "日志级别与环境策略"),
        ("优先级", "P1"),
        ("实施阶段", "Phase 0"),
        ("当前状态", "未开始"),
        ("现状与缺口", "无 LOG_LEVEL 环境变量；debug 与 info 未区分；perf 日志仅靠 PERF_LOG_INTERVAL_MS 节流。"),
        ("商业影响", "生产噪声大或排障信息不足；日志存储成本不可控。"),
        ("建设目标", "error 必留；warn 慢查询/降级；info 采样（phase/join）；debug 仅 development。"),
        ("方案-步骤1", "环境变量 LOG_LEVEL=error|warn|info|debug。"),
        ("方案-步骤2", "高频事件（bite_tick）默认 debug 或采样 1%。"),
        ("方案-步骤3", "文档化各级别事件清单。"),
        ("涉及文件", "logger.ts · runtimeConfig.ts"),
        ("依赖需求", "D-L1-01"),
        ("关联文档", "—"),
        ("关联xlsx", "—"),
        ("验收标准", "production 默认 info；debug 事件不出现在生产日志。"),
        ("预估工作量", "0.5 天"),
    ],
    "D-L1-03-日志落盘轮转": [
        ("需求编号", "D-L1-03"),
        ("数据层级", "L1 运行日志"),
        ("需求标题", "日志落盘与按日轮转"),
        ("优先级", "P0"),
        ("实施阶段", "Phase 0"),
        ("当前状态", "未开始"),
        ("现状与缺口", "日志仅输出终端；进程重启即丢；无 log 文件。"),
        ("商业影响", "无法事后审计；崩溃后无法追溯。"),
        ("建设目标", "单机部署：logs/ 目录按日轮转；Docker：stdout 供采集器读取。"),
        ("方案-步骤1", "pino/winston + daily-rotate-file（或 pino.destination）。"),
        ("方案-步骤2", "LOG_DIR 环境变量；默认 server/logs/。"),
        ("方案-步骤3", ".gitignore 排除 logs/；文档说明保留 7~30 天。"),
        ("涉及文件", "logger.ts · package.json"),
        ("依赖需求", "D-L1-01"),
        ("关联文档", "—"),
        ("关联xlsx", "—"),
        ("验收标准", "重启后历史日志文件仍存在；单文件超大小自动轮转。"),
        ("预估工作量", "0.5~1 天"),
    ],
    "D-L1-04-集中日志": [
        ("需求编号", "D-L1-04"),
        ("数据层级", "L1 运行日志"),
        ("需求标题", "集中日志平台接入"),
        ("优先级", "P1"),
        ("实施阶段", "Phase 1（内测/小规模商用）"),
        ("当前状态", "未开始"),
        ("现状与缺口", "无 Loki/ELK/云日志；多机无法统一检索。"),
        ("商业影响", "规模化运维不可持续；告警无法基于日志。"),
        ("建设目标", "stdout JSON → Promtail/Fluent Bit → Loki 或云日志；支持按 correlationId/playerId 检索。"),
        ("方案-步骤1", "选型：小规模 Grafana Loki + Grafana；或云厂商 SLS/CloudWatch。"),
        ("方案-步骤2", "docker-compose 增加 loki+promtail（可选）。"),
        ("方案-步骤3", "预置 Dashboard：error 率、慢查询、admin 耗时。"),
        ("涉及文件", "deploy/（新）· README 运维章节"),
        ("依赖需求", "D-L1-01 · D-L1-03"),
        ("关联文档", "—"),
        ("关联xlsx", "—"),
        ("验收标准", "Web UI 可按 correlationId 搜到完整请求链日志。"),
        ("预估工作量", "2~4 天（含部署）"),
    ],
    "D-L1-05-错误持久化": [
        ("需求编号", "D-L1-05"),
        ("数据层级", "L1 运行日志"),
        ("需求标题", "错误日志 DB 持久化"),
        ("优先级", "P0"),
        ("实施阶段", "Phase 0"),
        ("当前状态", "部分实现"),
        ("现状与缺口", "errorLog.ts 内存环形缓冲最多 200 条；/api/admin/logs 重启即空；无告警。"),
        ("商业影响", "生产异常无法追溯；合规审计缺失。"),
        ("建设目标", "error_logs 表持久化；Admin 分页查询；保留 90 天。"),
        ("方案-步骤1", "migration：error_logs(id, message, stack, context, correlation_id, created_at)。"),
        ("方案-步骤2", "logError 写 DB + 仍打 logger。"),
        ("方案-步骤3", "GET /api/admin/logs 改查 DB；支持 since/limit/context 过滤。"),
        ("涉及文件", "errorLog.ts · db.ts · migrations · admin.ts"),
        ("依赖需求", "D-L1-01"),
        ("关联文档", "docs/planning/product/v0.1.0-功能全景.md Admin 错误日志"),
        ("关联xlsx", "—"),
        ("验收标准", "重启后 /api/admin/logs 仍可查历史 error；未捕获异常均入库。"),
        ("预估工作量", "1 天"),
    ],
    "D-L1-06-全链路追踪": [
        ("需求编号", "D-L1-06"),
        ("数据层级", "L1 运行日志"),
        ("需求标题", "correlationId 全链路贯通"),
        ("优先级", "P1"),
        ("实施阶段", "Phase 1"),
        ("当前状态", "部分实现"),
        ("现状与缺口", "v0.5 R1-5 已部分落地：Socket correlationId + HTTP X-Request-Id；metrics 入库与 Admin timeline 仍缺 correlationId 过滤。"),
        ("商业影响", "复杂 bug 无法一键串链。"),
        ("建设目标", "HTTP/Socket/状态机/metrics/error 同一 correlationId；Admin 支持按 ID 查询。"),
        ("方案-步骤1", "recordFishingMetric payload 自动附加 correlationId。"),
        ("方案-步骤2", "Admin GET timeline?correlationId=。"),
        ("方案-步骤3", "客户端可选上报 correlationId（调试包）。"),
        ("涉及文件", "fishingObservability.ts · fishingMetrics.ts · admin.ts · sessionRegistry.ts"),
        ("依赖需求", "—"),
        ("关联文档", "docs/planning/specs/服务器架构优化路线图-v0.5.md R1-5"),
        ("关联xlsx", "服务器架构问题与修复方案-v0.5.xlsx → 09-R1-5-链路追踪"),
        ("验收标准", "一次 join→咬钩→断线 全链日志同一 correlationId。"),
        ("预估工作量", "1 天（补全）"),
    ],
    "D-L1-07-告警通知": [
        ("需求编号", "D-L1-07"),
        ("数据层级", "L1 运行日志"),
        ("需求标题", "运维告警规则与通知"),
        ("优先级", "P1"),
        ("实施阶段", "Phase 1"),
        ("当前状态", "未开始"),
        ("现状与缺口", "无 Prometheus/Alertmanager；错误率、断连率、磁盘无自动告警。"),
        ("商业影响", "故障发现依赖用户反馈。"),
        ("建设目标", "最低配：进程存活、5xx、未捕获异常、SQLite 磁盘、metrics 写入失败。"),
        ("方案-步骤1", "Prometheus exporter 或健康检查 cron。"),
        ("方案-步骤2", "Alertmanager → 钉钉/邮件/Webhook。"),
        ("方案-步骤3", "运行手册：告警 → 查 logs → 查 player timeline。"),
        ("涉及文件", "deploy/monitoring/（新）"),
        ("依赖需求", "D-L1-04 · D-L2-05"),
        ("关联文档", "docs/planning/specs/排查-挂机断线诊断阶段2-4.md SOP"),
        ("关联xlsx", "—"),
        ("验收标准", "模拟进程 crash 5 分钟内收到告警。"),
        ("预估工作量", "2~3 天"),
    ],
    "D-L1-08-健康监控": [
        ("需求编号", "D-L1-08"),
        ("数据层级", "L1 运行日志"),
        ("需求标题", "健康检查与进程指标"),
        ("优先级", "P0"),
        ("实施阶段", "Phase 0"),
        ("当前状态", "部分实现"),
        ("现状与缺口", "已有 GET /health → { ok: true }，但无 /ready、无 DB ping、无 metrics 队列深度、无 event loop lag/连接数。"),
        ("商业影响", "负载均衡/K8s 无法探活；容量规划盲目。"),
        ("建设目标", "GET /health（db ping + 队列深度）；GET /metrics Prometheus 格式（可选）。"),
        ("方案-步骤1", "/health 返回 ok、db、uptime、metricsQueueDepth。"),
        ("方案-步骤2", "process.memoryUsage、activeSockets、pondUserCount 暴露。"),
        ("方案-步骤3", "优雅停机时 /health 返回 503。"),
        ("涉及文件", "index.ts · fishingMetrics.ts"),
        ("依赖需求", "—"),
        ("关联xlsx", "服务器架构问题与修复方案-v0.5.xlsx → 08-R1-4-优雅停机"),
        ("验收标准", "K8s liveness 可用；停机时 health 失败。"),
        ("预估工作量", "0.5~1 天"),
    ],
    "D-L1-09-日志合规": [
        ("需求编号", "D-L1-09"),
        ("数据层级", "L1 运行日志"),
        ("需求标题", "日志脱敏与合规"),
        ("优先级", "P2"),
        ("实施阶段", "Phase 2"),
        ("当前状态", "未开始"),
        ("现状与缺口", "日志可能含 playerId、token 片段；无保留策略文档。"),
        ("商业影响", "商用合规风险（隐私/GDPR 类要求）。"),
        ("建设目标", "禁止记录 token/密码；playerId 可配置哈希；日志保留策略文档化。"),
        ("方案-步骤1", "logger 层 redact 字段列表。"),
        ("方案-步骤2", "运维文档：保留 30~90 天、访问权限。"),
        ("涉及文件", "logger.ts · docs/ops/（新）"),
        ("依赖需求", "D-L1-01"),
        ("验收标准", "代码审查无 token 明文日志；有数据保留政策文档。"),
        ("预估工作量", "1 天"),
    ],
    "D-L1-10-OpenTelemetry": [
        ("需求编号", "D-L1-10"),
        ("数据层级", "L1 运行日志"),
        ("需求标题", "OpenTelemetry 分布式追踪"),
        ("优先级", "P2"),
        ("实施阶段", "Phase 2"),
        ("当前状态", "未开始"),
        ("现状与缺口", "仅有 correlationId 字符串串联；无 span 树、无跨 handler 耗时分解、无 Jaeger/Tempo 可视化。"),
        ("商业影响", "复杂竞态/性能问题仍需人工拼日志；多服务扩展后链路不可见。"),
        ("建设目标", "关键路径 span：join_pond → takeSpot → phase_transition → bite_check → disconnect；trace_id 与 correlationId 对齐。"),
        ("方案-步骤1", "接入 @opentelemetry/sdk-node + OTLP exporter（Jaeger/Tempo 或云 APM）。"),
        ("方案-步骤2", "socketPondHandlers、fishingStateMachine、serverLoops 关键节点手动 span。"),
        ("方案-步骤3", "Grafana 或 Jaeger UI 预置「一次钓鱼会话」查询模板。"),
        ("涉及文件", "index.ts · socketPondHandlers.ts · fishingStateMachine.ts · deploy/otel/（新）"),
        ("依赖需求", "D-L1-01 · D-L1-06"),
        ("关联文档", "docs/planning/specs/三层数据体系-可观测性补充-v0.6.md §2.1"),
        ("关联xlsx", "服务器架构问题与修复方案-v0.5.xlsx → 09-R1-5-链路追踪"),
        ("验收标准", "Jaeger 中可按 correlationId 看到 join→waiting→bite 完整 span 链；P95 handler 耗时可读。"),
        ("预估工作量", "3~5 天"),
    ],
    "D-L1-11-动态调试采样": [
        ("需求编号", "D-L1-11"),
        ("数据层级", "L1 运行日志"),
        ("需求标题", "按玩家/鱼塘动态提级 DEBUG 采样"),
        ("优先级", "P1"),
        ("实施阶段", "Phase 1"),
        ("当前状态", "未开始"),
        ("现状与缺口", "生产只能全局 LOG_LEVEL；客诉时无法对单个 playerId 临时打开详细日志而不打爆磁盘。"),
        ("商业影响", "线上个案排查要么信息不足，要么全开 debug 成本失控。"),
        ("建设目标", "runtimeConfig 或环境变量指定 debugPlayers/debugPonds；命中目标 100% debug、其余维持 info；TTL 30 分钟自动失效。"),
        ("方案-步骤1", "logger 层 isDebugTarget(playerId?, pondId?) 判断。"),
        ("方案-步骤2", "Admin POST /api/admin/observability/debug-targets 增删（operator 角色）。"),
        ("方案-步骤3", "审计日志记录谁对谁开了 debug。"),
        ("涉及文件", "logger.ts · runtimeConfig.ts · admin.ts"),
        ("依赖需求", "D-L1-01 · D-L1-02"),
        ("关联文档", "docs/planning/specs/三层数据体系-可观测性补充-v0.6.md §2.2"),
        ("关联xlsx", "—"),
        ("验收标准", "对指定 playerId 开启后 5 分钟内日志含 bite_tick 等 debug 事件；过期自动恢复 info。"),
        ("预估工作量", "1~2 天"),
    ],
    "D-L1-12-Socket事件Tap": [
        ("需求编号", "D-L1-12"),
        ("数据层级", "L1 运行日志"),
        ("需求标题", "Socket.io 全事件 Tap（入站/出站）"),
        ("优先级", "P2"),
        ("实施阶段", "Phase 1"),
        ("当前状态", "未开始"),
        ("现状与缺口", "仅 instrumented 路径有日志；未覆盖的 client emit 或新增事件无观测。"),
        ("商业影响", "协议变更或客户端 bug 导致「发了但服务端没 handler」难以发现。"),
        ("建设目标", "socket.onAny + onAnyOutgoing 记录 event 名、payload 摘要（脱敏）、方向；生产默认 1% 采样，debug 目标 100%。"),
        ("方案-步骤1", "socketLifecycle.ts 注册 onAny/onAnyOutgoing middleware。"),
        ("方案-步骤2", "与 D-L1-11 debug 目标联动；payload 截断至 512 字符。"),
        ("方案-步骤3", "未知 event 名打 warn 级结构化日志。"),
        ("涉及文件", "socketLifecycle.ts · logger.ts"),
        ("依赖需求", "D-L1-01 · D-L1-11"),
        ("关联文档", "docs/planning/specs/三层数据体系-可观测性补充-v0.6.md §2.3"),
        ("关联xlsx", "—"),
        ("验收标准", "debug 模式下可见 client emit 的任意事件；生产采样不产生明显 CPU 抖动。"),
        ("预估工作量", "1 天"),
    ],
    "D-L2-01-事件字典": [
        ("需求编号", "D-L2-01"),
        ("数据层级", "L2 指标与事件"),
        ("需求标题", "事件字典与 Payload Schema 契约"),
        ("优先级", "P0"),
        ("实施阶段", "Phase 0"),
        ("当前状态", "部分实现"),
        ("现状与缺口", "v0.4.4 埋点表在 xlsx 有 33 条；代码 FishingMetricEvent 已扩展；缺机器可读 schema 与 CI 校验。"),
        ("商业影响", "埋点漂移导致报表不可信；前后端对不齐。"),
        ("建设目标", "shared/metrics-schema.ts 或 JSON Schema；与 v0.4.4-埋点表清单.xlsx 双向同步。"),
        ("方案-步骤1", "从 xlsx/文档导出 events.json schema。"),
        ("方案-步骤2", "recordFishingMetric 开发环境校验 payload 必填字段。"),
        ("方案-步骤3", "PR 检查：新增 event 必须更新 schema + xlsx。"),
        ("涉及文件", "shared/ · fishingMetrics.ts · docs/planning/reports/v0.4.4-埋点表清单.xlsx"),
        ("依赖需求", "—"),
        ("关联文档", "docs/planning/specs/v0.4.4-埋点缺口复核与补全.md"),
        ("关联xlsx", "v0.4.4-埋点表清单.xlsx（33 条埋点标准格式）"),
        ("验收标准", "每个 event_type 有文档化 payload；非法 payload 开发环境 warn。"),
        ("预估工作量", "1~2 天"),
    ],
    "D-L2-02-保留归档": [
        ("需求编号", "D-L2-02"),
        ("数据层级", "L2 指标与事件"),
        ("需求标题", "fishing_metrics 保留与归档策略"),
        ("优先级", "P0"),
        ("实施阶段", "Phase 0"),
        ("当前状态", "未开始"),
        ("现状与缺口", "SQLite fishing_metrics 无限增长；无 TTL、无归档、无聚合表。"),
        ("商业影响", "长期运行磁盘撑爆；查询变慢。"),
        ("建设目标", "原始事件保留 30~90 天；超期删除或导出 OSS；日聚合永久保留。"),
        ("方案-步骤1", "cron scripts/archive-metrics.mjs：导出 JSONL + DELETE WHERE created_at < ?。"),
        ("方案-步骤2", "环境变量 METRICS_RETENTION_DAYS。"),
        ("方案-步骤3", "启动时检查表大小告警。"),
        ("涉及文件", "scripts/archive-metrics.mjs（新）· db.ts"),
        ("依赖需求", "D-L2-03"),
        ("关联xlsx", "服务器架构问题与修复方案-v0.5.xlsx → 11-R2-2-SQLite写压"),
        ("验收标准", "90 天前数据可归档删除；聚合表仍可查历史趋势。"),
        ("预估工作量", "1~2 天"),
    ],
    "D-L2-03-日聚合表": [
        ("需求编号", "D-L2-03"),
        ("数据层级", "L2 指标与事件"),
        ("需求标题", "日聚合表 daily_player / daily_pond_stats"),
        ("优先级", "P0"),
        ("实施阶段", "Phase 0"),
        ("当前状态", "未开始"),
        ("现状与缺口", "仅有 getFishingMetricsSummary 内存聚合；无按日落库；无法看 30 天趋势。"),
        ("商业影响", "运营无法看日活、日钓量、断线率趋势。"),
        ("建设目标", "每日 00:05 cron 聚合：玩家日钓/断线/时长；鱼塘人口/咬钩率。"),
        ("方案-步骤1", "表 daily_player_stats、daily_pond_stats。"),
        ("方案-步骤2", "scripts/aggregate-daily-metrics.mjs。"),
        ("方案-步骤3", "Admin 或 HTML 只读趋势页。"),
        ("涉及文件", "migrations · scripts/ · admin.ts"),
        ("依赖需求", "D-L2-01"),
        ("验收标准", "可查询任意历史日的四塘日钓合计、断线次数。"),
        ("预估工作量", "2~3 天"),
    ],
    "D-L2-04-存储演进": [
        ("需求编号", "D-L2-04"),
        ("数据层级", "L2 指标与事件"),
        ("需求标题", "Metrics 存储演进（PostgreSQL/Timescale）"),
        ("优先级", "P2"),
        ("实施阶段", "Phase 2"),
        ("当前状态", "未开始"),
        ("现状与缺口", "全量 metrics 在 SQLite；写已批量但仍单库单连接。"),
        ("商业影响", "日活上千后写入与查询成为瓶颈。"),
        ("建设目标", "metrics 迁 PostgreSQL 或 TimescaleDB；生态仍可 SQLite 或一并迁移。"),
        ("方案-步骤1", "抽象 MetricsStore 接口。"),
        ("方案-步骤2", "双写验证后切读。"),
        ("方案-步骤3", "文档化连接池与备份。"),
        ("涉及文件", "fishingMetrics.ts · db.ts"),
        ("依赖需求", "D-L2-02 · D-L2-03"),
        ("关联xlsx", "服务器架构问题与修复方案-v0.5.xlsx → 11-R2-2 · 12-R2-3"),
        ("验收标准", "压测 100 并发写入无丢事件；查询 timeline <500ms。"),
        ("预估工作量", "1~2 周"),
    ],
    "D-L2-05-RED指标": [
        ("需求编号", "D-L2-05"),
        ("数据层级", "L2 指标与事件"),
        ("需求标题", "RED 指标与 Grafana 业务大盘"),
        ("优先级", "P1"),
        ("实施阶段", "Phase 1"),
        ("当前状态", "部分实现"),
        ("现状与缺口", "P2 性能日志已有 tick/snapshot/admin 耗时；无 Prometheus 指标；无 Grafana 大盘。"),
        ("商业影响", "容量与性能退化不可见。"),
        ("建设目标", "Rate/Errors/Duration：API、Socket、咬钩循环、snapshot；在线人数、各塘负载。"),
        ("方案-步骤1", "prom-client 暴露 /metrics。"),
        ("方案-步骤2", "Grafana dashboard JSON 入库 deploy/。"),
        ("方案-步骤3", "复用现有 perf 日志字段对齐 metric 名。"),
        ("涉及文件", "index.ts · gameState.ts · deploy/grafana/"),
        ("依赖需求", "D-L1-08"),
        ("关联xlsx", "v0.4.4-埋点表清单.xlsx → 性能 sheet"),
        ("验收标准", "Grafana 可看 24h 咬钩循环 P95 耗时、断线率。"),
        ("预估工作量", "2~3 天"),
    ],
    "D-L2-06-Admin增强": [
        ("需求编号", "D-L2-06"),
        ("数据层级", "L2 指标与事件"),
        ("需求标题", "Admin 时间线导出与排查模板"),
        ("优先级", "P1"),
        ("实施阶段", "Phase 1"),
        ("当前状态", "部分实现"),
        ("现状与缺口", "已有 timeline API + AdminMetricsPanel；无 CSV 导出、无 correlationId 过滤、无预置排查模板。"),
        ("商业影响", "客服/运营排查效率低。"),
        ("建设目标", "导出 CSV；按 correlationId 过滤；「断线」「咬钩异常」「pending 超时」一键模板。"),
        ("方案-步骤1", "GET timeline/export.csv。"),
        ("方案-步骤2", "Admin UI 增加导出与模板按钮。"),
        ("涉及文件", "admin.ts · AdminMetricsPanel.tsx"),
        ("依赖需求", "D-L1-06"),
        ("关联文档", "docs/planning/specs/排查-挂机断线诊断阶段2-4.md"),
        ("验收标准", "给定 playerId 可导出 24h CSV；SOP 步骤可在 Admin 完成。"),
        ("预估工作量", "1~2 天"),
    ],
    "D-L2-07-Admin-RBAC": [
        ("需求编号", "D-L2-07"),
        ("数据层级", "L2 指标与事件"),
        ("需求标题", "Admin RBAC 与生产只读角色"),
        ("优先级", "P1"),
        ("实施阶段", "Phase 1（公网前必须）"),
        ("当前状态", "部分实现"),
        ("现状与缺口", "单 ADMIN_SECRET；破坏性操作与查询同权；生产默认密钥已拦但无角色。"),
        ("商业影响", "管理面泄露等于全库重置。"),
        ("建设目标", "角色：readonly（查 metrics/logs/debug）、operator（调配置）、admin（reset/clear）。"),
        ("方案-步骤1", "多 key 或 JWT role claim。"),
        ("方案-步骤2", "destructive 操作写 audit_log。"),
        ("涉及文件", "admin.ts · auth.ts"),
        ("依赖需求", "D-L2-08 架构 R0-1 鉴权"),
        ("关联xlsx", "服务器架构问题与修复方案-v0.5.xlsx → 01-R0-1 · 13-R2-4"),
        ("验收标准", "readonly key 无法 POST ecology/reset。"),
        ("预估工作量", "1~2 天"),
    ],
    "D-L2-08-埋点缺口闭环": [
        ("需求编号", "D-L2-08"),
        ("数据层级", "L2 指标与事件"),
        ("需求标题", "v0.4.4 埋点表 33 条与代码完全对齐"),
        ("优先级", "P1"),
        ("实施阶段", "Phase 0~1"),
        ("当前状态", "部分实现"),
        ("现状与缺口", "patch2 文档列 3 项真缺失、2 别名、2 不完整；xlsx 状态可能滞后。"),
        ("商业影响", "按 eventType 统计漏数；性能与 session 问题不可见。"),
        ("建设目标", "33 条全部「已实现」或明确废弃；eventType canonical 名统一；verify 覆盖。"),
        ("方案-步骤1", "补齐 session_rebound、snapshot_build_duration_ms、sqlite_query_slow。"),
        ("方案-步骤2", "socket_disconnect / pending_catch_accept 写入 canonical 名。"),
        ("方案-步骤3", "同步更新 v0.4.4-埋点表清单.xlsx 状态列。"),
        ("涉及文件", "fishingMetrics.ts · gameState.ts · db.ts · verify-server-observability.ts"),
        ("关联文档", "docs/planning/specs/v0.4.4-埋点缺口复核与补全.md"),
        ("关联xlsx", "v0.4.4-埋点表清单.xlsx（总览 + 6 分类 sheet）"),
        ("验收标准", "npm run verify:server-observability 全绿；xlsx 状态与代码一致。"),
        ("预估工作量", "1~2 天（补缺口）"),
    ],
    "D-L2-09-业务健康看板": [
        ("需求编号", "D-L2-09"),
        ("数据层级", "L2 指标与事件"),
        ("需求标题", "业务健康看板（咬钩/断线/经济）"),
        ("优先级", "P1"),
        ("实施阶段", "Phase 1"),
        ("当前状态", "部分实现"),
        ("现状与缺口", "getFishingMetricsSummary 有 catch/escape/abandon/金币估算；无趋势、无分塘、无咬钩 tick 命中率。"),
        ("商业影响", "数值调参无法从线上验证。"),
        ("建设目标", "Admin 或 Grafana：日钓量、人均时产、断线率、空杆率、bait 消耗、各塘人口率。"),
        ("方案-步骤1", "基于 daily 聚合表出图。"),
        ("方案-步骤2", "bite_tick_hit / miss 比率看板。"),
        ("涉及文件", "fishingMetrics.ts · AdminMetricsPanel.tsx"),
        ("依赖需求", "D-L2-03"),
        ("验收标准", "可对比 v0.4.1 目标（~100 条/天）与线上 7 日均值。"),
        ("预估工作量", "2 天"),
    ],
    "D-L2-10-幂等去重": [
        ("需求编号", "D-L2-10"),
        ("数据层级", "L2 指标与事件"),
        ("需求标题", "关键事件 eventId 幂等"),
        ("优先级", "P2"),
        ("实施阶段", "Phase 2"),
        ("当前状态", "未开始"),
        ("现状与缺口", "每条 metric 新 UUID；重试可能双写；无去重键。"),
        ("商业影响", "财务/产出类统计可能偏高。"),
        ("建设目标", "catch_accept、pending_catch_accept 等带业务 eventId；DB 唯一索引可选。"),
        ("方案-步骤1", "payload.eventId = pendingCatchId 等。"),
        ("方案-步骤2", "聚合时 DISTINCT eventId。"),
        ("涉及文件", "fishingMetrics.ts · inventory.ts"),
        ("验收标准", "重复 accept 不产生双计数。"),
        ("预估工作量", "1 天"),
    ],
    "D-L2-11-DB备份": [
        ("需求编号", "D-L2-11"),
        ("数据层级", "L2 指标与事件"),
        ("需求标题", "SQLite 定时备份与恢复演练"),
        ("优先级", "P0"),
        ("实施阶段", "Phase 0"),
        ("当前状态", "未开始"),
        ("现状与缺口", "fish-social.db 无自动备份；误操作 reset 不可恢复。"),
        ("商业影响", "数据丢失即业务灾难。"),
        ("建设目标", "每日冷备；保留 7 天；文档化恢复步骤。"),
        ("方案-步骤1", "scripts/backup-db.mjs：sqlite .backup + 压缩。"),
        ("方案-步骤2", "cron / 云盘同步。"),
        ("方案-步骤3", "季度恢复演练 checklist。"),
        ("涉及文件", "scripts/ · docs/ops/"),
        ("验收标准", "可从昨日备份恢复到新实例并查 metrics。"),
        ("预估工作量", "0.5~1 天"),
    ],
    "D-L2-12-客户端上报": [
        ("需求编号", "D-L2-12"),
        ("数据层级", "L2 指标与事件"),
        ("需求标题", "客户端关键事件上报（可选）"),
        ("优先级", "P2"),
        ("实施阶段", "Phase 2"),
        ("当前状态", "部分实现"),
        ("现状与缺口", "pondLifecycleLog 仅 __DEV__ 内存 200 条；未上报服务端。"),
        ("商业影响", "纯服务端无法解释「客户端切后台」类问题。"),
        ("建设目标", "生产采样上报：connect/disconnect/join_fail/app_background；与 timeline 对照。"),
        ("方案-步骤1", "POST /api/client-events 批量、限流。"),
        ("方案-步骤2", "client_metrics 表或合入 fishing_metrics。"),
        ("涉及文件", "mobile/lib/pondLifecycleLog.ts · server routes"),
        ("关联文档", "docs/planning/specs/排查-挂机断线诊断阶段2-4.md 阶段2"),
        ("验收标准", "断线 case 可同时看到服务端 disconnect 与客户端 disconnect 事件。"),
        ("预估工作量", "2~3 天"),
    ],
    "D-L2-13-Live会话调试": [
        ("需求编号", "D-L2-13"),
        ("数据层级", "L2 指标与事件"),
        ("需求标题", "Live Session Inspector（实时会话调试）"),
        ("优先级", "P1"),
        ("实施阶段", "Phase 1"),
        ("当前状态", "未开始"),
        ("现状与缺口", "A2 fishing-debug 为 3s 缓存快照；无法实时订阅玩家 state 变化；策划/客服需 SSH 或反复刷新。"),
        ("商业影响", "「玩家在干什么」类客诉响应慢；与客户端 lifecycle 无法同屏对照。"),
        ("建设目标", "Admin 选定 playerId 后：每秒推送 enriched state + 最近 20 条 metrics；可选展示客户端上报事件（D-L2-12）；仅 staging/dev 或 readonly+审计。"),
        ("方案-步骤1", "GET /api/admin/sessions/:playerId/live SSE 或 WebSocket admin channel。"),
        ("方案-步骤2", "服务端内存 ring：每玩家最近 N 条 metric 事件。"),
        ("方案-步骤3", "Admin UI「实时监视」Tab：左服务端 state、右客户端 lifecycle（若已上报）。"),
        ("涉及文件", "admin.ts · sessionRegistry.ts · AdminMetricsPanel.tsx 或 admin-web/"),
        ("依赖需求", "D-L1-06 · D-L2-06"),
        ("关联文档", "docs/planning/specs/A2-Debug面板.md · 三层数据体系-可观测性补充-v0.6.md §2.4"),
        ("关联xlsx", "—"),
        ("验收标准", "waiting 阶段可实时看到 sessionFishingMs 递增；断线瞬间可见 phase→disconnected。"),
        ("预估工作量", "2~3 天"),
    ],
    "D-L2-14-admin-web运维台": [
        ("需求编号", "D-L2-14"),
        ("数据层级", "L2 指标与事件"),
        ("需求标题", "独立 admin-web 运维控制台"),
        ("优先级", "P2"),
        ("实施阶段", "Phase 1~2"),
        ("当前状态", "未开始"),
        ("现状与缺口", "Admin 嵌在 Expo mobile/app/admin.tsx；大屏检索、多 Tab 日志、图表体验差。"),
        ("商业影响", "运维/策划排查效率低；难以作为商业化交付的「运营后台」。"),
        ("建设目标", "Vite+React admin-web/：鱼塘 debug、metrics 大盘、player timeline、日志检索入口（对接 Loki 或 Admin API）。"),
        ("方案-步骤1", "抽离 adminApi 为 shared 包或复制类型；admin-web 复用 REST。"),
        ("方案-步骤2", "路由：鱼塘 / 玩家 / 日志 / 配置；保留 mobile Admin 作现场抽查。"),
        ("方案-步骤3", "与 D-L2-13 Live Inspector 同页集成。"),
        ("涉及文件", "admin-web/（新）· mobile/lib/adminApi.ts"),
        ("依赖需求", "D-L2-06 · D-L1-04（日志检索可选）"),
        ("关联文档", "docs/planning/specs/三层数据体系-可观测性补充-v0.6.md §2.5"),
        ("关联xlsx", "—"),
        ("验收标准", "桌面浏览器可完成 SOP 全流程（timeline 导出 + fishing-debug + 配置只读）。"),
        ("预估工作量", "3~5 天"),
    ],
    "D-L3-01-日批流水线": [
        ("需求编号", "D-L3-01"),
        ("数据层级", "L3 产品/运营分析"),
        ("需求标题", "日批分析流水线（metrics → 报表）"),
        ("优先级", "P0"),
        ("实施阶段", "Phase 0~1"),
        ("当前状态", "部分实现"),
        ("现状与缺口", "docs/analytics 仅有离线 simulate-pond-day；与线上 fishing_metrics 无连接。"),
        ("商业影响", "设计目标（100 条/天）无法与真实数据对照。"),
        ("建设目标", "npm run analytics:daily：读 DB 聚合 → 写入 docs/analytics/daily/ 或 internal HTML。"),
        ("方案-步骤1", "scripts/analytics/daily-pipeline.mjs。"),
        ("方案-步骤2", "产出：日钓量、断线率、经济、人口率。"),
        ("方案-步骤3", "与 simulate 报告同版式对比栏。"),
        ("涉及文件", "scripts/analytics/ · docs/analytics/daily/"),
        ("依赖需求", "D-L2-03"),
        ("验收标准", "cron 每日自动生成昨日报告 HTML。"),
        ("预估工作量", "2~3 天"),
    ],
    "D-L3-02-线上模拟对照": [
        ("需求编号", "D-L3-02"),
        ("数据层级", "L3 产品/运营分析"),
        ("需求标题", "线上实测 vs 蒙特卡洛模拟对照"),
        ("优先级", "P1"),
        ("实施阶段", "Phase 1"),
        ("当前状态", "部分实现"),
        ("现状与缺口", "bite-escape-calibration 为理论公式；pond-day 为模拟；缺「线上 7 日均钓 vs 模拟 7 日」。"),
        ("商业影响", "调参不知是否偏离设计。"),
        ("建设目标", "报告一页：模拟日均/线上日均/偏差%；按规则版本切分。"),
        ("方案-步骤1", "daily 聚合 + compact.json 同指标定义。"),
        ("方案-步骤2", "analysis-report 增加「线上对照」章节。"),
        ("涉及文件", "generate-analysis-report.mjs · daily-pipeline.mjs"),
        ("依赖需求", "D-L3-01"),
        ("验收标准", "可展示 5人/塘场景模拟 118 vs 线上 X 条/天。"),
        ("预估工作量", "1~2 天"),
    ],
    "D-L3-03-留存会话": [
        ("需求编号", "D-L3-03"),
        ("数据层级", "L3 产品/运营分析"),
        ("需求标题", "留存与会话时长分析"),
        ("优先级", "P1"),
        ("实施阶段", "Phase 1"),
        ("当前状态", "未开始"),
        ("现状与缺口", "无 D1/D7 留存；fishing_start/stop 可推导但未聚合。"),
        ("商业影响", "无法评估商业化留存与付费转化。"),
        ("建设目标", "daily_active_players、session 时长分布、D1/D7  cohort。"),
        ("方案-步骤1", "players 表 last_seen；每日活跃标记。"),
        ("方案-步骤2", "weekly-retention.mjs。"),
        ("涉及文件", "migrations · scripts/analytics/"),
        ("依赖需求", "D-L2-03 · D-L2-01"),
        ("验收标准", "可输出上周注册玩家 D7 留存率。"),
        ("预估工作量", "2~3 天"),
    ],
    "D-L3-04-经济日报": [
        ("需求编号", "D-L3-04"),
        ("数据层级", "L3 产品/运营分析"),
        ("需求标题", "经济系统 faucet/sink 日报"),
        ("优先级", "P1"),
        ("实施阶段", "Phase 1"),
        ("当前状态", "部分实现"),
        ("现状与缺口", "summary 有 faucetCoinsEstimate/sinkCoinsEstimate 但仅滑动窗口、无日趋势。"),
        ("商业影响", "通胀/通缩不可见。"),
        ("建设目标", "按日：卖饵/修竿 sink、卖鱼/任务 faucet；净变化曲线。"),
        ("方案-步骤1", "补全 catch 金币、faucet 类事件埋点。"),
        ("方案-步骤2", "daily_economy_stats 表。"),
        ("涉及文件", "fishingMetrics.ts · shop.ts · scripts/analytics/"),
        ("依赖需求", "D-L3-01"),
        ("验收标准", "可发现连续 3 日 faucet>sink 告警。"),
        ("预估工作量", "2 天"),
    ],
    "D-L3-05-生态日报": [
        ("需求编号", "D-L3-05"),
        ("数据层级", "L3 产品/运营分析"),
        ("需求标题", "鱼塘生态健康日报"),
        ("优先级", "P1"),
        ("实施阶段", "Phase 1"),
        ("当前状态", "部分实现"),
        ("现状与缺口", "生态在 SQLite；无每日人口率、品质结构、补充量线上报表。"),
        ("商业影响", "枯竭或稀有鱼曝光不足无法预警。"),
        ("建设目标", "每日快照各塘人口/上限、品质分布、补充条数；对比模拟阈值 85%。"),
        ("方案-步骤1", "ecology_daily_snapshot 表或 JSON 归档。"),
        ("方案-步骤2", "接入 analytics 报告样式。"),
        ("涉及文件", "pondEcology.ts · scripts/analytics/"),
        ("依赖需求", "D-L3-01"),
        ("验收标准", "人口率 <70% 自动标红。"),
        ("预估工作量", "2 天"),
    ],
    "D-L3-06-BI对接": [
        ("需求编号", "D-L3-06"),
        ("数据层级", "L3 产品/运营分析"),
        ("需求标题", "BI / 数据仓库对接"),
        ("优先级", "P2"),
        ("实施阶段", "Phase 2"),
        ("当前状态", "未开始"),
        ("现状与缺口", "数据困在 SQLite 单文件；无 ETL。"),
        ("商业影响", "运营无法自助分析；多产品数据无法融合。"),
        ("建设目标", "每日导出 Parquet/CSV 到 OSS；可选 BigQuery/ClickHouse 同步。"),
        ("方案-步骤1", "export-warehouse.mjs 导出 daily_* + 维度表。"),
        ("方案-步骤2", "Metabase/Superset 只读连接。"),
        ("涉及文件", "scripts/ · deploy/"),
        ("依赖需求", "D-L2-04 · D-L3-01"),
        ("验收标准", "运营可在 BI 工具拖图表看 30 日趋势。"),
        ("预估工作量", "1~2 周"),
    ],
    "D-L3-07-版本维度": [
        ("需求编号", "D-L3-07"),
        ("数据层级", "L3 产品/运营分析"),
        ("需求标题", "规则版本与配置版本切分分析"),
        ("优先级", "P1"),
        ("实施阶段", "Phase 1"),
        ("当前状态", "部分实现"),
        ("现状与缺口", "模拟有 rulesVersion；线上 metrics 未系统记录 BITE_BASE_SCALE 等配置版本。"),
        ("商业影响", "调参后无法对比前后效果。"),
        ("建设目标", "metrics payload 或 daily 表含 rulesVersion、configSnapshotId。"),
        ("方案-步骤1", "runtimeConfig 变更写 config_audit；日报按版本 group by。"),
        ("方案-步骤2", "与 docs/analytics/runs 归档 ID 对齐。"),
        ("涉及文件", "runtimeConfig.ts · fishingMetrics.ts"),
        ("验收标准", "可对比 v0.4.0 vs v0.4.1 上线前后日钓量。"),
        ("预估工作量", "1 天"),
    ],
    "D-L3-08-运营告警": [
        ("需求编号", "D-L3-08"),
        ("数据层级", "L3 产品/运营分析"),
        ("需求标题", "运营指标告警（产量/生态）"),
        ("优先级", "P1"),
        ("实施阶段", "Phase 1"),
        ("当前状态", "未开始"),
        ("现状与缺口", "无「日钓超 150」「人口 <70%」类业务告警。"),
        ("商业影响", "数值失控后玩家体验已受损才被发现。"),
        ("建设目标", "daily 流水线后检查阈值 → Webhook 通知策划。"),
        ("方案-步骤1", "alert-rules.json：dailyCatchMax、popRatioMin。"),
        ("方案-步骤2", "集成钉钉/飞书。"),
        ("涉及文件", "scripts/analytics/daily-pipeline.mjs"),
        ("依赖需求", "D-L3-01 · D-L3-05"),
        ("验收标准", "模拟超标数据触发一条运营告警。"),
        ("预估工作量", "1 天"),
    ],
    "D-L3-09-analytics打通": [
        ("需求编号", "D-L3-09"),
        ("数据层级", "L3 产品/运营分析"),
        ("需求标题", "docs/analytics 索引纳入线上日报"),
        ("优先级", "P1"),
        ("实施阶段", "Phase 1"),
        ("当前状态", "部分实现"),
        ("现状与缺口", "index.html 仅归档 simulate runs；无 live-daily 入口。"),
        ("商业影响", "分析资产分散。"),
        ("建设目标", "index 增加「线上日报」区；与 runs 并列；manifest 含 type=live-daily。"),
        ("方案-步骤1", "扩展 build-index.mjs · archive-run 支持 live-daily。"),
        ("方案-步骤2", "compare 页可选模拟 vs 线上同日。"),
        ("涉及文件", "scripts/analytics/build-index.mjs"),
        ("依赖需求", "D-L3-01"),
        ("验收标准", "index 可打开最新线上日报告。"),
        ("预估工作量", "1 天"),
    ],
    "D-L3-10-数据合规": [
        ("需求编号", "D-L3-10"),
        ("数据层级", "L3 产品/运营分析"),
        ("需求标题", "用户数据导出/删除与指标脱敏"),
        ("优先级", "P2"),
        ("实施阶段", "Phase 2"),
        ("当前状态", "未开始"),
        ("现状与缺口", "无 GDPR 式删除；metrics 永久关联 playerId。"),
        ("商业影响", "商用合规风险。"),
        ("建设目标", "账号删除时匿名化 metrics；提供玩家数据导出 API。"),
        ("方案-步骤1", "DELETE player 流程 cascade 或 hash player_id。"),
        ("方案-步骤2", "隐私政策文档。"),
        ("涉及文件", "admin.ts · migrations"),
        ("依赖需求", "D-L1-09"),
        ("验收标准", "删除账号后 timeline 不可还原身份。"),
        ("预估工作量", "2~3 天"),
    ],
}

OVERVIEW_ROWS: list[list] = [
    # 编号, 层级, 标题, 优先级, 阶段, 状态, 缺口摘要, sheet名, 依赖, 关联架构, 关联埋点
    ["D-L1-01", "L1", "统一结构化 Logger", "P0", "Phase 0", "部分实现", "console 分散、无 JSON 标准", "D-L1-01-结构化日志", "D-L1-06", "R1-5", "—"],
    ["D-L1-02", "L1", "日志级别与环境策略", "P1", "Phase 0", "未开始", "无 LOG_LEVEL", "D-L1-02-日志级别策略", "D-L1-01", "—", "—"],
    ["D-L1-03", "L1", "日志落盘与轮转", "P0", "Phase 0", "未开始", "仅终端输出", "D-L1-03-日志落盘轮转", "D-L1-01", "—", "—"],
    ["D-L1-04", "L1", "集中日志平台", "P1", "Phase 1", "未开始", "无 Loki/ELK", "D-L1-04-集中日志", "D-L1-01,03", "—", "—"],
    ["D-L1-05", "L1", "错误日志 DB 持久化", "P0", "Phase 0", "部分实现", "内存 200 条", "D-L1-05-错误持久化", "D-L1-01", "—", "—"],
    ["D-L1-06", "L1", "correlationId 全链路", "P1", "Phase 1", "部分实现", "metrics 未全带 ID", "D-L1-06-全链路追踪", "—", "R1-5", "—"],
    ["D-L1-07", "L1", "运维告警", "P1", "Phase 1", "未开始", "无 Alertmanager", "D-L1-07-告警通知", "D-L1-04,L2-05", "—", "—"],
    ["D-L1-08", "L1", "健康检查/进程指标", "P0", "Phase 0", "部分实现", "浅 /health 无 /ready", "D-L1-08-健康监控", "—", "R1-4", "—"],
    ["D-L1-09", "L1", "日志脱敏合规", "P2", "Phase 2", "未开始", "无 redact 策略", "D-L1-09-日志合规", "D-L1-01", "—", "—"],
    ["D-L1-10", "L1", "OpenTelemetry 追踪", "P2", "Phase 2", "未开始", "无 span 可视化", "D-L1-10-OpenTelemetry", "D-L1-01,06", "R1-5", "—"],
    ["D-L1-11", "L1", "动态 DEBUG 采样", "P1", "Phase 1", "未开始", "无法按玩家提级", "D-L1-11-动态调试采样", "D-L1-01,02", "—", "—"],
    ["D-L1-12", "L1", "Socket 全事件 Tap", "P2", "Phase 1", "未开始", "未 instrument 路径盲区", "D-L1-12-Socket事件Tap", "D-L1-01,11", "—", "—"],
    ["D-L2-01", "L2", "事件字典 Schema", "P0", "Phase 0", "部分实现", "xlsx 与代码无 CI 校验", "D-L2-01-事件字典", "—", "—", "33条埋点表"],
    ["D-L2-02", "L2", "metrics 保留归档", "P0", "Phase 0", "未开始", "SQLite 无限增长", "D-L2-02-保留归档", "D-L2-03", "R2-2", "—"],
    ["D-L2-03", "L2", "日聚合表", "P0", "Phase 0", "未开始", "无 daily_* 表", "D-L2-03-日聚合表", "D-L2-01", "—", "—"],
    ["D-L2-04", "L2", "Metrics 迁 PG/Timescale", "P2", "Phase 2", "未开始", "SQLite 瓶颈", "D-L2-04-存储演进", "D-L2-02,03", "R2-2,R2-3", "—"],
    ["D-L2-05", "L2", "RED 指标 Grafana", "P1", "Phase 1", "部分实现", "有 perf 日志无 Prometheus", "D-L2-05-RED指标", "D-L1-08", "—", "性能sheet"],
    ["D-L2-06", "L2", "Admin 导出/排查模板", "P1", "Phase 1", "部分实现", "无 CSV/correlation 过滤", "D-L2-06-Admin增强", "D-L1-06", "—", "timeline API"],
    ["D-L2-07", "L2", "Admin RBAC", "P1", "Phase 1", "部分实现", "单密钥全权限", "D-L2-07-Admin-RBAC", "R0-1", "R0-1,R2-4", "—"],
    ["D-L2-08", "L2", "v0.4.4 埋点对齐", "P1", "Phase 0~1", "部分实现", "3 缺失+别名", "D-L2-08-埋点缺口闭环", "—", "—", "33条全表"],
    ["D-L2-09", "L2", "业务健康看板", "P1", "Phase 1", "部分实现", "无趋势/分塘", "D-L2-09-业务健康看板", "D-L2-03", "—", "D类咬钩事件"],
    ["D-L2-10", "L2", "eventId 幂等", "P2", "Phase 2", "未开始", "重试双写风险", "D-L2-10-幂等去重", "—", "—", "—"],
    ["D-L2-11", "L2", "DB 定时备份", "P0", "Phase 0", "未开始", "无冷备", "D-L2-11-DB备份", "—", "—", "—"],
    ["D-L2-12", "L2", "客户端事件上报", "P2", "Phase 2", "部分实现", "仅 DEV 内存", "D-L2-12-客户端上报", "—", "—", "连接类A"],
    ["D-L2-13", "L2", "Live Session Inspector", "P1", "Phase 1", "未开始", "仅静态 fishing-debug", "D-L2-13-Live会话调试", "D-L1-06,L2-06", "—", "A2"],
    ["D-L2-14", "L2", "admin-web 运维台", "P2", "Phase 1~2", "未开始", "Admin 在移动端", "D-L2-14-admin-web运维台", "D-L2-06", "—", "—"],
    ["D-L3-01", "L3", "日批分析流水线", "P0", "Phase 0~1", "部分实现", "模拟与线上割裂", "D-L3-01-日批流水线", "D-L2-03", "—", "—"],
    ["D-L3-02", "L3", "线上 vs 模拟对照", "P1", "Phase 1", "部分实现", "无对照页", "D-L3-02-线上模拟对照", "D-L3-01", "—", "—"],
    ["D-L3-03", "L3", "留存会话分析", "P1", "Phase 1", "未开始", "无 D1/D7", "D-L3-03-留存会话", "D-L2-03", "—", "fishing_start"],
    ["D-L3-04", "L3", "经济 faucet/sink 日报", "P1", "Phase 1", "部分实现", "无日趋势", "D-L3-04-经济日报", "D-L3-01", "—", "bait_buy等"],
    ["D-L3-05", "L3", "生态健康日报", "P1", "Phase 1", "部分实现", "无线上人口报表", "D-L3-05-生态日报", "D-L3-01", "—", "—"],
    ["D-L3-06", "L3", "BI/数仓对接", "P2", "Phase 2", "未开始", "无 ETL", "D-L3-06-BI对接", "D-L2-04,L3-01", "R2-3", "—"],
    ["D-L3-07", "L3", "规则版本切分", "P1", "Phase 1", "部分实现", "metrics 无版本维", "D-L3-07-版本维度", "—", "—", "—"],
    ["D-L3-08", "L3", "运营指标告警", "P1", "Phase 1", "未开始", "无产量/枯竭告警", "D-L3-08-运营告警", "D-L3-01,05", "—", "—"],
    ["D-L3-09", "L3", "analytics 索引打通", "P1", "Phase 1", "部分实现", "index 无 live-daily", "D-L3-09-analytics打通", "D-L3-01", "—", "—"],
    ["D-L3-10", "L3", "用户数据合规", "P2", "Phase 2", "未开始", "无删除/导出", "D-L3-10-数据合规", "D-L1-09", "—", "—"],
]

GAP_MATRIX = [
    ["能力域", "L1 日志", "L2 指标", "L3 分析", "商业化就绪度"],
    ["结构化可检索", "△ console 分散", "✓ metrics 有 schema", "✓ 模拟报告完善", "不足"],
    ["持久化", "✗ 终端/内存", "△ SQLite 无 TTL", "△ 仅离线 simulate", "不足"],
    ["集中平台", "✗ 无", "△ Admin API", "△ HTML 本地", "不足"],
    ["排障闭环", "△ 部分 correlation", "✓ player timeline", "✗ 无线上日报", "部分"],
    ["实时介入", "✗ 无 OTel/Live", "△ fishing-debug 快照", "—", "不足（D-L1-10~12,L2-13）"],
    ["告警", "✗ 无", "△ summary alerts 少", "✗ 无", "不足"],
    ["合规备份", "✗", "✗ 无备份策略", "✗", "不足"],
    ["与架构 xlsx 关系", "R1-5,R1-4", "R2-2,R0-1,R2-4", "独立建设", "需协同"],
    ["与埋点 xlsx 关系", "性能日志 E 类", "A~E 33 条", "聚合口径", "D-L2-08 对齐"],
]

REF_INDEX = [
    ["文件名", "路径", "用途", "与本清单关系"],
    ["v0.4.4-埋点表清单.xlsx", "docs/planning/reports/", "33 条埋点事件标准表（6 sheet）", "D-L2-01/08 需保持同步；L2 事件字典权威来源"],
    ["服务器架构问题与修复方案-v0.5.xlsx", "docs/planning/reports/", "14 项架构问题（鉴权/在线态/SQLite 等）", "L1/L2 基础设施依赖 R0/R1/R2 项"],
    ["三层数据体系-开发需求清单.xlsx", "docs/planning/reports/", "本文件", "三层缺口总清单 + 每需求 sheet"],
    ["服务器架构缺陷与埋点设计-v0.4.4.md", "docs/planning/specs/", "埋点设计原文", "事件分类 A~E、字段规范"],
    ["v0.4.4-埋点缺口复核与补全.md", "docs/planning/specs/", "33 条复核结论", "D-L2-08 实施依据"],
    ["排查-挂机断线诊断阶段2-4.md", "docs/planning/specs/", "排障 SOP", "D-L2-06/12 用户故事"],
    ["三层数据体系-可观测性补充-v0.6.md", "docs/planning/specs/", "实时排障补充需求与 xlsx 对齐说明", "D-L1-10~12 · D-L2-13~14 权威说明"],
    ["docs/analytics/", "docs/analytics/", "离线模拟与校准报告", "D-L3-01~09 线上扩展基础"],
]

PHASE_PLAN = [
    ["阶段", "时间建议", "目标", "本清单需求（编号）", "里程碑"],
    ["Phase 0", "上线前 1~2 周", "能排障、不丢数据", "L1-01,03,05,08 · L2-01,02,03,08,11 · L3-01", "日志落盘+错误入库+日聚合+备份+埋点对齐"],
    ["Phase 1", "内测~小规模商用", "可观测、可对照、可实时介入", "L1-04,06,07,11,12 · L2-05~09,13,14 · L3-02~05,07~09", "Grafana+集中日志+Live Inspector+线上日报"],
    ["Phase 2", "稳定增长", "可扩展、合规、分布式追踪", "L1-09,10 · L2-04,10,12 · L3-06,10", "OTel+PG 迁移+BI+合规"],
]

# 与对话方案对齐说明（非 sheet，供维护者参考）
RECONCILIATION_NOTE = """
权威来源：本 xlsx（31→36 条需求，v0.6 增补 5 条）。
对话中「Phase A/B/C 可观测性建议」已并入：
  Phase A ≈ L1-01,03,05,08 + L2-06
  Phase B ≈ L1-04,07 + L2-05
  Phase C ≈ L1-10,11,12 + L2-13,14
未采纳重复建设：自建 structured_logs SQLite tail（有 Loki 后不需要）。
Sentry vs error_logs：默认 D-L1-05 自建表；可选用 Sentry 作为 L1-05 步骤 2 可选项。
"""


def build_workbook() -> Workbook:
    wb = Workbook()

    # 总览
    ws = wb.active
    ws.title = "总览"
    headers = ["编号", "层级", "标题", "优先级", "阶段", "当前状态", "缺口摘要", "对应sheet", "依赖", "关联架构项", "关联埋点"]
    ws.append(headers)
    for row in OVERVIEW_ROWS:
        ws.append(row)
    col_widths = {"A": 10, "B": 6, "C": 22, "D": 8, "E": 10, "F": 10, "G": 28, "H": 22, "I": 14, "J": 12, "K": 12}
    for col, w in col_widths.items():
        ws.column_dimensions[col].width = w
    for c in ws[1]:
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="1F4E78")
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        layer = row[1].value
        fill = LAYER_FILL.get(str(layer), None)
        if fill:
            for c in row:
                c.fill = fill
                c.alignment = Alignment(vertical="top", wrap_text=True)

    # 缺口矩阵
    ws2 = wb.create_sheet("缺口矩阵")
    for r in GAP_MATRIX:
        ws2.append(r)
    for col in "ABCDE":
        ws2.column_dimensions[col].width = 22
    for c in ws2[1]:
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="4472C4")

    # 关联索引
    ws3 = wb.create_sheet("关联文档索引")
    for r in REF_INDEX:
        ws3.append(r)
    ws3.column_dimensions["A"].width = 36
    ws3.column_dimensions["B"].width = 28
    ws3.column_dimensions["C"].width = 32
    ws3.column_dimensions["D"].width = 40
    for c in ws3[1]:
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="548235")

    # 阶段路线图
    ws4 = wb.create_sheet("阶段路线图")
    for r in PHASE_PLAN:
        ws4.append(r)
    for col, w in {"A": 10, "B": 16, "C": 20, "D": 48, "E": 28}.items():
        ws4.column_dimensions[col].width = w
    for c in ws4[1]:
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="BF8F00")

    # 方案对齐（对话建议 vs 本清单）
    ws5 = wb.create_sheet("方案对齐说明")
    ws5.append(["说明", RECONCILIATION_NOTE.strip()])
    ws5.column_dimensions["A"].width = 120
    ws5["A1"].font = Font(bold=True)
    ws5["A2"].alignment = Alignment(wrap_text=True, vertical="top")
    ws5.row_dimensions[2].height = 200

    # 各需求 sheet
    sheet_order = [r[7] for r in OVERVIEW_ROWS]
    for name in sheet_order:
        add_kv_sheet(wb, name, REQUIREMENTS[name])

    return wb


def main() -> None:
    REPORTS.mkdir(parents=True, exist_ok=True)
    wb = build_workbook()
    try:
        wb.save(OUT)
        print(f"Wrote {OUT} ({len(wb.sheetnames)} sheets)")
    except PermissionError:
        alt = REPORTS / "三层数据体系-开发需求清单-完整.xlsx"
        wb.save(alt)
        print(f"Wrote {alt} ({len(wb.sheetnames)} sheets) — 原文件可能被占用")


if __name__ == "__main__":
    main()
