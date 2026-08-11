#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Build 埋点表清单.xlsx — events catalog per metrics schema (root + docs copy)."""
import shutil
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

ROOT = Path(__file__).resolve().parents[1]
REPORTS = ROOT / "docs/planning/reports"
# 仓库根目录为权威交付物；docs/planning/reports/ 为同步副本（兼容旧链接）
OUT_NAME = "v0.4.4-埋点表清单.xlsx"
OUT = ROOT / OUT_NAME
OUT_DOCS = REPORTS / OUT_NAME

# 类别, 事件名, 类型, 触发时机, 必填字段, payload示例, 优先级, 状态, 落点文件
EVENTS = [
    ("A连接会话", "socket_connect", "日志", "socket 建连成功", "ts\nplayerId\nsocketId", '{"eventType":"socket_connect"}', "P0", "已实现", "fishingObservability.ts"),
    ("A连接会话", "socket_disconnect", "metrics+日志", "socket 断开", "ts\nplayerId\nreason", '{"eventType":"disconnect"}', "P0", "已实现(别名disconnect)", "fishingStateMachine.ts"),
    ("A连接会话", "socket_connect_error", "日志", "建连失败", "ts\nreason", "{}", "P0", "已实现", "socketLifecycle.ts"),
    ("A连接会话", "join_pond_attempt", "日志", "收到 join_pond", "pondId\nplayerId", "{}", "P0", "已实现", "gameState.ts"),
    ("A连接会话", "join_pond_success", "metrics+日志", "join/reconnect 成功", "pondId\njoinKind", '{"joinKind":"reconnect"}', "P0", "已实现", "gameState.ts"),
    ("A连接会话", "join_pond_fail", "metrics+日志", "join 失败", "ackError", "{}", "P0", "已实现", "gameState.ts"),
    ("A连接会话", "leave_pond", "metrics+日志", "主动离塘", "reason", '{"reason":"user_leave"}', "P0", "已实现", "gameState.ts"),
    ("A连接会话", "disconnect_timeout", "metrics+日志", "60s 未重连清场", "playerId\npondId", "{}", "P0", "已实现", "fishingStateMachine.ts"),
    ("A连接会话", "session_rebound", "日志", "playerId 绑定新 socket", "playerId\nsocketId", "{}", "P1", "已实现", "sessionRegistry.ts"),
    ("B钓位鱼塘", "spot_take_success", "metrics", "占座成功", "spotId", "{}", "P1", "已实现", "gameState.ts"),
    ("B钓位鱼塘", "spot_take_fail", "metrics", "占座失败", "reason", "{}", "P1", "已实现", "gameState.ts"),
    ("B钓位鱼塘", "spot_release", "metrics", "释放钓位", "spotId\nreason", "{}", "P1", "已实现", "gameState.ts"),
    ("B钓位鱼塘", "pond_full_reject", "metrics", "鱼塘满", "pondId", "{}", "P1", "已实现", "gameState.ts"),
    ("B钓位鱼塘", "bot_evicted_for_human", "metrics", "踢 bot 腾位", "botUserId", "{}", "P1", "已实现", "gameState.ts"),
    ("C状态机", "fishing_phase_transition", "metrics+日志", "phase 变更；D-L2-16 metrics 仅短码 f/t/c；ADMIN-OBS-1.3 默认 bot 不落库（METRICS_BOT_PHASE=1 恢复）；日志仍用全称", "f\nt\nc", '{"f":4,"t":5,"c":"bite_hook"}', "P0", "已实现", "fishingObservability.ts"),
    ("C状态机", "phase_transition_invalid", "日志+metrics", "非法跳转；metrics 短码 f/t/c；日志可读全称", "f\nt\nc", '{"f":0,"t":5,"c":"bite_hook"}', "P1", "已实现", "fishingStateMachine.ts"),
    ("D咬钩产出", "bite_tick_miss", "metrics", "waiting 未咬钩（D-L2-15 默认不落库；METRICS_BITE_TICK_PERSIST=1）", "pondId", "{}", "P1", "已废弃默认", "fishingStateMachine.ts"),
    ("D咬钩产出", "bite_tick_hit", "metrics", "waiting 咬钩（D-L2-15 默认不落库，并入 bite_hook）", "speciesId\nquality", "{}", "P1", "已废弃默认", "fishingStateMachine.ts"),
    ("D咬钩产出", "bite_hook", "metrics", "上钩+会话累计（sessionHooks/Escapes/MissTicks）", "sessionHooks\nsessionMissTicks", "{}", "P0", "已实现", "fishingStateMachine.ts"),
    ("D咬钩产出", "pending_catch_created", "metrics", "待领取创建", "fishId", "{}", "P1", "已实现", "inventory.ts"),
    ("D咬钩产出", "pending_catch_accept", "metrics", "领取成功", "fishId", "{}", "P1", "已实现(别名catch_accept)", "socketPondHandlers.ts"),
    ("D咬钩产出", "pending_catch_expired", "metrics", "待领取超时", "fishId", "{}", "P1", "已实现", "inventory.ts"),
    ("D咬钩产出", "bait_depleted", "metrics", "饵耗尽", "baitId", "{}", "P1", "已实现", "fishingStateMachine.ts"),
    ("D咬钩产出", "escape", "metrics", "脱钩+会话累计", "sessionEscapes\nsessionHooks", "{}", "P0", "已实现", "fishingStateMachine.ts"),
    ("D咬钩产出", "fishing_start", "metrics", "开始钓鱼", "pondId", "{}", "P1", "已实现", "fishingStateMachine.ts"),
    ("D咬钩产出", "fishing_stop", "metrics", "停止钓鱼", "pondId", "{}", "P1", "已实现", "fishingStateMachine.ts"),
    ("D咬钩产出", "abandon_fishing", "metrics", "<30s 停止", "pondId", "{}", "P1", "已实现", "fishingMetrics.ts"),
    ("E性能", "tick_fishing_phases_duration_ms", "聚合日志", "200ms tick 耗时", "durationMs", "{}", "P2", "已实现", "serverLoops.ts"),
    ("E性能", "bite_check_loop_duration_ms", "聚合日志", "咬钩循环耗时", "durationMs", "{}", "P2", "已实现", "serverLoops.ts"),
    ("E性能", "snapshot_build_duration_ms", "聚合日志", "buildSnapshot 耗时", "durationMs\npondId", "{}", "P2", "已实现", "gameState.ts"),
    ("E性能", "socket_broadcast_fanout", "指标+可选日志", "广播 fanout；默认 Prometheus counter，FANOUT_LOG_INFO=1 才打 info", "fanoutCount\nchannel\npondId", "{}", "P2", "已实现", "serverLoops.ts"),
    ("E性能", "sqlite_query_slow", "日志", "慢 SQL", "label\ndurationMs\nrows", "{}", "P2", "已实现", "db.ts"),
    ("E性能", "admin_route_duration_ms", "日志", "Admin 接口耗时", "route\ndurationMs", "{}", "P2", "已实现", "admin.ts"),
]

CAT_SHEETS = {
    "A连接会话": "A-连接与会话",
    "B钓位鱼塘": "B-钓位与鱼塘",
    "C状态机": "C-状态机",
    "D咬钩产出": "D-咬钩与产出",
    "E性能": "E-系统性能",
}


def style_header(ws, ncol: int) -> None:
    for c in ws[1]:
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="1F4E78")
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def main() -> None:
    REPORTS.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "总览"
    headers = ["类别", "事件名", "类型", "触发时机", "必填字段(核心)", "示例payload", "优先级", "状态", "推荐落点文件"]
    ws.append(headers)
    for e in EVENTS:
        ws.append(list(e))
    style_header(ws, len(headers))
    widths = {"A": 12, "B": 28, "C": 12, "D": 28, "E": 18, "F": 22, "G": 8, "H": 18, "I": 28}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for c in row:
            c.alignment = Alignment(vertical="top", wrap_text=True)

    for cat, sheet in CAT_SHEETS.items():
        wsc = wb.create_sheet(sheet)
        wsc.append(headers)
        for e in EVENTS:
            if e[0] == cat:
                wsc.append(list(e))
        style_header(wsc, len(headers))
        for col, w in widths.items():
            wsc.column_dimensions[col].width = w
        for row in wsc.iter_rows(min_row=2, max_row=wsc.max_row):
            for c in row:
                c.alignment = Alignment(vertical="top", wrap_text=True)

    # 慢测验证 sheet
    ws_v = wb.create_sheet("慢测验证")
    ws_v.append(["用例", "脚本", "状态", "说明"])
    ws_v.append(["server observability", "npm run verify:server-observability", "已实现", "join/phase/timeline"])
    ws_v.append(["pending timeout", "npx tsx scripts/verify-pending-timeout.ts", "可选", "慢测 pending_catch_expired"])
    ws_v.append(["afk diag", "npx tsx scripts/verify-afk-diag.ts", "已实现", "断线宽限"])
    style_header(ws_v, 4)

    wb.save(OUT)
    print(f"Wrote {OUT} ({len(wb.sheetnames)} sheets)")
    try:
        REPORTS.mkdir(parents=True, exist_ok=True)
        shutil.copy2(OUT, OUT_DOCS)
        print(f"Synced copy: {OUT_DOCS}")
    except OSError as e:
        print(f"WARN: could not sync docs copy: {e}")


if __name__ == "__main__":
    main()
